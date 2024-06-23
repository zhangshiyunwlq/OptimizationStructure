import copy
import random
import xlrd
import matplotlib.pyplot as plt
import openpyxl
# import evaluation_DNN as ed
import pandas as pd

def all_modular_infor():
    node1 = [(0, 0), (6, 0), (6, 6), (0, 6)]
    story_node = []
    for i in range(modular_length_num):
        for j in range(4):
            story_node.append((node1[j][0] + 7 * i, node1[j][1]))

    nodes_all = []
    for i in range(story_num):
        for j in range(len(story_node)):
            nodes_all.append((story_node[j][0], story_node[j][1] + 7 * i))

    room1_indx = [0, 1, 3, 2]
    all_room_indx = []
    for i in range(modular_length_num * story_num):
        temp = []
        for j in range(len(room1_indx)):
            temp.append(room1_indx[j] + 4 * i)
        all_room_indx.append(temp)

    frame = [[2, 3], [0, 1], [0, 2], [1, 3]]

    joint_hor = [[1, 4], [2, 7]]
    story_joint_hor = []
    for i in range(modular_length_num - 1):
        for j in range(len(joint_hor)):
            story_joint_hor.append([joint_hor[j][0] + 4 * i, joint_hor[j][1] + 4 * i])
    all_joint_hor = []
    for i in range(story_num):
        for j in range(len(story_joint_hor)):
            all_joint_hor.append([story_joint_hor[j][0] + 4 * modular_length_num * i,
                                  story_joint_hor[j][1] + 4 * modular_length_num * i])

    joint_ver = [[3, 32], [2, 33]]
    story_joint_ver = []
    for i in range(modular_length_num):
        for j in range(len(joint_ver)):
            story_joint_ver.append([joint_ver[j][0] + 4 * i, joint_ver[j][1] + 4 * i])
    all_joint_ver = []
    for i in range(story_num - 1):
        for j in range(len(story_joint_ver)):
            all_joint_ver.append([story_joint_ver[j][0] + 4 * modular_length_num * i,
                                  story_joint_ver[j][1] + 4 * modular_length_num * i])

    # 绘制支撑的点
    brace_node1 = [(0, 0), (6, 0), (6, 6), (0, 6), (2, 0), (4, 0), (2, 6), (4, 6), (3, 6)]
    brace_story_node = []
    for i in range(modular_length_num):
        for j in range(9):
            brace_story_node.append((brace_node1[j][0] + 7 * i, brace_node1[j][1]))

    brace_nodes_all = []
    for i in range(story_num):
        for j in range(len(brace_story_node)):
            brace_nodes_all.append((brace_story_node[j][0], brace_story_node[j][1] + 7 * i))

    brace_room1_indx = [0, 1, 3, 2, 4, 5, 6, 7, 8]
    brace_all_room_indx = []
    for i in range(modular_length_num * story_num):
        temp = []
        for j in range(len(brace_room1_indx)):
            temp.append(brace_room1_indx[j] + 9 * i)
        brace_all_room_indx.append(temp)

    brace_frame = [[[0, 8], [1, 8]], [[1, 2], [0, 3]],
                   [[0, 6], [2, 4], [1, 7], [3, 5]]]
    return nodes_all, all_room_indx, all_joint_hor, all_joint_ver, brace_nodes_all, brace_all_room_indx, frame, brace_frame

def get_info(iter_num):
    pop_room = []
    pop_brace = []
    wb = openpyxl.load_workbook(
        filename=f'D:\desktop\os\optimization of structure\optimization of structure\optimization of structure\out_all_infor_case4\\run_infor_{num_var}_{modular_num}_{al_time}.xlsx',
    )
    sheet1 = wb['pop1_all']
    for z in range(3*zone_num):
        rows = sheet1.cell(iter_num*(POP_SIZE+1)+2, z + 1).value
        pop_room.append(rows)
    sheet3 = wb['pop2_all']
    brace_dis = sheet3.cell(iter_num*(POP_SIZE+1)+2, num_var+1).value
    sheet2 = wb['pop3_all']
    for z in range(modular_all):
        rows = sheet2.cell(iter_num * (POP_SIZE + 1) + 2, z + 1).value
        pop_brace.append(rows)


    return pop_room,pop_brace,brace_dis

def decoding_modular_section(pop2):

    pop_all = copy.deepcopy(pop2)
    modular_type1 = [i for i in range(3)]
    #生成对每个模块的截面编号索引
    modular_type_all= []
    for i in range(modular_num):
        modular_type_temp = []
        for j in range(len(modular_type1)):
            modular_type_temp.append(num_var+num_room_type+modular_type1[j]+3*i)
        modular_type_all.append(modular_type_temp)


    #提取截面表
    pop1_all = []
    for i in range(len(pop_all)):
        pop1_section = []
        for j in range(num_var+num_room_type+section_num+brace_num,num_var+num_room_type+section_num+brace_num+zone_num):
            for z in range(3):
                sec = int(pop_all[i][j])
                pop1_section.append(pop_all[i][int(modular_type_all[sec][z])])
        pop1_all.append(pop1_section)

    #解码pop1_1all
    pop1_decoding = []
    for i in range(len(pop1_all)):
        pop1_temp = []
        for j in range(len(pop1_all[i])):
            pop1_temp.append(pop_all[i][int(pop1_all[i][j])])
        pop1_decoding.append(pop1_temp)


    #生成支撑表
    brace_sort = [i for i in range(num_var+num_room_type+section_num,num_var+num_room_type+section_num+brace_num)]
    pop3_all = []
    for i in range(len(pop_all)):
        pop3_brace = []
        for j in range(num_var+num_room_type+section_num+brace_num,num_var+num_room_type+section_num+brace_num+zone_num):
            bra = int(pop_all[i][j])
            if pop_all[i][int(brace_sort[bra])] ==0:
                pop3_brace.append(0)
            else:
                pop3_brace.append(pop_all[i][num_var])
        pop3_all.append(pop3_brace)
    pop_3 =[]
    for j in range(len(pop_all)):
        temp2 = pop3_all[j]
        brace_all = []
        for i in range(len(labels)):
            temp1 = int(labels[i])
            brace_all.append(temp2[temp1])
        pop_3.append(brace_all)

    return pop1_decoding,pop_3


def draw_corr_conn(cons_all,nodes_all,corrs_all):
    for line in cons_all:
        x_values = [nodes_all[line[0]][0], nodes_all[line[1]][0]]
        y_values = [nodes_all[line[0]][1], nodes_all[line[1]][1]]
        plt.plot(x_values, y_values, 'b-', linewidth=2.5, color='grey')

    for line in corrs_all:
        x_values = [nodes_all[line[0]][0], nodes_all[line[1]][0]]
        y_values = [nodes_all[line[0]][1], nodes_all[line[1]][1]]
        plt.plot(x_values, y_values, 'b-', linewidth=2.5, color='grey')
    for node in nodes_all:
        plt.plot(node[0], node[1], 'ro', markersize=2.5, color='grey')


def draw_frame(all_indx_draw,pop_room,pop_brace):
    for i in range(len(all_indx_draw)):

        modular_type1 = [ii for ii in range(3)]
        modular_type_all = []
        for ii in range(zone_num):
            modular_type_temp = []
            for j in range(len(modular_type1)):
                modular_type_temp.append(modular_type1[j] + 3 * ii)
            modular_type_all.append(modular_type_temp)

        column_frame = [[all_room_indx[i][frame[2][0]], all_room_indx[i][frame[2][1]]],
                        [all_room_indx[i][frame[3][0]], all_room_indx[i][frame[3][1]]]]
        bottom_frame = [[all_room_indx[i][frame[1][0]], all_room_indx[i][frame[1][1]]]]
        top_frame = [[all_room_indx[i][frame[0][0]], all_room_indx[i][frame[0][1]]]]

        for ii in range(len(top_frame)):
            member_indx = pop_room[modular_type_all[labels[int(all_indx_draw[i])]][0]]
            if member_indx <= 6:
                c = 'red'
                x_values = [nodes_all[top_frame[ii][0]][0],
                            nodes_all[top_frame[ii][1]][0]]
                y_values = [nodes_all[top_frame[ii][0]][1],
                            nodes_all[top_frame[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=int(list_new[int(member_indx)] * 0.6 + 1))
            elif member_indx >= 6:
                c = 'b'
                x_values = [nodes_all[top_frame[ii][0]][0],
                            nodes_all[top_frame[ii][1]][0]]
                y_values = [nodes_all[top_frame[ii][0]][1],
                            nodes_all[top_frame[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=int(list_new[int(member_indx)] * 0.6 + 1))

        for ii in range(len(bottom_frame)):
            member_indx = pop_room[modular_type_all[labels[int(all_indx_draw[i])]][1]]
            if member_indx <= 6:
                c = 'red'
                x_values = [nodes_all[bottom_frame[ii][0]][0],
                            nodes_all[bottom_frame[ii][1]][0]]
                y_values = [nodes_all[bottom_frame[ii][0]][1],
                            nodes_all[bottom_frame[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=int(list_new[int(member_indx)] * 0.6 + 1))
            elif member_indx >= 6:
                c = 'b'
                x_values = [nodes_all[bottom_frame[ii][0]][0],
                            nodes_all[bottom_frame[ii][1]][0]]
                y_values = [nodes_all[bottom_frame[ii][0]][1],
                            nodes_all[bottom_frame[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=int(list_new[int(member_indx)] * 0.6 + 1))

        for ii in range(len(column_frame)):
            member_indx = pop_room[modular_type_all[labels[int(all_indx_draw[i])]][2]]
            if member_indx <= 6:
                c = 'red'
                x_values = [nodes_all[column_frame[ii][0]][0],
                            nodes_all[column_frame[ii][1]][0]]
                y_values = [nodes_all[column_frame[ii][0]][1],
                            nodes_all[column_frame[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=int(list_new[int(member_indx)] * 0.6 + 1))
            elif member_indx >= 6:
                c = 'b'
                x_values = [nodes_all[column_frame[ii][0]][0],
                            nodes_all[column_frame[ii][1]][0]]
                y_values = [nodes_all[column_frame[ii][0]][1],
                            nodes_all[column_frame[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=int(list_new[int(member_indx)] * 0.6 + 1))

    for i in range(len(all_indx_draw)):
        brace_frame_all = []
        if int(pop_brace[int(all_indx_draw[i])]) != 0:
            brace_type = brace_frame[int(pop_brace[int(all_indx_draw[i])]) - 1]
            for ii in range(len(brace_type)):
                temp = [brace_all_room_indx[i][brace_type[ii][0]], brace_all_room_indx[i][brace_type[ii][1]]]
                brace_frame_all.append(temp)
            for ii in range(len(brace_frame_all)):
                c = 'grey'
                x_values = [brace_nodes_all[brace_frame_all[ii][0]][0],
                            brace_nodes_all[brace_frame_all[ii][1]][0]]
                y_values = [brace_nodes_all[brace_frame_all[ii][0]][1],
                            brace_nodes_all[brace_frame_all[ii][1]][1]]
                plt.plot(x_values, y_values, 'b-', color=c, linewidth=2.5)


def all_indx(all_it):
    for ite in all_it:
        # 读取优化信息
        pop_room, pop_brace, brace_dis = get_info(ite)
        # 生成一个侧面的房间编号
        temp_indx = [i for i in range(8)]
        all_indx_draw = []
        for i in range(story_num):
            te = []
            for j in range(len(temp_indx)):
                te.append(temp_indx[j] + i * modular_length_num * 2)
            all_indx_draw.extend(te)

        draw_frame(all_indx_draw, pop_room, pop_brace)
        draw_corr_conn(all_joint_hor, nodes_all, all_joint_ver)

        # 设置坐标轴范围
        plt.xlim(-3, 80)
        plt.ylim(-5, 90)

        # 设置坐标轴标签
        plt.xlabel('X')
        plt.ylabel('Y')

        # 显示图形
        plt.gca().set_aspect('equal', adjustable='box')
        plt.show()

def draw_pred(pop1,pop2,pop3):
    for ite in range(len(pop1)):
        # 读取优化信息
        pop_room=pop1[ite]
        brace_dis=pop2[ite][num_var]
        pop_brace =pop3[ite]
        # 生成一个侧面的房间编号
        temp_indx = [i for i in range(8)]
        all_indx_draw = []
        for i in range(story_num):
            te = []
            for j in range(len(temp_indx)):
                te.append(temp_indx[j] + i * modular_length_num * 2)
            all_indx_draw.extend(te)

        draw_frame(all_indx_draw, pop_room, pop_brace)
        draw_corr_conn(all_joint_hor, nodes_all, all_joint_ver)

        # 设置坐标轴范围
        plt.xlim(-3, 80)
        plt.ylim(-5, 90)

        # 设置坐标轴标签
        plt.xlabel('X')
        plt.ylabel('Y')

        # 显示图形
        plt.gca().set_aspect('equal', adjustable='box')
        plt.show()


'''model data'''
# modular size
#建筑参数
modular_length = 8000
modular_width = [4000,4000,5400,3600,3600,4400,4400,4000]
modular_heigth = 3000
modular_length_num = 8
modular_dis = 400
corridor_width = 4000
story_num = 12
story_zone = 4#每组模块的分区数量
story_group = 3#每组模块的楼层数
modular_num = 3#整个建筑的模块种类
num_room_type =1


zone_num = int(story_num / story_group * story_zone)
section_num = 3 * modular_num
brace_num = modular_num
group_num = int(story_num / story_group)
modular_all = modular_length_num * 2 *story_num
POP_SIZE = 30
#生成建筑信息
nodes_all,all_room_indx,all_joint_hor,all_joint_ver,brace_nodes_all,brace_all_room_indx,frame,brace_frame = all_modular_infor()

labels = []
labels1 = []
for i in range(group_num):
    temp = []
    for j in range(story_zone):
        for z in range(int(modular_length_num/story_zone)):
            temp.append(i*story_zone+j)
    for j in range(2*story_group):
        labels.extend(temp)
        labels1.append(temp)

#构件界面排序
area = [2080,3642,4392,5292,6660,2256,3584,4400,5600,6800,7600,8400]
fit_ini = copy.deepcopy(area)
luyi = copy.deepcopy(area)
luyi.sort()
sort_num = []
lst = list(range(1, len(fit_ini) + 1))
list_new = []
for i in range(len(fit_ini)):
    sort_num.append(fit_ini.index(luyi[i]))
for i in range(len(fit_ini)):
    list_new.append(lst[sort_num[i]])

num_var = 5
al_time = 15
#绘制某个pop中的最有个体
# all_in=[19]
# all_indx(all_in)
#数据文件中生成pop2
# fit1,fit2,pop2 = ed.get_DNN_GA(15,27,20)
# pop1,pop3=ed.decoding_modular_section(pop2)
# draw_pred(pop1,pop2,pop3)

#数据文件中获得pop2
path_memo = f"D:\desktop\os\optimization of structure\optimization of structure\optimization of structure\DNN_test_data\\all_data_5.xlsx"
pop2 = pd.read_excel(io=path_memo, sheet_name="pop2_all", header=None)
pop2 = pop2.values.tolist()
pop1,pop3=decoding_modular_section(pop2)
draw_pred(pop1,pop2,pop3)
