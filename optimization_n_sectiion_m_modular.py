import numpy as np
import random
from random import randint
import copy
import data_to_json as dj
import xlwt
import threading
import queue
import math as m
import os
import pandas as pd
import sys
import xlrd
import matplotlib.pyplot as plt
import data_to_json as dj
import modular_utils as md
import model_to_sap as ms
import sap_run as sr
import configparser
import comtypes.client
import gc
import xlsxwriter
import csv
import openpyxl
from CNN import create_model
# os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
'''
功能：指定n种截面，m种模块然后整栋建筑的的模块都随机使用指定的模块，优化算法为HIGA,案例为case4

'''

def out_put_result(pop1_all,pop2_all,pop3_all,fitness_all,weight_all,pop_all_fitness,pop_all_weight,time):
    APIPath = os.path.join(os.getcwd(), 'out_all_infor_case4')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    path1 = os.path.join(APIPath, f'run_infor_{num_var}_{modular_num}_{time}')


    wb1 = xlsxwriter.Workbook(f'{path1}.xlsx')
    out_pop1_all = wb1.add_worksheet('pop1_all')

    loc = 0
    for i in range(len(pop1_all)):
        out_pop1_all.write(loc, 0, f'{[i]}')
        for j in range(len(pop1_all[i])):
            loc += 1
            for z in range(len(pop1_all[i][j])):
                out_pop1_all.write(loc, z, pop1_all[i][j][z])
        loc += 1

    out_pop2_all = wb1.add_worksheet('pop2_all')
    loc = 0
    for i in range(len(pop2_all)):
        out_pop2_all.write(loc, 0, f'{[i]}')
        for j in range(len(pop2_all[i])):
            loc += 1
            for z in range(len(pop2_all[i][j])):
                out_pop2_all.write(loc, z, pop2_all[i][j][z])
        loc += 1

    out_pop3_all = wb1.add_worksheet('pop3_all')
    loc = 0
    for i in range(len(pop3_all)):
        out_pop3_all.write(loc, 0, f'{[i]}')
        for j in range(len(pop3_all[i])):
            loc += 1
            for z in range(len(pop3_all[i][j])):
                out_pop3_all.write(loc, z, pop3_all[i][j][z])
        loc += 1

    pop_all_fit = wb1.add_worksheet('pop_all_fitness')
    loc = 0
    for i in range(len(pop_all_fitness)):
        pop_all_fit.write(loc, 0, f'{[i]}')
        loc += 1
        for j in range(len(pop_all_fitness[i])):
            pop_all_fit.write(loc, j, pop_all_fitness[i][j])
        loc += 1


    pop_all_wei = wb1.add_worksheet('pop_all_weight')
    loc = 0
    for i in range(len(pop_all_weight)):
        pop_all_wei.write(loc, 0, f'{[i]}')
        loc += 1
        for j in range(len(pop_all_weight[0])):
            pop_all_wei.write(loc, j, pop_all_weight[i][j])
        loc += 1


    outmaxfitness = wb1.add_worksheet('max_fitness')
    loc = 0
    outmaxfitness.write(loc, 0, 'max_fitness_all')
    loc += 1
    for i in range(len(fitness_all)):
        outmaxfitness.write(loc, i, fitness_all[i])
    loc += 1
    outmaxfitness.write(loc, 0, 'min_weight_all')
    loc += 1
    for i in range(len(weight_all)):
        outmaxfitness.write(loc, i, weight_all[i])




    wb1.close()

def out_put_fitness_prediction(DNN_fit_pred):
    APIPath = os.path.join(os.getcwd(), 'out_all_DNN_fitness_case4')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    path1 = os.path.join(APIPath, f'prediction_fitness_{num_var}_{modular_num}_{time}')
    wb1 = xlsxwriter.Workbook(f'{path1}.xlsx')
    fit_pred = wb1.add_worksheet(f'DNN_fitness')
    loc = 0
    for ii in range(len(DNN_fit_pred)):

        for i in range(len(DNN_fit_pred[ii])):
            fit_pred.write(loc,i, DNN_fit_pred[ii][i])
        loc += 1

    wb1.close()

def out_put_prediction_gx(gx_all,time_pr):
    APIPath = os.path.join(os.getcwd(), 'out_all_prediction_case4')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    path1 = os.path.join(APIPath, f'prediction_infor_{num_var}_{modular_num}_{time}')

    wb1 = xlsxwriter.Workbook(f'{path1}.xlsx')

    for ii in range(len(gx_all)):
        gx_pred = wb1.add_worksheet(f'gx_prediction_{ii}')
        loc = 0
        for i in range(len(gx_all[ii])):
            for j in range(len(gx_all[ii][i])):
                gx_pred.write(loc, j, gx_all[ii][i][j])
            loc += 1

    wb1.close()

def out_put_memorize(memorize_pool,memorize_fit,memorize_weight,memorize_gx,memorize_loss,memorize_mae,memorize_gx_nor,memorize_num,gx_prediction):
    APIPath = os.path.join(os.getcwd(), 'out_all_memorize_case4')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    path1 = os.path.join(APIPath, f'memorize_infor_{num_var}_{modular_num}_{time}')

    wb1 = xlsxwriter.Workbook(f'{path1}.xlsx')



    out_pop1_all = wb1.add_worksheet('memorize_pool')
    loc = 0

    for i in range(len(memorize_pool)):
        pool_list = copy.deepcopy(memorize_pool[i])
        # pool_list.tolist()
        for j in range(len(pool_list)):
            out_pop1_all.write(loc, j, pool_list[j])
        loc += 1

    pop_all_fit = wb1.add_worksheet('memorize_fit')

    for i in range(len(memorize_fit)):
        pop_all_fit.write(i, 0, memorize_fit[i])

    pop_all_wei = wb1.add_worksheet('memorize_weight')

    for i in range(len(memorize_weight)):
        pop_all_wei.write(i, 0, memorize_weight[i])


    memo_gx = wb1.add_worksheet('memorize_gx')
    loc = 0
    for i in range(len(memorize_gx)):
        for j in range(len(memorize_gx[i])):
            memo_gx.write(loc, j, memorize_gx[i][j])
        loc += 1

    gx_pred = wb1.add_worksheet('gx_prediction')
    loc = 0
    for i in range(len(gx_prediction)):
        for j in range(len(gx_prediction[i])):
            gx_pred.write(loc, j, gx_prediction[i][j])
        loc += 1



    # for i in range(len(memorize_loss)):
    #     memo_loss.write(i, 0, memorize_loss[i])
    memo_loss = wb1.add_worksheet('memorize_loss')
    loc = 0
    for i in range(len(memorize_loss)):
        for j in range(len(memorize_loss[i])):
            memo_loss.write(loc, j, memorize_loss[i][j])
        loc += 1

    memo_mae = wb1.add_worksheet('memorize_mae')
    loc = 0
    for i in range(len(memorize_mae)):
        for j in range(len(memorize_mae[i])):
            memo_mae.write(loc, j, memorize_mae[i][j])
        loc += 1

    # for i in range(len(memorize_mae)):
    #     memo_mae.write(i, 0, memorize_mae[i])

    memo_gx_nor = wb1.add_worksheet('memorize_gx_nor')
    loc = 0
    for i in range(len(memorize_gx_nor)):
        for j in range(len(memorize_gx_nor[i])):
            memo_gx_nor.write(loc, j, memorize_gx_nor[i][j])
        loc += 1

    memo_num = wb1.add_worksheet('memorize_num')

    for i in range(len(memorize_num)):
        memo_num.write(i, 0, memorize_num[i])




    wb1.close()

def draw_loss(num_var,time):
    fig = plt.figure(1)
    ax1 = fig.add_subplot()
    dev_x = np.arange(0, len(history_loss))
    dev_y = history_loss
    ax1.set_xlabel("Epoch",fontsize=10)
    ax1.set_ylabel("loss", fontsize=10)

    APIPath = os.path.join(os.getcwd(), f'loss_mae')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass
    path1 = os.path.join(APIPath, f'loss{num_var}_{time}')
    ax1.plot(dev_x, dev_y)
    plt.savefig(path1, dpi=300)
    plt.close()
    #绘制mae图
    fig2 = plt.figure(2)
    ax2 = fig2.add_subplot()
    ax2.set_xlabel("Epoch",fontsize=10)
    ax2.set_ylabel("Mae", fontsize=10)
    path1 = os.path.join(APIPath, f'mae{num_var}_{time}')
    dev_x = np.arange(0, len(history_loss))
    dev_y = history_mae
    ax2.plot(dev_x, dev_y)
    plt.savefig(path1, dpi=300)
    plt.close()

def mulit_get_sap(num_thread):
    case_name = []
    APIPath_name = []
    mySapObject_name = []
    SapModel_name = []
    ModelPath_name = []
    for i in range(num_thread):
        case_name.append(f"cases{i}")
        APIPath_name.append(os.path.join(os.getcwd(), f"cases{i}"))
        mySapObject_name.append(f"mySapObject{i}")
        SapModel_name.append(f"SapModel{i}")
        ModelPath_name.append(f"ModelPath{i}")
        mySapObject_name[i], ModelPath_name[i], SapModel_name[i] = SAPanalysis_GA_run2(APIPath_name[i])
    return mySapObject_name,ModelPath_name,SapModel_name

def generate_coding_modular_section(x):

    room_nu = np.linspace(1, 3, 3)

    pop = np.zeros((POP_SIZE,num_var+num_room_type+section_num+brace_num+zone_num))
    for i in range(len(pop)):
        sec = list(map(int, random.sample(x.tolist(), num_var)))
        sec.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]

        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = random.randint(1,3)

        for j in range(num_var+num_room_type,num_var+num_room_type+section_num):
            pop[i][j] = random.randint(0,num_var-1)

        for j in range(num_var+num_room_type+section_num,num_var+num_room_type+section_num+brace_num):
            pop[i][j] = randint(0,1)
        for j in range(num_var+num_room_type+section_num+brace_num,num_var+num_room_type+section_num+brace_num+zone_num):
            pop[i][j] = randint(0,modular_num-1)

    return pop

def decoding_modular_section(pop2):

    pop_all = copy.deepcopy(pop2)
    modular_type1 = [i for i in range(3)]
    #生成对每个模块的截面编号索引
    modular_type_all= []
    for i in range(modular_num):
        modular_type_temp = []
        for j in range(len(modular_type1)):
            modular_type_temp.append(num_var+num_room_type+modular_type1[j]+3*i)
        modular_type_all.append(modular_type_temp)


    #提取截面表
    pop1_all = []
    for i in range(len(pop_all)):
        pop1_section = []
        for j in range(num_var+num_room_type+section_num+brace_num,num_var+num_room_type+section_num+brace_num+zone_num):
            for z in range(3):
                sec = int(pop_all[i][j])
                pop1_section.append(pop_all[i][int(modular_type_all[sec][z])])
        pop1_all.append(pop1_section)

    #解码pop1_1all
    pop1_decoding = []
    for i in range(len(pop1_all)):
        pop1_temp = []
        for j in range(len(pop1_all[i])):
            pop1_temp.append(pop_all[i][int(pop1_all[i][j])])
        pop1_decoding.append(pop1_temp)


    #生成支撑表
    brace_sort = [i for i in range(num_var+num_room_type+section_num,num_var+num_room_type+section_num+brace_num)]
    pop3_all = []
    for i in range(len(pop_all)):
        pop3_brace = []
        for j in range(num_var+num_room_type+section_num+brace_num,num_var+num_room_type+section_num+brace_num+zone_num):
            bra = int(pop_all[i][j])
            if pop_all[i][int(brace_sort[bra])] ==0:
                pop3_brace.append(0)
            else:
                pop3_brace.append(pop_all[i][num_var])
        pop3_all.append(pop3_brace)
    pop_3 =[]
    for j in range(len(pop_all)):
        temp2 = pop3_all[j]
        brace_all = []
        for i in range(len(labels)):
            temp1 = int(labels[i])
            brace_all.append(temp2[temp1])
        pop_3.append(brace_all)

    return pop1_decoding,pop_3

def mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop_room,pop_room_label):
    # 建立房间信息
    sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                      cfg_file_name="Steel_section_data_I_cube.ini")
    modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
    # 按房间分好节点
    modulars_of_building = modular_building.building_modulars

    modular_infos = {}
    # 每个房间定义梁柱截面信息
    sr.run_column_room_modular_section(labels,pop_room_label, modular_length_num * 2 * story_num, sections_data_c1, modular_infos, pop_room,story_num,zone_num)
    #
    for i in range(len(modulars_of_building)):
        modulars_of_building[i].Add_Info_And_Update_Modular(
            # modular_infos[modulars_of_building[i].modular_label - 1])
            modular_infos[i])
    modular_building.Building_Assembling(modulars_of_building)

    all_data = ms.Run_GA_sap_2(mySapObject, ModelPath, SapModel, modular_building,pop_room_label,200,modular_length_num,story_num)
    aa, bb, cc, dd, ee, ff, gg, hh, ii = all_data

    ret = SapModel.SetModelIsLocked(False)
    return aa,bb,cc,dd,ee,ff,gg,hh,ii


def mulitrun_GA_1(ModelPath,mySapObject, SapModel,pop1,pop_all,pop3,q,result,weight_1,col_up,beam_up,sap_run_time00):
    while True:
        if q.empty():
            break
        time = q.get()
        pop = pop1[time]
        pop_room_label = pop3[time]
        pop2= pop_all[time]
        value = 0
        for i in range(len(memorize_pool)):
            sum_code = sum(pop2)
            if sum_code==memorize_sum[i]:
                pop2_list = copy.deepcopy(pop2)
                memorize_list = copy.deepcopy(memorize_pool[i])
                if pop2_list.tolist()==memorize_list.tolist():
                    result[time]=memorize_fit[i]
                    weight_1[time]=memorize_weight[i]
                    col_up[time]=memorize_col[i]
                    beam_up[time]=memorize_beam[i]
                    value = 1
                    break

        if value == 0:
            we, co, be, r1, r2, r3, r4, dis_all, force_all = mulit_Sap_analy_allroom(ModelPath, mySapObject, SapModel,
                                                                                             pop,
                                                                                             pop_room_label)
            num_zero = pop_room_label.count(0)
            nonzero_rate = (len(pop_room_label)-num_zero)/len(pop_room_label)
            res1, res2,gx,gx_demo = Fun_1(we, co, be, dis_all, force_all, 10000,nonzero_rate)

            # num3 += 1
            weight_1[time] = res2
            col_up[time] = co
            beam_up[time] = be
            result[time] = res1
            #全局记忆池
            memorize_sum.append(sum(pop2))
            memorize_pool.append(pop2)
            memorize_fit.append(res1)
            memorize_weight.append(res2)
            memorize_col.append(col_up[time])
            memorize_beam.append(beam_up[time])
            memorize_gx.append(gx)
            memorize_gx_nor.append(gx_demo)
            # 局部记忆池
            memorize_sum_local.append(sum(pop2))
            memorize_pool_local.append(pop2)
            memorize_fit_local.append(res1)
            memorize_weight_local.append(res2)
            memorize_col_local.append(col_up[time])
            memorize_beam_local.append(beam_up[time])
            memorize_gx_local.append(gx)

def mulitrun_GA_continue(ModelPath,mySapObject, SapModel,pop1,pop_all,pop3,q,result,weight_1,col_up,beam_up,sap_run_time00):
    while True:
        if q.empty():
            break
        time = q.get()
        pop = pop1[time]
        pop_room_label = pop3[time]
        pop2= pop_all[time]


        we, co, be, r1, r2, r3, r4, dis_all, force_all = mulit_Sap_analy_allroom(ModelPath, mySapObject, SapModel,
                                                                                         pop,
                                                                                         pop_room_label)
        num_zero = pop_room_label.count(0)
        nonzero_rate = (len(pop_room_label)-num_zero)/len(pop_room_label)
        res1, res2,gx,gx_demo = Fun_1(we, co, be, dis_all, force_all, 10000,nonzero_rate)

        # num3 += 1
        weight_1[time] = res2
        col_up[time] = co
        beam_up[time] = be
        result[time] = res1
        #全局记忆池
        memorize_sum.append(sum(pop2))
        memorize_pool.append(pop2)
        memorize_fit.append(res1)
        memorize_weight.append(res2)
        memorize_col.append(col_up[time])
        memorize_beam.append(beam_up[time])
        memorize_gx.append(gx)
        memorize_gx_nor.append(gx_demo)
        # 局部记忆池
        memorize_sum_local.append(sum(pop2))
        memorize_pool_local.append(pop2)
        memorize_fit_local.append(res1)
        memorize_weight_local.append(res2)
        memorize_col_local.append(col_up[time])
        memorize_beam_local.append(beam_up[time])
        memorize_gx_local.append(gx)

def GA_examine(ModelPath,mySapObject, SapModel,pop1,pop3):
    result = []
    num1 = 0
    num2 = 0
    num3 = 0
    weight_1 = []
    col_up = []
    beam_up = []
    # if len(pop_all) ==0:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         pop_all.append(pop)
    #         we,co,be,r1,r2,r3,r4 =Sap_analy(pop)
    #         res1,res2 = Fun(we, co, be, 10000)
    #         num3 += 1
    #         weight_1.append(res2)
    #         result.append(res1)
    #         pop_fun_all.append(res1)
    #         pop_weight_all.append(res2)
    # else:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         for i in range(len(pop_all)):
    #             if all(pop == pop_all[i]):
    #                 num1 = 1
    #                 num2 = i
    #                 # result.append(pop_fun_all[i])
    #         if num1 == 0:
    #             pop_all.append(pop)
    #             we, co, be, r1, r2, r3, r4 = Sap_analy(pop)
    #             res1, res2 = Fun(we, co, be, 10000)
    #             num3 += 1
    #             weight_1.append(res2)
    #             pop_weight_all.append(res2)
    #             result.append(res1)
    #             pop_fun_all.append(res1)
    #         elif num1 == 1:
    #             result.append(pop_fun_all[num2])
    #             weight_1.append(pop_weight_all[num2])
    # wb2_examine_ind = openpyxl.Workbook()
    # wb2_examine_ind = openpyxl.load_workbook('examine_individual.xlsx')
    # wb2_pop1_indivi = wb2_examine_ind.create_sheet('pop1_indivi', index=0)
    # wb2_pop3_indivi = wb2_examine_ind.create_sheet('pop3_indivi', index=1)
    loc_3 = 1
    for time in range(len(pop1)):
        pop = pop1[time]
        pop_room_label = pop3[time]
        # _ = wb2_pop1_indivi.cell(row=loc_3, column=1, value=f'{pop1[time]}')
        # _ = wb2_pop3_indivi.cell(row=loc_3, column=1, value=f'{pop3[time]}')
        # loc_3 += 1
        # wb2_examine_ind.save('examine_individual.xlsx')
        # pop_all.append(pop)
        we,co,be,r1,r2,r3,r4,dis_all,force_all =mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop,pop_room_label)
        num_zero = pop_room_label.count(0)
        nonzero_rate = (len(pop_room_label) - num_zero) / len(pop_room_label)
        res1,res2,gx,gx_demo = Fun_1(we, co, be,dis_all,force_all, 10000,nonzero_rate)
        # num3 += 1
        weight_1.append(res2)
        col_up.append(co)
        beam_up.append(be)
        result.append(res1)
        # pop_fun_all.append(res1)
        # pop_weight_all.append(res2)

    # wb_clear_in1 = openpyxl.load_workbook('examine_individual.xlsx')
    # ws_clear_in1 = wb_clear_in1['pop1_indivi']
    # for row in ws_clear_in1:
    #     for cell in row:
    #         cell.value = None
    # ws_clear_in3 = wb_clear_in1['pop3_indivi']
    # for row in ws_clear_in3:
    #     for cell in row:
    #         cell.value = None
    # wb2_examine_ind.save('examine_individual.xlsx')
    return result,weight_1,col_up,beam_up

def SAPanalysis_GA_run2(APIPath):


    cfg = configparser.ConfigParser()
    cfg.read("Configuration.ini", encoding='utf-8')
    ProgramPath = cfg['SAP2000PATH']['dirpath']
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    AttachToInstance = False
    SpecifyPath = True

    # ModelPath = os.path.join(APIPath, 'API_1-001.sdb')
    helper = comtypes.client.CreateObject('SAP2000v1.Helper')
    helper = helper.QueryInterface(comtypes.gen.SAP2000v1.cHelper)
    if AttachToInstance:
        # attach to a running instance of SAP2000
        try:
            # get the active SapObject
            mySapObject = helper.Getobject("CSI.SAP2000.API.SapObject")
        except (OSError, comtypes.COMError):
            print("No running instance of the program found or failed to attach.")
            sys.exit(-1)
    else:
        if SpecifyPath:
            try:
                # 'create an instance of the SAPObject from the specified path
                mySapObject = helper.CreateObject(ProgramPath)
            except (OSError, comtypes.COMError):
                print("Cannot start a new instance of the program from" + ProgramPath)
                sys.exit(-1)
        else:
            try:
                # create an instance of the SapObject from the latest installed SAP2000
                mySapObject = helper.CreateObjectProgID("CSI.SAP2000.API.SapObject")
            except (OSError, comtypes.COMError):
                print("Cannot start a new instance of the program")
                sys.exit(-1)

        # start SAP2000 application
        mySapObject.ApplicationStart()

    # create SapModel object
    SapModel = mySapObject.SapModel
    # initialize model
    SapModel.InitializeNewModel()
    ModelPath = os.path.join(APIPath, 'API_1-001.sdb')
    # create new blank model
    return mySapObject,ModelPath, SapModel

def thread_sap(ModelPath_name,mySapObject_name,SapModel_name,num,pop1,pop2,pop3,result,weight_1,col_up,beam_up):


    pop_n = [0 for i in range(len(pop2[0]))]

    q = queue.Queue()
    threads = []
    for i in range(len(pop1)):
        q.put(i)
    for i in range(num_thread):
        if len(ModelPath_name)!=num_thread:
            for j in range(len(ModelPath_name),num_thread):
                mySapObject_name.append(f"mySapObject{j}")
                SapModel_name.append(f"SapModel{j}")
                ModelPath_name.append(f"ModelPath{j}")
                mySapObject_name[j], ModelPath_name[j], SapModel_name[j] = SAPanalysis_GA_run2(os.path.join(os.getcwd(), f"cases{j}"))
        t = threading.Thread(target=mulitrun_GA_1, args=(ModelPath_name[i],mySapObject_name[i],SapModel_name[i],pop1,pop2,pop3,q,result,weight_1,col_up,beam_up,sap_run_time00))
        t.start()
        threads.append(t)
    for i in threads:
        i.join()
    return result,weight_1,col_up,beam_up

def thread_sap_continue(ModelPath_name,mySapObject_name,SapModel_name,num,pop1,pop2,pop3,result,weight_1,col_up,beam_up):


    pop_n = [0 for i in range(len(pop2[0]))]

    q = queue.Queue()
    threads = []
    for i in range(len(pop1)):
        q.put(i)
    for i in range(num_thread):
        if len(ModelPath_name)!=num_thread:
            for j in range(len(ModelPath_name),num_thread):
                mySapObject_name.append(f"mySapObject{j}")
                SapModel_name.append(f"SapModel{j}")
                ModelPath_name.append(f"ModelPath{j}")
                mySapObject_name[j], ModelPath_name[j], SapModel_name[j] = SAPanalysis_GA_run2(os.path.join(os.getcwd(), f"cases{j}"))
        t = threading.Thread(target=mulitrun_GA_continue, args=(ModelPath_name[i],mySapObject_name[i],SapModel_name[i],pop1,pop2,pop3,q,result,weight_1,col_up,beam_up,sap_run_time00))
        t.start()
        threads.append(t)
    for i in threads:
        i.join()
    return result,weight_1,col_up,beam_up


def select_2(pop, fitness):  # nature selection wrt pop's fitness

    fit_ini = copy.deepcopy(fitness)
    luyi = copy.deepcopy(fitness)
    luyi.sort(reverse=True)
    sort_num = []
    for i in range(len(fit_ini)):
        sort_num.append(luyi.index(fit_ini[i]))
    # print(sort_num)
    # print(f'{len(sort_num)}_{len(pop)}')
    for i in range(len(sort_num)):
        if sort_num[i]==0:
            sort_num[i]+=0.01
    # pop_last.append(pop)



    # for i in range(len(list_new)):
    #     list_new[i] = m.e ** (list_new[i] * 1.5)
    idx = np.random.choice(np.arange(len(pop)), size=len(pop), replace=True,
                           p=np.array(sort_num) / (sum(sort_num)))
    pop2 = np.zeros((len(pop), len(pop[0])))
    for i in range(len(pop2)):
        pop2[i] = pop[int(idx[i])]
    return pop2

def crossover_and_mutation_GA_for_DNN(pop2,num_var,CROSSOVER_RATE,MUTATION_RATE):
    pop = pop2

    new_pop = np.zeros((len(pop),len(pop[0])))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(len(pop2))]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1_stort_modular_section(num_room_type,num_var,child,MUTATION_RATE)
        new_pop[i] = child

    for i in range(len(new_pop)):

        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]

        for mutate_point in range(num_var):
            x_var = list(map(int, x.tolist()))
            for mutate_point_1 in range(num_var):
                if pop[i][mutate_point_1] in x_var:
                    x_var.remove(pop[i][mutate_point_1])
            if pop[i][mutate_point] == pop[i][
                mutate_point + 1] and mutate_point <= num_var - 2:  # 以MUTATION_RATE的概率进行变异
                pop[i][mutate_point] = np.random.choice(x_var)

        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]

    return new_pop

def mutation_1_stort_modular_section(num_room_type,num_var,child,MUTATION_RATE):

    num_var = int(num_var)
    room_nu = np.linspace(1, 12, 12)
    num_room_type = int(num_room_type)
    if num_var!=len(x):
        for mutate_point in range(num_var):
            x_var = list(map(int, x.tolist()))
            for mutate_point_1 in range(num_var):
                if child[mutate_point_1] in x_var:
                    x_var.remove(child[mutate_point_1])
            if np.random.rand() < MUTATION_RATE:  # 以MUTATION_RATE的概率进行变异
                child[mutate_point] = np.random.choice(x_var)

    for j in range(num_var,num_var+num_room_type):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(1,3)
    for j in range(num_var+num_room_type,num_var+num_room_type+section_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,num_var-1)
    for j in range(num_var+num_room_type+section_num,num_var+num_room_type+section_num+brace_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,1)
    for j in range(num_var + num_room_type + section_num+ brace_num, num_var + num_room_type + section_num + brace_num+zone_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0, modular_num-1)
            # child[j] = 0


def generation_population_modular_section(best_indivi,rate):

    best_in = copy.deepcopy(best_indivi)
    best_in.tolist()
    pop = np.zeros((50,len(best_in)))
    for i in range(len(pop)):
        if num_var != len(x):
            for mutate_point in range(num_var):
                x_var = list(map(int, x.tolist()))
                for mutate_point_1 in range(num_var):
                    if pop[i][mutate_point_1] in x_var:
                        x_var.remove(pop[i][mutate_point_1])
                if np.random.rand() <0.25:  # 以MUTATION_RATE的概率进行变异
                    pop[i][mutate_point] = np.random.choice(x_var)
                else:
                    pop[i][mutate_point] = best_in[mutate_point]
        for j in range(num_var, num_var + num_room_type):
            if np.random.rand() < 0.25:
                pop[i][j] = randint(1, 3)
            else:
                pop[i][j] = best_in[j]
        for j in range(num_var + num_room_type, num_var + num_room_type + section_num):
            if np.random.rand() < 0.25:
                pop[i][j] = randint(0, num_var - 1)
            else:
                pop[i][j] = best_in[j]
        for j in range(num_var + num_room_type + section_num, num_var + num_room_type + section_num + brace_num):
            if np.random.rand() < 0.25:
                pop[i][j] = randint(0, 1)
            else:
                pop[i][j] = best_in[j]
        for j in range(num_var + num_room_type + section_num + brace_num,
                       num_var + num_room_type + section_num + brace_num + zone_num):
            if np.random.rand() < 0.25:
                pop[i][j] = randint(0, modular_num - 1)
            else:
                pop[i][j] = best_in[j]

    new_pop = copy.deepcopy(pop)
    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]




    return new_pop

def gx_nonNormalization(gx,gx_data_select):
    gx_demo = copy.deepcopy(gx)
    for i in range(len(gx_demo)):
        for j in range(len(gx_data_select)):
            if gx_data_select[j] == 0:
                gx_demo[i][j] = gx_demo[i][j] * 3 - 1
            elif gx_data_select[j] == 1:
                gx_demo[i][j] = gx_demo[i][j] * 4 - 1
            elif gx_data_select[j] == 2:
                gx_demo[i][j] = gx_demo[i][j] * 0.01
            elif gx_data_select[j] == 3:
                gx_demo[i][j] = gx_demo[i][j] * 0.01
            elif gx_data_select[j] == 5:
                gx_demo[i][j] = gx_demo[i][j] * 600
        # gx_demo[i][1] = gx_demo[i][1] * 1.5-1
        # gx_demo[i][2] = gx_demo[i][2] * 0.1
        # gx_demo[i][3] = gx_demo[i][3] * 0.1
        # gx_demo[i][5] = gx_demo[i][5] * 200+400
        # gx_demo[i][0] = gx_demo[i][0] * 600
    return gx_demo
#修改区间
def Gx_convert(fitness1,gx_data_select):
    fitness3 = copy.deepcopy(fitness1)
    fitness4 = []  # 储存所有gx
    fitness2 = []  # 所有gx的和
    for j in range(len(fitness3)):
        fitness4.append(fitness3[j].tolist())
    fitness4=gx_nonNormalization(fitness4,gx_data_select)
    fitness5 = copy.deepcopy(fitness4)
    for j in range(len(fitness3)):
        # fitness2.append(sum(fitness4[j]))
        for z in range(len(gx_data_select)):
            if gx_data_select[z] == 0:
                if fitness5[j][z]<=0:
                    fitness5[j][z] =0
            elif gx_data_select[z] == 1:
                if fitness5[j][z] <= 0:
                    fitness5[j][z] = 0
            elif gx_data_select[z] == 2:
                if fitness5[j][z]<=0.00167 and fitness5[j][z] >= -0.00167:
                    fitness5[j][z] =0
                else:
                    fitness5[j][z] = 100*abs(fitness5[j][z])
            elif gx_data_select[z] == 3:
                if fitness5[j][z] <= 0.004 and fitness5[j][z] >= -0.004:
                    fitness5[j][z] = 0
                else:
                    fitness5[j][z] = 100*abs(fitness5[j][z])
            elif gx_data_select[z] == 4:
                if fitness5[j][z] <= 0.4:
                    fitness5[j][z] = 0
                else:
                    fitness5[j][z] = 100 * abs(fitness5[j][z])
        value = 0
        for z in range(len(gx_data_select)):
            if gx_data_select[z] != 5:
                value += 10000*fitness5[j][z]
            elif gx_data_select[z] == 5:
                value += fitness5[j][z]
        fitness2.append(value)
        # if fitness5[j][0]<=0:
        #     fitness5[j][0] =0
        # if fitness5[j][1] <= 0:
        #     fitness5[j][1] = 0
        # if fitness5[j][2]<=0.00167 and fitness5[j][2] >= -0.00167:
        #     fitness5[j][2] =0
        # else:
        #     fitness5[j][2] = abs(fitness5[j][2])
        # if fitness5[j][3] <= 0.004 and fitness5[j][3] >= -0.004:
        #     fitness5[j][3] = 0
        # else:
        #     fitness5[j][3] = abs(fitness5[j][3])
        # if fitness5[j][4] <= 0.4:
        #     fitness5[j][4] = 0
        # fitness2.append(fitness5[j][5]+10000*(fitness5[j][0]+fitness5[j][1]+fitness5[j][2]*100+fitness5[j][3]*100+100*abs(fitness5[j][4])))
        # if fitness5[j][0]<=0:
        #     fitness5[j][0] = abs(fitness5[j][0])*100
        # fitness2.append(fitness5[j][0])
    return fitness2
#修改区间
def gx_Normalization(gx,gx_data_select):
    gx_demo = copy.deepcopy(gx)
    for i in range(len(gx_demo)):
        for j in range(len(gx_data_select)):
            if gx_data_select[j] == 0:
                if gx_demo[i][j]>=2:
                    gx_demo[i][j]=1
                elif gx_demo[i][j]<=-1:
                    gx_demo[i][j] = 0
                elif gx_demo[i][j]<=2 and gx_demo[i][j]>=-1:
                    gx_demo[i][j]=(gx_demo[i][j]+1)/3
            elif gx_data_select[j] == 1:
                if gx_demo[i][j]>=3:
                    gx_demo[i][j]=1
                elif gx_demo[i][j]<=-1:
                    gx_demo[i][j] = 0
                elif gx_demo[i][j]<=3 and gx_demo[i][j]>=-1:
                    gx_demo[i][j]=(gx_demo[i][j]+1)/4
            elif gx_data_select[j] == 2:
                if gx_demo[i][j] >= 0.01:
                    gx_demo[i][j] = 1
                else:
                    gx_demo[i][j] = gx_demo[i][j] / 0.01
            elif gx_data_select[j] == 3:
                if gx_demo[i][j] >= 0.01:
                    gx_demo[i][j] = 1
                else:
                    gx_demo[i][j] = gx_demo[i][j] / 0.01
            elif gx_data_select[j] == 5:
                if gx_demo[i][j] >= 600:
                    gx_demo[i][j] = 1
                elif gx_demo[i][j] <= 0:
                    gx_demo[i][j] = 0
                elif gx_demo[i][j] <= 600 and gx_demo[i][j] >= 0:
                    gx_demo[i][j] = gx_demo[i][j] / 600
    return gx_demo
# def crossover_and_mutation_GA_for_DNN(pop2,num_var,CROSSOVER_RATE,MUTATION_RATE):
#     pop = pop2
#
#     new_pop = np.zeros((len(pop),len(pop[0])))
#     for i in range(len(pop)):
#         father = pop[i]
#         child = father
#         if np.random.rand() < CROSSOVER_RATE:
#             mother = pop[np.random.randint(POP_SIZE)]
#             cross_points1 = np.random.randint(low=0, high=len(pop[0]))
#             cross_points2 = np.random.randint(low=0, high=len(pop[0]))
#             while cross_points2==cross_points1:
#                 cross_points2 = np.random.randint(low=0, high=len(pop[0]))
#             exchan = []
#             exchan.append(cross_points2)
#             exchan.append(cross_points1)
#             for j in range(min(exchan),max(exchan)):
#                 child[j] = mother[j]
#         mutation_1_stort_modular_section(num_room_type,num_var,child,MUTATION_RATE)
#         new_pop[i] = child
#
#
#     for i in range(len(new_pop)):
#         sec_sort = []
#         room_sort = []
#         for j in range(num_var):
#             sec_sort.append(new_pop[i][j])
#         sec_sort.sort()
#         for j in range(num_var):
#             new_pop[i][j] = sec_sort[j]
#
#
#
#     return new_pop

#用于GA_forDNN中的交叉变异
def mutation_GA_for_DNN_modular(child,num_var,MUTATION_RATE):
    section_num = 3*modular_num
    brace_num = modular_num

    for j in range(section_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = random.randint(0,13)
    for j in range(section_num,section_num+brace_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(1,3)
    for j in range(section_num+brace_num,section_num+brace_num+modular_length_num*2*story_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,modular_num-1)


def GA_for_DNN(run_time,pop2,model):
    fitness_pred = []
    for i in range(run_time):
        temp = []
        fitness1 = model.predict(pop2,verbose=0)
        fitness2 = Gx_convert(fitness1,gx_data_select)#归一化还原，并将每个染色体对应的gx累加
        # all_fit_pred_GA.append(fitness2)
        mm = fitness2.index(min(fitness2))
        min1 = min(fitness2)
        temp.append(fitness1[mm])
        fitness_pred.append(min1)
        mm2_all = pop2[mm]
        #选择
        pop2 = select_2(pop2, fitness2)
        # 交叉变异
        pop2 = crossover_and_mutation_GA_for_DNN(pop2, num_var,CROSSOVER_RATE,MUTATION_RATE)
        fit_pred = model.predict(pop2,verbose=0)
        fit_pred2=Gx_convert(fit_pred,gx_data_select)
        if min1 <= fit_pred2[0]:
            pop2[0] = mm2_all
    # gx_pred_best.append(temp)
    DNN_prediction_fitness.append(fitness_pred)
    # fitness_best.append(min(min1,fit_pred2[0]))
    return pop2

def DNN_GA(num_var,num_room_type,num_ind,best_indivi,run_time):
    #局部训练
    if len(memorize_pool_local)!=0:
        pool_local = copy.deepcopy(memorize_pool_local)
        x_train1_local = np.array(pool_local)
        x_train_local = x_train1_local#提取用于训练的x_train部分
        gx_local = copy.deepcopy(memorize_gx_local)
        y_train_local = np.array(gx_local)
        y_train_local= gx_Normalization(y_train_local,gx_data_select)#归一化
        model= create_model(len(x_train_local[0]), len(y_train_local[0]))#创建模型
        #verbose取消打印损失
        model.fit(x_train_local, y_train_local, epochs=200, batch_size=32,verbose=0)#训练模型

    #全局训练
    pool_global = copy.deepcopy(memorize_pool)
    gx_global = copy.deepcopy(memorize_gx)
    x_train1 = np.array(pool_global)
    x_train = x_train1#提取用于训练的x_train部分
    y_train = np.array(gx_global)
    y_train = gx_Normalization(y_train,gx_data_select)#归一化
    model = create_model(len(x_train[0]),len(y_train[0]))#创建模型
    history=model.fit(x_train, y_train, epochs=200, batch_size=32,verbose=0)#训练模型
    # history_loss.extend(history.history['loss'])
    # history_mae.extend(history.history['mae'])
    # history_loss.append(history.history['loss'][len(history.history['loss'])-1])
    # history_mae.append(history.history['mae'][len(history.history['loss'])-1])
    history_loss.append(history.history['loss'])
    history_mae.append(history.history['mae'])
    pop_best = []
    for i in range(num_ind):
        pop1 = generation_population_modular_section(best_indivi, 0.25)#根据最好个体生成种群
        pop2 = copy.deepcopy(pop1)
        pop2 = GA_for_DNN(run_time, pop2, model)
        pop_best.append(pop2[0].tolist())
    pop_best = np.array(pop_best)
    return pop_best,model

def GA_DNN_run_modular(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time):
    pop2= generate_coding_modular_section(x)
    pop_decoe_1 = copy.deepcopy(pop2)
    pop1,pop3 = decoding_modular_section(pop2)

    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []
    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    col_up_all= []
    beam_up_all=[]
    pop_all_weight=[]
    pop_all_fitness=[]
    weight_min=[]
    min_ru = []
    sap_run_time = 0
    predict_time = 0
    all_pred = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)

        # 计算fitness等参数
        fit = [0 for i in range(len(pop2))]
        weight = [0 for i in range(len(pop2))]
        clo_val = [0 for i in range(len(pop2))]
        beam_val = [0 for i in range(len(pop2))]
        result1,weight_pop,clo_up_1,beam_up_1=thread_sap(ModelPath_name,mySapObject_name,SapModel_name,num_thread, pop1, pop2, pop3, fit, weight, clo_val, beam_val)

        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        fitness2 =copy.deepcopy(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        min1 = min(fitness2)
        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop3编码
        min_ru.append(min(fitness2))# 统计历代最小值
        #选择
        pop2 = select_2(pop2, fitness2)
        #交叉变异
        pop2 = crossover_and_mutation_GA_for_DNN(pop2,num_var,CROSSOVER_RATE,MUTATION_RATE)

        # 引入新个体
        run_time +=1
        if run_time % 20 == 0:
            pop2_new,model = DNN_GA(num_var,num_room_type,int(1 * len(pop2)),pop2[0],400)
            exchange_num = int(1*len(pop2))
            # for ex_num in range(exchange_num):
            #     pop2[len(pop2) - 1 - ex_num] = pop2_new[ex_num]
            pop2 = copy.deepcopy(pop2_new)
            memorize_num.append(len(memorize_pool))
            memorize_sum_loacl = []
            memorize_pool_loacl = []
            memorize_fit_loacl = []
            memorize_weight_loacl = []
            memorize_col_loacl = []
            memorize_beam_loacl = []
            memorize_gx_loacl = []

            # 使用深度神经网络对记忆池中的所有个体进行预测
            memorize_pool_temp = copy.deepcopy(memorize_pool)
            memorize_pool_temp = np.array(memorize_pool_temp)
            x_data_prediction = memorize_pool_temp


            fitness_prediction = model.predict(x_data_prediction, verbose=0)
            all_pred.append(fitness_prediction)


        if run_time % 20 == 0:
            print(run_time)
            print(f'记忆池数量:{len(memorize_pool)}')
        pop1, pop3 = decoding_modular_section(pop2)

        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if min1 <=GA_examine(ModelPath_name[0],mySapObject_name[0], SapModel_name[0],aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    out_put_prediction_gx(all_pred, predict_time)

    for i in range(len(mySapObject_name)):
        ret = mySapObject_name[i].ApplicationExit(False)
        SapModel_name[i] = None
        mySapObject_name[i] = None


    out_put_result(pop_zhongqun_all, pop_zhongqun_all_2, pop_zhongqun_all_3,min_ru,weight_min,pop_all_fitness,pop_all_weight,time)
    out_put_fitness_prediction(DNN_prediction_fitness)


    return pop_zhongqun_all,pop_zhongqun_all_2,pop_zhongqun_all_3,fitness_prediction
#GA算法
def GA_run_modular(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time):
    pop2= generate_coding_modular_section(x)
    pop_decoe_1 = copy.deepcopy(pop2)
    pop1,pop3 = decoding_modular_section(pop2)

    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []
    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    col_up_all= []
    beam_up_all=[]
    pop_all_weight=[]
    pop_all_fitness=[]
    weight_min=[]
    min_ru = []
    sap_run_time = 0
    predict_time = 0
    all_pred = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)

        # 计算fitness等参数
        fit = [0 for i in range(len(pop2))]
        weight = [0 for i in range(len(pop2))]
        clo_val = [0 for i in range(len(pop2))]
        beam_val = [0 for i in range(len(pop2))]
        result1,weight_pop,clo_up_1,beam_up_1=thread_sap(ModelPath_name,mySapObject_name,SapModel_name,num_thread, pop1, pop2, pop3, fit, weight, clo_val, beam_val)

        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        fitness2 =result1
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        min1 = min(fitness2)
        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop3编码
        min_ru.append(min(fitness2))# 统计历代最小值
        #选择
        pop2 = select_2(pop2, fitness2)
        #交叉变异
        pop2 = crossover_and_mutation_GA_for_DNN(pop2,num_var,CROSSOVER_RATE,MUTATION_RATE)



        if run_time % 20 == 0:
            print(run_time)
            print(f'记忆池数量:{len(memorize_pool)}')
        pop1, pop3 = decoding_modular_section(pop2)

        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if min1 <=GA_examine(ModelPath_name[0],mySapObject_name[0], SapModel_name[0],aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    # out_put_prediction_gx(all_pred, predict_time)
    for i in range(len(mySapObject_name)):
        ret = mySapObject_name[i].ApplicationExit(False)
        SapModel_name[i] = None
        mySapObject_name[i] = None


    out_put_result(pop_zhongqun_all, pop_zhongqun_all_2, pop_zhongqun_all_3,min_ru,weight_min,pop_all_fitness,pop_all_weight,time)



    return pop_zhongqun_all,pop_zhongqun_all_2,pop_zhongqun_all_3

def Fun_1(weight,g_col,g_beam,dis_all,all_force,u,rate):

    g_col_max= max(g_col)
    g_beam_max = max(g_beam)

    dis_all5_abs = copy.deepcopy(dis_all[5])
    for i in range(len(dis_all5_abs)):
        dis_all5_abs[i] = abs(dis_all5_abs[i])

    dis_all7_abs = copy.deepcopy(dis_all[7])
    for i in range(len(dis_all7_abs)):
        dis_all7_abs[i] = abs(dis_all7_abs[i])

    dis_all_max = max(dis_all5_abs)
    interdis_max = max(dis_all7_abs)

    rate_nonzero = copy.deepcopy(rate)
    if rate_nonzero<=0.4:
        rate_nonzero =0
    else:
        rate_nonzero=rate_nonzero

    if g_col_max<=0:
        g_col_fit = 0
    else:
        g_col_fit = g_col_max

    if g_beam_max<=0:
        g_beam_fit =0
    else:
        g_beam_fit = g_beam_max

    if dis_all_max<= 0.00167 and dis_all_max >= -0.00167:
        dis_all_fit = 0
    else:
        dis_all_fit = dis_all_max

    if interdis_max<= 0.004 and interdis_max >= -0.004:
        interdis_all_fit = 0
    else:
        interdis_all_fit = interdis_max

    G_value=u * (g_col_fit + g_beam_fit + 100*dis_all_fit + 100*interdis_all_fit +rate_nonzero)
    value_jisuan = [g_col_fit,g_beam_fit,100*dis_all_fit,100*interdis_all_fit,rate_nonzero,weight]
    # gx = [g_col_max,g_beam_max,dis_all_max,interdis_max,rate,weight]
    gx = []
    for z in range(len(gx_data_select)):
        if gx_data_select[z] ==0:
            gx.append(g_col_max)
        elif gx_data_select[z] ==1:
            gx.append(g_beam_max)
        elif gx_data_select[z] ==2:
            gx.append(dis_all_max)
        elif gx_data_select[z] ==3:
            gx.append(interdis_max)
        elif gx_data_select[z] ==4:
            gx.append(rate)
        elif gx_data_select[z] ==5:
            gx.append(weight)
    # gx_Normalization = [g_col_all,g_beam_all,Y_dis_radio_all,Y_interdis_all]
    # result = weight + G_value

    value = 0
    for z in range(len(gx_data_select)):
        if gx_data_select[z] != 5:
            value += 10000 * value_jisuan[int(gx_data_select[z])]
        elif gx_data_select[z] == 5:
            value += weight

    result = value

    gx_demo = copy.deepcopy(gx)
    for j in range(len(gx_demo)):
        if gx_data_select[j] == 0:
            if gx_demo[j] >= 2:
                gx_demo[j] = 1
            elif gx_demo[j] <= -1:
                gx_demo[j] = 0
            elif gx_demo[j] <= 2 and gx_demo[j] >= -1:
                gx_demo[j] = (gx_demo[j] + 1) / 3
        elif gx_data_select[j] == 1:
            if gx_demo[j] >= 3:
                gx_demo[j] = 1
            elif gx_demo[j] <= -1:
                gx_demo[j] = 0
            elif gx_demo[j] <= 3 and gx_demo[j] >= -1:
                gx_demo[j] = (gx_demo[j] + 1) / 4
        elif gx_data_select[j] == 2:
            if gx_demo[j] >= 0.01:
                gx_demo[j] = 1
            else:
                gx_demo[j] = gx_demo[j] / 0.01
        elif gx_data_select[j] == 3:
            if gx_demo[j] >= 0.01:
                gx_demo[j] = 1
            else:
                gx_demo[j] = gx_demo[j] / 0.01
        elif gx_data_select[j] == 5:
            if gx_demo[j] >= 600:
                gx_demo[j] = 1
            elif gx_demo[j] <= 0:
                gx_demo[j] = 0
            elif gx_demo[j] <= 600 and gx_demo[j] >= 0:
                gx_demo[j] = (gx_demo[j]) / 600
    # if gx_demo[0]>=2:
    #     gx_demo[0]=1
    # elif gx_demo[0]<=-1:
    #     gx_demo[0] = 0
    # elif gx_demo[0]<=2 and gx_demo[0]>=-1:
    #     gx_demo[0]=(gx_demo[0]+1)/3
    # if gx_demo[1]>=0.5:
    #     gx_demo[1]=1
    # elif gx_demo[1]<=-1:
    #     gx_demo[1] = 0
    # elif gx_demo[1]<=0.5 and gx_demo[1]>=-1:
    #     gx_demo[1]=(gx_demo[1]+1)/1.5
    # if gx_demo[2] >= 0.1:
    #     gx_demo[2] = 1
    # else:
    #     gx_demo[2] = gx_demo[2] / 0.1
    # if gx_demo[3] >= 0.1:
    #     gx_demo[3] = 1
    # else:
    #     gx_demo[3] = gx_demo[3] / 0.1
    # if gx_demo[0] >= 600:
    #     gx_demo[0] = 1
    # elif gx_demo[0] <= 0:
    #     gx_demo[0] = 0
    # elif gx_demo[0] <= 600 and gx_demo[0] >= 0:
    #     gx_demo[0] = (gx_demo[0])/600
    return result,weight,gx,gx_demo
def get_continue_data(file_time,num_continue):
    path_memo = f"D:\desktop\os\optimization of structure\optimization of structure\optimization of structure\out_all_memorize_case4\memorize_infor_{num_var}_{modular_num}_{file_time}.xlsx"
    path_infor = f"D:\desktop\os\optimization of structure\optimization of structure\optimization of structure\out_all_infor_case4\\run_infor_{num_var}_{modular_num}_{file_time}.xlsx"
    gx_nor = pd.read_excel(io=path_memo, sheet_name="memorize_gx_nor")
    gx_nor_data = gx_nor.values.tolist()

    memorize_pool_pop1 = pd.read_excel(io=path_memo, sheet_name="memorize_pool")
    memorize_pool = memorize_pool_pop1.values.tolist()

    memorize_fit1 = pd.read_excel(io=path_memo, sheet_name="memorize_fit")
    memorize_fit2 = memorize_fit1.values.tolist()
    memorize_fit = []
    for i in range(len(memorize_fit2)):
        memorize_fit.append(memorize_fit2[i][0])

    memorize_weight1 = pd.read_excel(io=path_memo, sheet_name="memorize_weight")
    memorize_weight2 = memorize_weight1.values.tolist()
    memorize_weight = []
    for i in range(len(memorize_weight2)):
        memorize_weight.append(memorize_weight2[i][0])

    memorize_gx1 = pd.read_excel(io=path_memo, sheet_name="memorize_gx")
    memorize_gx = memorize_gx1.values.tolist()

    gx_prediction1 = pd.read_excel(io=path_memo, sheet_name="gx_prediction")
    gx_prediction = gx_prediction1.values.tolist()

    memorize_loss1 = pd.read_excel(io=path_memo, sheet_name="memorize_loss")
    memorize_loss = memorize_loss1.values.tolist()

    memorize_mae1 = pd.read_excel(io=path_memo, sheet_name="memorize_mae")
    memorize_mae = memorize_mae1.values.tolist()

    memorize_gx_nor1 = pd.read_excel(io=path_memo, sheet_name="memorize_gx_nor")
    memorize_gx_nor = memorize_gx_nor1.values.tolist()

    memorize_num1 = pd.read_excel(io=path_memo, sheet_name="memorize_num")
    memorize_num2 = memorize_num1.values.tolist()
    memorize_num = []
    for i in range(len(memorize_num2)):
        memorize_num.append(memorize_num2[i][0])

    pop2_best1 = pd.read_excel(io=path_infor, sheet_name="pop2_all")
    pop2_fitness1 = pd.read_excel(io=path_infor, sheet_name="pop_all_fitness")
    pop2_pool_all = pop2_best1.values.tolist()
    fitness_pool_all = pop2_fitness1.values.tolist()
    pop2_remove = []
    fitness_remove = []
    for i in range(len(pop2_pool_all)):
        if i <= len(pop2_pool_all):
            if type(pop2_pool_all[i][0]) == str:
                pop2_remove.append(i)

    for i in range(len(fitness_pool_all)):
        if i <= len(pop2_pool_all):
            if type(fitness_pool_all[i][0]) == str:
                fitness_remove.append(i)

    for i in range(len(pop2_remove)):
        pop2_pool_all.remove(pop2_pool_all[int(pop2_remove[len(pop2_remove) - 1 - i])])

    for i in range(len(fitness_remove)):
        fitness_pool_all.remove(fitness_pool_all[int(fitness_remove[len(fitness_remove) - 1 - i])])

    pop2_best = pop2_pool_all[(num_continue - 1) * POP_SIZE]
    fitness_best = fitness_pool_all[num_continue - 1][0]
    return pop2_best,memorize_pool,memorize_fit,memorize_weight,memorize_gx,gx_prediction,memorize_loss,memorize_mae,memorize_gx_nor,memorize_num
#续跑算法
def continue_DNN_GA(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time,N1,N2,best_individual):
    pop2, model = DNN_GA(num_var, num_room_type, POP_SIZE, best_individual, 400)
    pop2[0] = best_individual
    pop1, pop3 = decoding_modular_section(pop2)

    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []

    col_up_all= []
    beam_up_all=[]
    pop_all_weight=[]
    pop_all_fitness=[]
    weight_min=[]
    min_ru = []
    sap_run_time = 0
    predict_time = 0
    all_pred = []
    for run_time in range(N1,N2):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)

        # 计算fitness等参数
        fit = [0 for i in range(len(pop2))]
        weight = [0 for i in range(len(pop2))]
        clo_val = [0 for i in range(len(pop2))]
        beam_val = [0 for i in range(len(pop2))]
        result1,weight_pop,clo_up_1,beam_up_1=thread_sap_continue(ModelPath_name,mySapObject_name,SapModel_name,num_thread, pop1, pop2, pop3, fit, weight, clo_val, beam_val)

        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        fitness2 =copy.deepcopy(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        min1 = min(fitness2)
        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop3编码
        min_ru.append(min(fitness2))# 统计历代最小值
        #选择
        pop2 = select_2(pop2, fitness2)
        #交叉变异
        pop2 = crossover_and_mutation_GA_for_DNN(pop2,num_var,CROSSOVER_RATE,MUTATION_RATE)

        # 引入新个体
        run_time +=1
        if run_time % 20 == 0:
            pop2_new,model = DNN_GA(num_var,num_room_type,int(0.9 * len(pop2)),pop2[0],400)
            exchange_num = int(0.9*len(pop2))
            for ex_num in range(exchange_num):
                pop2[len(pop2) - 1 - ex_num] = pop2_new[ex_num]
            memorize_num.append(len(memorize_pool))
            memorize_sum_loacl = []
            memorize_pool_loacl = []
            memorize_fit_loacl = []
            memorize_weight_loacl = []
            memorize_col_loacl = []
            memorize_beam_loacl = []
            memorize_gx_loacl = []

            # 使用深度神经网络对记忆池中的所有个体进行预测
            memorize_pool_temp = copy.deepcopy(memorize_pool)
            memorize_pool_temp = np.array(memorize_pool_temp)
            x_data_prediction = memorize_pool_temp


            fitness_prediction = model.predict(x_data_prediction, verbose=0)
            all_pred.append(fitness_prediction)


        if run_time % 20 == 0:
            print(run_time)
            print(f'记忆池数量:{len(memorize_pool)}')
        pop1, pop3 = decoding_modular_section(pop2)

        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if min1 <=GA_examine(ModelPath_name[0],mySapObject_name[0], SapModel_name[0],aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    out_put_prediction_gx(all_pred, predict_time)

    for i in range(len(mySapObject_name)):
        ret = mySapObject_name[i].ApplicationExit(False)
        SapModel_name[i] = None
        mySapObject_name[i] = None


    out_put_result(pop_zhongqun_all, pop_zhongqun_all_2, pop_zhongqun_all_3,min_ru,weight_min,pop_all_fitness,pop_all_weight,time)
    out_put_fitness_prediction(DNN_prediction_fitness)


    return pop_zhongqun_all,pop_zhongqun_all_2,pop_zhongqun_all_3,fitness_prediction
'''model data'''
# modular size
#建筑参数
modular_length = 8000
modular_width = [4000,4000,5400,3600,3600,4400,4400,4000]
modular_heigth = 3000
modular_length_num = 8
modular_dis = 400
corridor_width = 4000
story_num = 12
story_zone = 4#每组模块的分区数量
story_group = 3#每组模块的楼层数
modular_num = 3#整个建筑的模块种类

zone_num = int(story_num / story_group * story_zone)
section_num = 3 * modular_num
brace_num = modular_num
group_num = int(story_num / story_group)
modular_all = modular_length_num * 2 *story_num

# steel section information
sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                  cfg_file_name="Steel_section_data_I_cube.ini")

# generate model
model_data = dj.generate_model_data(modular_length,modular_width,modular_heigth,modular_length_num,modular_dis,story_num,corridor_width)
nodes = model_data[0]
edges_all = model_data[1]
labels = model_data[2]
cor_edges = model_data[3]
joint_hor = model_data[4]
joint_ver = model_data[5]
room_indx = model_data[6]
#优化参数
DNA_SIZE = 4*story_num+modular_length_num*2*story_num
POP_SIZE = 30
CROSSOVER_RATE = 0.6
MUTATION_RATE = 0.1
N_GENERATIONS = 140
num_thread =10

min_genera = []

num_room_type=1


gx_data_select = [0,1,3,4,5]


x = np.linspace(0, 11, 12)
# x = np.array([2,4,6,8,10,12])

num_room_type=1
#全局记忆池
memorize_pool = []
memorize_fit = []
memorize_weight = []
memorize_col = []
memorize_beam = []
memorize_sum = []
memorize_gx = []
memorize_gx_nor = []
memorize_num = []
sap_run_time00 = 0
#局部记忆池

memorize_sum_local=[]
memorize_pool_local=[]
memorize_fit_local=[]
memorize_weight_local=[]
memorize_col_local=[]
memorize_beam_local=[]
memorize_gx_local=[]
history_loss = []
history_mae = []
DNN_prediction_fitness= []
# label=[1,1,1,1,2,2,2,2]
# labels = []
# for i in range(12):
#     labels.extend(label)
labels = []
labels1 = []
for i in range(group_num):
    temp = []
    for j in range(story_zone):
        for z in range(int(modular_length_num/story_zone)):
            temp.append(i*story_zone+j)
    for j in range(2*story_group):
        labels.extend(temp)
        labels1.append(temp)

for num_var in [5]:
    for time in range(19,20):
        memorize_pool = []
        memorize_fit = []
        memorize_weight = []
        memorize_col = []
        memorize_beam = []
        memorize_sum = []
        memorize_gx = []
        memorize_num = []
        sap_run_time00 = 0

        memorize_sum_local = []
        memorize_pool_local = []
        memorize_fit_local = []
        memorize_weight_local = []
        memorize_col_local = []
        memorize_beam_local = []
        memorize_gx_local = []
        history_loss = []
        history_mae = []
        DNN_prediction_fitness = []
        mySapObject_name, ModelPath_name, SapModel_name =mulit_get_sap(num_thread)
        # zhan,jia,qi=run(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time)
        #跑HIGA用
        zhan, jia, qi,fitness_prediction = GA_DNN_run_modular(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time)
        out_put_memorize(memorize_pool, memorize_fit, memorize_weight, memorize_gx, history_loss, history_mae,
                         memorize_gx_nor, memorize_num, fitness_prediction)
        #跑GA用
        # zhan, jia, qi = GA_run_modular(ModelPath_name, mySapObject_name, SapModel_name, num_var,
        #                                                    num_room_type, x, labels, time)
        #续跑HIGA用
        # best_individual,memorize_pool,memorize_fit,memorize_weight,memorize_gx,gx_prediction,memorize_loss,memorize_mae,memorize_gx_nor,memorize_num = get_continue_data(0,num_continue)
        # zhan, jia, qi,fitness_prediction = continue_DNN_GA(ModelPath_name, mySapObject_name, SapModel_name, num_var, num_room_type, x, labels, time, 140,
        #                 200,np.array(best_individual))
        # out_put_memorize(memorize_pool, memorize_fit, memorize_weight, memorize_gx, history_loss, history_mae,
        #                  memorize_gx_nor, memorize_num, fitness_prediction)
        # draw_loss(num_var, time)
        gc.collect()
