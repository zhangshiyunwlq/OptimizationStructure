import copy
import xlrd
import numpy as np
import pyvista as pv
import openpyxl

def generative_data():
    Joint = [[0,0,0],[3000,0,0],[3000,8000,0],[0,8000,0],[0,0,3000],[3000,0,3000],[3000,8000,3000],[0,8000,3000]]
    mid_point = [[0,4000,0],[3000,4000,0],[0,4000,3000],[3000,4000,0]]
    tri_point = [[0,3000,0],[0,5000,0],[3000,3000,0],[3000,5000,0],[0,3000,3000],[0,5000,3000],[3000,3000,3000],[3000,5000,3000]]

    all_point = []
    all_point.extend(Joint)
    all_point.extend(mid_point)
    all_point.extend(tri_point)
    #构件编号
    column_num = [[0,4],[1,5],[2,6],[3,7]]
    top_beam_num = [[5,6],[4,7],[4,5],[6,7]]
    bottom_beam_num = [[0,3],[1,2],[0,1],[2,3]]
    #支撑编号
    person_brace = [[0,10],[3,10],[1,18],[2,18]]
    exchange_brace = [[3,4],[0,7],[1,6],[2,5]]
    double_exchange_brace = [[0,16],[4,12],[3,17],[7,13],[1,18],[5,14],[2,19],[6,15]]

    all_draw_node = []
    for i in range(len(location)):
        all_point_temp = copy.deepcopy(all_point)
        all_node = []
        for j in range(len(all_point_temp)):
            all_node.append([all_point_temp[j][0] + location[i][0], all_point_temp[j][1] + location[i][1],
                             all_point_temp[j][2] + location[i][2]])
        all_draw_node.extend(all_node)
    all_draw_node = np.array(all_draw_node)

    brace_data = [person_brace,exchange_brace,double_exchange_brace]

    # 模块对应编号
    modular1 = [i for i in range(20)]
    modular_num_all = []
    for i in range(len(location)):
        modular_temp = []
        for j in range(len(modular1)):
            modular_temp.append(modular1[j] + 20 * i)
        modular_num_all.append(modular_temp)

    # 模块对应编号
    modular_se1 = [i for i in range(3)]
    modular_sectione_all = []
    for i in range(len(location)):
        modular_temp = []
        for j in range(len( modular_se1)):
            modular_temp.append(modular_se1[j] + 3 * i)
        modular_sectione_all.append(modular_temp)

    return all_point,column_num,top_beam_num,bottom_beam_num,all_draw_node,brace_data,modular_num_all,modular_sectione_all

def area_sort():
    area = [2080,3642,4392,5292,6660,2256,3584,4400,5600,6800,7600,8400]
    fit_ini = copy.deepcopy(area)
    luyi = copy.deepcopy(area)
    luyi.sort()
    sort_num = []
    lst = list(range(1, len(fit_ini) + 1))
    list_new = []
    for i in range(len(fit_ini)):
        sort_num.append(fit_ini.index(luyi[i]))
    for i in range(len(fit_ini)):
        list_new.append(lst[sort_num[i]])
    list_new = [0, 3, 4, 5, 8, 1, 2, 6, 7, 9, 10, 11]
    return list_new


def draw_3d_modular(brace_dis,brace_type,member_section):
    p = pv.Plotter(shape=(1, 1))
    x = all_draw_node[:, 0]
    y = all_draw_node[:, 1]
    z = all_draw_node[:, 2]
    for j in range(len(modular_num_all)):
        if brace_dis[j] == 1:
            brace = brace_data[int(brace_type)-1]
            brace = np.array(brace)
            for zz in range(len(brace)):
                tube2 = pv.Tube((x[modular_num_all[j][brace[zz, 0]]],
                                 y[modular_num_all[j][brace[zz, 0]]],
                                 z[modular_num_all[j][brace[zz, 0]]]),
                                (x[modular_num_all[j][brace[zz, 1]]],
                                 y[modular_num_all[j][brace[zz, 1]]],
                                 z[modular_num_all[j][brace[zz, 1]]]), radius=40)
                p.add_mesh(tube2, color=[0.5, 0.5, 0.5], show_edges=False)

        column_num_temp = copy.deepcopy(column_num)
        column_num_temp = np.array(column_num_temp).astype(int)

        top_beam_num_temp = copy.deepcopy(top_beam_num)
        top_beam_num_temp = np.array(top_beam_num_temp).astype(int)

        bottom_beam_num_temp = copy.deepcopy(bottom_beam_num)
        bottom_beam_num_temp = np.array(bottom_beam_num_temp).astype(int)

        for zz in range(len(column_num_temp)):
            section_num = int(member_section[modular_sectione_all[j][2]])
            if section_num <= 6:
                color_member = [255, 0, 0]
            else:
                color_member = [0, 0, 255]
            section_size = list_new[section_num] * 10
            tube2 = pv.Tube((x[modular_num_all[j][column_num_temp[zz, 0]]],
                             y[modular_num_all[j][column_num_temp[zz, 0]]],
                             z[modular_num_all[j][column_num_temp[zz, 0]]]),
                            (x[modular_num_all[j][column_num_temp[zz, 1]]],
                             y[modular_num_all[j][column_num_temp[zz, 1]]],
                             z[modular_num_all[j][column_num_temp[zz, 1]]]), radius=section_size)
            p.add_mesh(tube2, color=color_member, show_edges=False)

        for zz in range(len(top_beam_num_temp)):
            section_num = int(member_section[modular_sectione_all[j][0]])
            if section_num <= 6:
                color_member = [255, 0, 0]
            else:
                color_member = [0, 0, 255]
            section_size = list_new[section_num] * 10
            tube2 = pv.Tube((x[modular_num_all[j][top_beam_num_temp[zz, 0]]],
                             y[modular_num_all[j][top_beam_num_temp[zz, 0]]],
                             z[modular_num_all[j][top_beam_num_temp[zz, 0]]]),
                            (x[modular_num_all[j][top_beam_num_temp[zz, 1]]],
                             y[modular_num_all[j][top_beam_num_temp[zz, 1]]],
                             z[modular_num_all[j][top_beam_num_temp[zz, 1]]]), radius=section_size)
            p.add_mesh(tube2, color=color_member, show_edges=False)

        for zz in range(len(bottom_beam_num_temp)):
            section_num = int(member_section[modular_sectione_all[j][1]])
            if section_num <= 6:
                color_member = [255, 0, 0]
            else:
                color_member = [0, 0, 255]
            section_size = list_new[section_num] * 10
            tube2 = pv.Tube((x[modular_num_all[j][bottom_beam_num_temp[zz, 0]]],
                             y[modular_num_all[j][bottom_beam_num_temp[zz, 0]]],
                             z[modular_num_all[j][bottom_beam_num_temp[zz, 0]]]),
                            (x[modular_num_all[j][bottom_beam_num_temp[zz, 1]]],
                             y[modular_num_all[j][bottom_beam_num_temp[zz, 1]]],
                             z[modular_num_all[j][bottom_beam_num_temp[zz, 1]]]), radius=section_size)
            p.add_mesh(tube2, color=color_member, show_edges=False)
    p.set_background('white')
    #摄像机位置、焦点、旋转
    p.camera_position = [(30000, -20000, 12000), (10000, 10000, 0), (0, 0, 10000)]
    p.show()

def draw_all():
    member_section = []
    wb = openpyxl.load_workbook(
        filename=f'D:\desktop\os\optimization of structure\out_all_infor_case4\\run_infor_{num_var}_{modular_num}_{time}.xlsx',
        )
    for pop_num in all_it:
        #获得梁柱截面编号
        sheet2 = wb['pop2_all']
        all_code =[]
        for z in range(num_var):
            rows = sheet2.cell(pop_num * (pop_size + 1) + 2, z + 1).value
            all_code.append(rows)

        member_section_temp = []
        for z in range(num_var+num_room_type,num_var+num_room_type+section_num):
            rows = sheet2.cell(pop_num * (pop_size + 1) + 2, z + 1).value
            member_section_temp.append(rows)

        for i in range(len(member_section_temp)):
            member_section.append(all_code[int(member_section_temp[i])])
        #获得支撑类型
        brace_type = sheet2.cell(pop_num * (pop_size + 1) + 2, num_var + 1).value
        #获得支撑分布情况
        brace_dis = []
        for z in range(num_var+num_room_type+section_num,num_var+num_room_type+section_num+brace_num):
            rows = sheet2.cell(pop_num * (pop_size + 1) + 2, z + 1).value
            brace_dis.append(rows)

        draw_3d_modular(brace_dis,brace_type,member_section)

#模块定位点
# location = [[0,0,0],[6000,0,0],[12000,0,0],[18000,0,0],[24000,0,0],[0,0,12000],[6000,0,12000],[12000,0,12000],[18000,0,12000],[24000,0,12000]]
location = [[0,0,0],[6000,0,0]]
#一个模块所有点坐标，柱编号、顶梁编号、底梁编号、所有模块的节点、支撑数据,模块节点编号
all_point,column_num,top_beam_num,bottom_beam_num,all_draw_node,brace_data,modular_num_all,modular_sectione_all = generative_data()
#按照截面大小排序
list_new = area_sort()

modular_length_num= 8
story_num = 12
story_zone = 4#每组模块的分区数量
story_group = 3#每组模块的楼层数
modular_num = 2#整个建筑的模块种类

zone_num = int(story_num / story_group * story_zone)
section_num = 3 * modular_num
brace_num = modular_num
group_num = int(story_num / story_group)
modular_all = modular_length_num * 2 *story_num


num_var = 4
time = 3
num_room_type = 1
num_room = 1
pop_size = 30#种群数量
all_it = [139]

draw_all()

