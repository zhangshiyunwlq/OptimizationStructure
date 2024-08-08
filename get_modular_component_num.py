import numpy as np
import pandas as pd
import openpyxl
from collections import Counter


def output_data(data):
    for run_time in range(len(data)):
        modular_num = data[run_time][1]
        num_var = data[run_time][0]
        file_name = data[run_time][2]
        iteration = 139

        story_num = 12
        story_zone = 4#每组模块的分区数量
        story_group = 3#每组模块的楼层数
        modular_length_num = 8
        zone_num = int(story_num / story_group * story_zone)
        section_num = 3 * modular_num
        brace_num = modular_num
        group_num = int(story_num / story_group)
        modular_all = modular_length_num * 2 *story_num
        num_room_type = 1

        all_mod = 192
        all_sect = 4*4*3

        wb = openpyxl.load_workbook(
            filename=f"D:\desktop\os\optimization of structure\out_all_infor_case4\\run_infor_{num_var}_{modular_num}_{file_name}.xlsx",
            )
        # sheet1 = wb['pop1_all']
        # for z in range(48):
        #     rows = sheet1.cell(4311,z+1).value
        #     pop_room.append(rows)
        # sheet1 = wb['pop3_all']
        # for z in range(modular_length_num*2*story_num):
        #     rows = sheet1.cell(4311,z+1).value
        #     pop_room_label.append(rows)

        pop1 = []
        sheet1 = wb['pop1_all']
        for z in range(all_sect):
            rows = sheet1.cell(iteration*31+2,z+1).value
            pop1.append(rows)

        pop2_all = []
        sheet1 = wb['pop2_all']
        for z in range(num_var+num_room_type+section_num+brace_num+zone_num):
            rows = sheet1.cell(iteration*31+2,z+1).value
            pop2_all.append(rows)

        pop2 = pop2_all[-16:]


        for z in range(num_var+num_room_type+section_num+brace_num+zone_num):
            rows = sheet1.cell(iteration*31+2,z+1).value
            pop2_all.append(rows)


        modular_section = pop2_all[num_var+num_room_type:num_var+num_room_type+modular_num*3]
        modular_section_decoding = []

        brace_type = pop2_all[num_var+num_room_type+modular_num*3:num_var+num_room_type+modular_num*3+brace_num]
        brace_type_decoding =[]
        for i in range(len(brace_type)):
            if brace_type[i]==0:
                brace_type_decoding.append(0)
            else:
                brace_type_decoding.append(pop2_all[num_var])


        for i in range(len(modular_section)):
            modular_section_decoding.append(pop2_all[modular_section[i]])
        mod_sec = [modular_section_decoding[i:i+3] for i in range(0,len(modular_section_decoding),3)]



        pop1_counter = Counter(pop1)
        pop3_counter = Counter(pop2)

        k = sorted(pop1_counter.items(), key=lambda x: x[0])


        for i in range(len(k)):
            k[i] = list(k[i])
            k[i][1] = k[i][1]*24

        pop2_temp = sorted(pop3_counter.items(), key=lambda x: x[0])


        for i in range(len(pop2_temp)):
            pop2_temp[i] = list(pop2_temp[i])
            pop2_temp[i][1] = pop2_temp[i][1]*6

        print(f'优化结果OP{run_time+1}_{num_var}_{modular_num}_{file_name}')
        print(f'模块种类{mod_sec}')
        print(f'截面数量{k}')
        print(f'模块数量{pop2_temp}')
        print(f'支撑类型{brace_type_decoding}')

data = [[2,3,1],[2,3,0],[2,3,3],[3,3,5],[3,3,4],[3,3,6],[4,3,0],[4,3,1],[4,3,2],[5,3,19],[5,3,0],[5,3,21]]
output_data(data)