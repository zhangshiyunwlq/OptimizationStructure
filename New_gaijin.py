import numpy as np
import random
from random import randint
import copy
import data_to_json as dj
import xlwt
import threading
import queue
import math as m
import os
import sys
import xlrd
import matplotlib.pyplot as plt
import data_to_json as dj
import modular_utils as md
import model_to_sap as ms
import sap_run as sr
import configparser
import comtypes.client
import gc
import openpyxl
from CNN import create_model
# os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'


def generate_DNA_coding_story5(num_var,num_room_type,x):
    all_room_num = 2*story_num*5
    room_nu =np.linspace(1, 3, 3)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+all_room_num))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        sec.sort()
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        room_ty.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = randint(1,3)
        for j in range(num_var+num_room_type,num_var+num_room_type+2*story_num*3):
            pop[i][j] = randint(0,num_var-1)
        for j in range(num_var+num_room_type+2*story_num*3,num_var+num_room_type+2*story_num*5):
            pop[i][j] = randint(0,1)
    return pop

#固定截面尺寸生成编码

def decoding(pop,num_var,num_room_type,labels):
    pop1_jiequ = pop[:,num_var+num_room_type:num_var+num_room_type+2*story_num*3]
    pop1_method = pop[:, num_var+num_room_type+2*story_num*3:num_var+num_room_type+2*story_num*5]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            posi = int(pop1_jiequ[i][j])
            pop_all[i][j] = pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_method[i][z*2+int(labels[j])-1])
                if posi == 0 and np.random.rand() < MUTATION_RATE*2.5:
                    pop_room_label[i][j] = 0
                    # pop_room_label[i][j] = randint(0,3)
                else:
                    pop_room_label[i][j] = pop[i][num_var]
    return pop_all,pop_room_label

def mulitrun_GA(pop1,pop_all,pop3,q,result,weight_1,col_up,beam_up,memory_pools_all,memory_pools_fit,memory_pools_weight,memory_pools_col,memory_pools_beam):
    while True:
        if q.empty():
            break

        time = q.get()
        pop = pop1[time]
        pop_room_label = pop3[time]
        pop2= pop_all[time]
        # 判断记忆池
        for j in range(len(memory_pools_all)):
            tf = []
            for z in range(len(pop2)):
                if pop2[z] != memory_pools_all[j][z]:
                    tf.append(z)
            if len(tf) == 0:
                break
        # 记忆池中存在满足条件的个体
        if len(tf) == 0:
            result[time] = memory_pools_fit[j]
            weight_1[time] = memory_pools_weight[j]
            col_up[time] = memory_pools_col[j]
            beam_up[time] = memory_pools_beam[j]
        # 记忆池中不存在满足条件的个体
        else:
            # pop_all.append(pop)

            res1, res2 = fun(pop,pop_room_label)

            # num3 += 1
            weight_1[time] = res2
            col_up[time] = 1
            beam_up[time] = 1
            result[time] = res1
        # 记忆池更新
            memory_pools_all.append(pop2)
            memory_pools_fit.append(res1)
            memory_pools_weight.append(res2)
            memory_pools_col.append(1)
            memory_pools_beam.append(1)

def mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop_room,pop_room_label):

    sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                      cfg_file_name="Steel_section_data.ini")
    modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
    # 按房间分好节点
    modulars_of_building = modular_building.building_modulars
    modular_nums = len(labels)
    modular_infos = {}
    # 每个房间定义梁柱截面信息
    sr.run_column_room_story1(labels,pop_room_label, modular_length_num * 2 * story_num, sections_data_c1, modular_infos, pop_room)
    #
    for i in range(len(modulars_of_building)):
        modulars_of_building[i].Add_Info_And_Update_Modular(
            # modular_infos[modulars_of_building[i].modular_label - 1])
            modular_infos[i])
    modular_building.Building_Assembling(modulars_of_building)

    all_data = ms.Run_GA_sap_2(mySapObject, ModelPath, SapModel, modular_building,pop_room_label,200,modular_length_num,story_num)
    aa, bb, cc, dd, ee, ff, gg, hh, ii = all_data

    ret = SapModel.SetModelIsLocked(False)
    return aa,bb,cc,dd,ee,ff,gg,hh,ii

def Fun_1(weight,g_col,g_beam,dis_all,all_force,u):
    g_col_all = 0
    g_beam_all = 0
    Y_dis_radio_all = 0
    Y_interdis_all = 0
    Y_interdis_radio_all = 0
    floor_radio = 0
    g_col1 = copy.deepcopy(g_col)
    g_beam1 = copy.deepcopy(g_beam)
    dis_all5 = copy.deepcopy(dis_all[5])
    dis_all7 = copy.deepcopy(dis_all[7])
    for i in range(len(g_col1)):
        if g_col1[i]<= 0:
            g_col1[i] = 0
        else:
            g_col1[i] = g_col1[i]
        g_col_all += g_col1[i]
    for i in range(len(g_beam1)):
        if g_beam1[i]<= 0:
            g_beam1[i] = 0
        else:
            g_beam1[i] = g_beam1[i]
        g_beam_all += g_beam1[i]
    #y dis ratio
    for i in range(len(dis_all5)):
        if dis_all5[i] <= 0.00167 and dis_all5[i] >= -0.00167:
            dis_all5[i] = 0
        else:
            dis_all5[i] = dis_all5[i]
        # Y_dis_radio_all += dis_all5[i]
    Y_dis_radio_all = max(dis_all5)
    Y_dis_radio_all = Y_dis_radio_all*100
    # y interdis max
    for i in range(len(dis_all7)):
        if dis_all7[i] <= 0.004 and dis_all7[i] >= -0.004:
            dis_all7[i] = 0
        else:
            dis_all7[i] = dis_all7[i]
        # Y_interdis_all += dis_all7[i]
    Y_interdis_all = max(dis_all7)
    Y_interdis_all = Y_interdis_all*100
    # # y interdis radio
    # for i in range(len(dis_all[11])):
    #     if dis_all[11][i] <= 1.5 and dis_all[11][i] >= -1.5:
    #         dis_all[11][i] = 0
    #     else:
    #         dis_all[11][i] = dis_all[11][i]
    #     Y_interdis_radio_all += dis_all[11][i]
    # # x interdis ratio
    # for i in range(len(all_force[10])):
    #     if all_force[10][i] <= 1.5 and all_force[10][i] >= -1.5:
    #         all_force[10][i] = 0
    #     else:
    #         all_force[10][i] = all_force[10][i]
    #     floor_radio += all_force[10][i]
    g_col_max= max(g_col)
    g_beam_max = max(g_beam)
    dis_all_max = max(dis_all[5])
    interdis_max = max(dis_all[7])
    g_all_max = max(g_col_max,g_beam_max)
    G_value=u * (abs(g_col_all) + abs(g_beam_all) + abs(Y_dis_radio_all) + abs(Y_interdis_all) + abs(Y_interdis_radio_all))
    gx = [g_col_max,g_beam_max,abs(dis_all_max),abs(interdis_max)]
    # gx_Normalization = [g_col_all,g_beam_all,Y_dis_radio_all,Y_interdis_all]
    result = weight + G_value

    gx_demo = copy.deepcopy(gx)
    if gx_demo[0]>=5:
        gx_demo[0]=1
    elif gx_demo[0]<=-1:
        gx_demo[0] = -1
    elif gx_demo[0]<=5 and gx_demo[0]>=-1:
        gx_demo[0]=(gx_demo[0]+1)/6
    if gx_demo[1]>=2:
        gx_demo[1]=1
    elif gx_demo[1]<=-1:
        gx_demo[1] = -1
    elif gx_demo[1]<=2 and gx_demo[1]>=-1:
        gx_demo[1]=(gx_demo[1]+1)/3
    if gx_demo[2] >= 0.05:
        gx_demo[2] = 0.05
    else:
        gx_demo[2] = gx_demo[2] / 0.05
    if gx_demo[3] >= 0.05:
        gx_demo[3] = 0.05
    else:
        gx_demo[3] = gx_demo[3] / 0.05
    return result,weight,gx,gx_demo

#不加记忆池
def mulitrun_GA_1(ModelPath,mySapObject, SapModel,pop1,pop_all,pop3,q,result,weight_1,col_up,beam_up):
    while True:
        if q.empty():
            break
        time = q.get()
        pop = pop1[time]
        pop_room_label = pop3[time]
        pop2= pop_all[time]
        value = 0
        for i in range(len(memorize_pool)):
            sum_code = sum(pop2)
            if sum_code==memorize_sum[i]:
                pop2_list = copy.deepcopy(pop2)
                memorize_list = copy.deepcopy(memorize_pool[i])
                if pop2_list.tolist()==memorize_list.tolist():
                    result[time]=memorize_fit[i]
                    weight_1[time]=memorize_weight[i]
                    col_up[time]=memorize_col[i]
                    beam_up[time]=memorize_beam[i]
                    value = 1
                    break
        if value ==0:
            we, co, be, r1, r2, r3, r4, dis_all, force_all = mulit_Sap_analy_allroom(ModelPath, mySapObject, SapModel,
                                                                                     pop,
                                                                                     pop_room_label)
            res1, res2,gx,gx_demo = Fun_1(we, co, be, dis_all, force_all, 10000)

            # num3 += 1
            weight_1[time] = res2
            col_up[time] = co
            beam_up[time] = be
            result[time] = res1
            #全局记忆池
            memorize_sum.append(sum(pop2))
            memorize_pool.append(pop2)
            memorize_fit.append(res1)
            memorize_weight.append(res2)
            memorize_col.append(col_up[time])
            memorize_beam.append(beam_up[time])
            memorize_gx.append(gx)
            memorize_gx_nor.append(gx_demo)
            # 全局记忆池
            memorize_sum_local.append(sum(pop2))
            memorize_pool_local.append(pop2)
            memorize_fit_local.append(res1)
            memorize_weight_local.append(res2)
            memorize_col_local.append(col_up[time])
            memorize_beam_local.append(beam_up[time])
            memorize_gx_local.append(gx)


def thread_sap(ModelPath_name,mySapObject_name,SapModel_name,num,pop1,pop2,pop3,result,weight_1,col_up,beam_up):


    pop_n = [0 for i in range(len(pop2[0]))]

    q = queue.Queue()
    threads = []
    for i in range(len(pop1)):
        q.put(i)
    for i in range(num_thread):
        if len(ModelPath_name)!=num_thread:
            for j in range(len(ModelPath_name),num_thread):
                mySapObject_name.append(f"mySapObject{j}")
                SapModel_name.append(f"SapModel{j}")
                ModelPath_name.append(f"ModelPath{j}")
                mySapObject_name[j], ModelPath_name[j], SapModel_name[j] = SAPanalysis_GA_run2(os.path.join(os.getcwd(), f"cases{j}"))
        t = threading.Thread(target=mulitrun_GA_1, args=(ModelPath_name[i],mySapObject_name[i],SapModel_name[i],pop1,pop2,pop3,q,result,weight_1,col_up,beam_up))
        t.start()
        threads.append(t)
    for i in threads:
        i.join()
    return result,weight_1,col_up,beam_up

def select_2(pop, fitness):  # nature selection wrt pop's fitness

    fit_ini = fitness
    luyi = fitness
    luyi.sort(reverse=True)
    sort_num = []
    lst = list(range(1, len(fit_ini)+1))
    list_new = []
    for i in range(len(fit_ini)):
        sort_num.append(fit_ini.index(luyi[i]))
    for i in range(len(fit_ini)):
        list_new.append(lst[sort_num[i]])
    for i in range(len(list_new)):
        list_new[i] = m.e ** (list_new[i] * 1.5)
    idx = np.random.choice(np.arange(POP_SIZE), size=POP_SIZE, replace=True,
                           p=np.array(list_new) / (sum(list_new)))
    pop2 = np.zeros((POP_SIZE, len(pop[0])))
    for i in range(len(pop2)):
        pop2[i] = pop[int(idx[i])]
    return pop2

def crossover_and_mutation_coding_story5(pop2,num_var,num_room_type,CROSSOVER_RATE):
    pop = pop2

    new_pop = np.zeros((len(pop),len(pop[0])))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1_stort5(child,x,num_var,num_room_type,MUTATION_RATE)
        new_pop[i] = child

    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]



    return new_pop

def mutation_1_stort5(child, x,num_var,num_room_type,MUTATION_RATE):
    num_var = int(num_var)
    room_nu = np.linspace(1, 12, 12)
    num_room_type = int(num_room_type)
    if num_var!=len(x):
        for mutate_point in range(num_var):
            x_var = list(map(int, x.tolist()))
            for mutate_point_1 in range(num_var):
                if child[mutate_point_1] in x_var:
                    x_var.remove(child[mutate_point_1])
            if np.random.rand() < MUTATION_RATE:  # 以MUTATION_RATE的概率进行变异
                child[mutate_point] = np.random.choice(x_var)
    for j in range(num_room_type + num_var, num_room_type + num_var + DNA_SIZE):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,num_var-1)
    for j in range(num_var, num_room_type + num_var ):
        if np.random.rand() < MUTATION_RATE:
            # child[j] = randint(1,3)
            child[j] = 0
def mulit_get_sap(num_thread):
    case_name = []
    APIPath_name = []
    mySapObject_name = []
    SapModel_name = []
    ModelPath_name = []
    for i in range(num_thread):
        case_name.append(f"cases{i}")
        APIPath_name.append(os.path.join(os.getcwd(), f"cases{i}"))
        mySapObject_name.append(f"mySapObject{i}")
        SapModel_name.append(f"SapModel{i}")
        ModelPath_name.append(f"ModelPath{i}")
        mySapObject_name[i], ModelPath_name[i], SapModel_name[i] = SAPanalysis_GA_run2(APIPath_name[i])
    return mySapObject_name,ModelPath_name,SapModel_name

def SAPanalysis_GA_run2(APIPath):


    cfg = configparser.ConfigParser()
    cfg.read("Configuration.ini", encoding='utf-8')
    ProgramPath = cfg['SAP2000PATH']['dirpath']
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    AttachToInstance = False
    SpecifyPath = True

    # ModelPath = os.path.join(APIPath, 'API_1-001.sdb')
    helper = comtypes.client.CreateObject('SAP2000v1.Helper')
    helper = helper.QueryInterface(comtypes.gen.SAP2000v1.cHelper)
    if AttachToInstance:
        # attach to a running instance of SAP2000
        try:
            # get the active SapObject
            mySapObject = helper.Getobject("CSI.SAP2000.API.SapObject")
        except (OSError, comtypes.COMError):
            print("No running instance of the program found or failed to attach.")
            sys.exit(-1)
    else:
        if SpecifyPath:
            try:
                # 'create an instance of the SAPObject from the specified path
                mySapObject = helper.CreateObject(ProgramPath)
            except (OSError, comtypes.COMError):
                print("Cannot start a new instance of the program from" + ProgramPath)
                sys.exit(-1)
        else:
            try:
                # create an instance of the SapObject from the latest installed SAP2000
                mySapObject = helper.CreateObjectProgID("CSI.SAP2000.API.SapObject")
            except (OSError, comtypes.COMError):
                print("Cannot start a new instance of the program")
                sys.exit(-1)

        # start SAP2000 application
        mySapObject.ApplicationStart()

    # create SapModel object
    SapModel = mySapObject.SapModel
    # initialize model
    SapModel.InitializeNewModel()
    ModelPath = os.path.join(APIPath, 'API_1-001.sdb')
    # create new blank model
    return mySapObject,ModelPath, SapModel


def GA_examine(ModelPath,mySapObject, SapModel,pop1,pop3):
    result = []
    num1 = 0
    num2 = 0
    num3 = 0
    weight_1 = []
    col_up = []
    beam_up = []
    # if len(pop_all) ==0:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         pop_all.append(pop)
    #         we,co,be,r1,r2,r3,r4 =Sap_analy(pop)
    #         res1,res2 = Fun(we, co, be, 10000)
    #         num3 += 1
    #         weight_1.append(res2)
    #         result.append(res1)
    #         pop_fun_all.append(res1)
    #         pop_weight_all.append(res2)
    # else:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         for i in range(len(pop_all)):
    #             if all(pop == pop_all[i]):
    #                 num1 = 1
    #                 num2 = i
    #                 # result.append(pop_fun_all[i])
    #         if num1 == 0:
    #             pop_all.append(pop)
    #             we, co, be, r1, r2, r3, r4 = Sap_analy(pop)
    #             res1, res2 = Fun(we, co, be, 10000)
    #             num3 += 1
    #             weight_1.append(res2)
    #             pop_weight_all.append(res2)
    #             result.append(res1)
    #             pop_fun_all.append(res1)
    #         elif num1 == 1:
    #             result.append(pop_fun_all[num2])
    #             weight_1.append(pop_weight_all[num2])
    # wb2_examine_ind = openpyxl.Workbook()
    # wb2_examine_ind = openpyxl.load_workbook('examine_individual.xlsx')
    # wb2_pop1_indivi = wb2_examine_ind.create_sheet('pop1_indivi', index=0)
    # wb2_pop3_indivi = wb2_examine_ind.create_sheet('pop3_indivi', index=1)
    loc_3 = 1
    for time in range(len(pop1)):
        pop = pop1[time]
        pop_room_label = pop3[time]
        # _ = wb2_pop1_indivi.cell(row=loc_3, column=1, value=f'{pop1[time]}')
        # _ = wb2_pop3_indivi.cell(row=loc_3, column=1, value=f'{pop3[time]}')
        # loc_3 += 1
        # wb2_examine_ind.save('examine_individual.xlsx')
        # pop_all.append(pop)
        we,co,be,r1,r2,r3,r4,dis_all,force_all =mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop,pop_room_label)
        res1,res2,gx,gx_demo = Fun_1(we, co, be,dis_all,force_all, 10000)
        # num3 += 1
        weight_1.append(res2)
        col_up.append(co)
        beam_up.append(be)
        result.append(res1)
        # pop_fun_all.append(res1)
        # pop_weight_all.append(res2)

    # wb_clear_in1 = openpyxl.load_workbook('examine_individual.xlsx')
    # ws_clear_in1 = wb_clear_in1['pop1_indivi']
    # for row in ws_clear_in1:
    #     for cell in row:
    #         cell.value = None
    # ws_clear_in3 = wb_clear_in1['pop3_indivi']
    # for row in ws_clear_in3:
    #     for cell in row:
    #         cell.value = None
    # wb2_examine_ind.save('examine_individual.xlsx')
    return result,weight_1,col_up,beam_up


def run(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time):
    pop2= generate_DNA_coding_story1(num_var, num_room_type, x)
    pop_decoe_1 = copy.deepcopy(pop2)
    pop1,pop3 = decoding1(pop_decoe_1,num_var,num_room_type,labels)

    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []
    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    col_up_all= []
    beam_up_all=[]
    pop_all_weight=[]
    pop_all_fitness=[]
    weight_min=[]
    min_ru = []
    sap_run_time = 0
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)

        # 计算fitness等参数
        fit = [0 for i in range(len(pop2))]
        weight = [0 for i in range(len(pop2))]
        clo_val = [0 for i in range(len(pop2))]
        beam_val = [0 for i in range(len(pop2))]
        result1,weight_pop,clo_up_1,beam_up_1=thread_sap(ModelPath_name,mySapObject_name,SapModel_name,num_thread, pop1, pop2, pop3, fit, weight, clo_val, beam_val)

        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        fitness2 =result1
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        min1 = min(fitness2)
        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        min_ru.append(min(fitness2))# 统计历代最小值
        #选择
        pop2 = select_2(pop2, fitness2)
        #交叉变异
        pop2 = crossover_and_mutation_coding_story5(pop2, num_var, num_room_type, CROSSOVER_RATE)

        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop2_new = generate_DNA_coding_story1(num_var, num_room_type, x)
            exchange_num = int(0.3*len(pop2_new))
            for ex_num in range(exchange_num):
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]

        if run_time %5==0:
            print(run_time)
            print(f'记忆池数量:{len(memorize_pool)}')
        pop1, pop3 = decoding1(pop2, num_var, num_room_type, labels)

        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if min1 <=GA_examine(ModelPath_name[0],mySapObject_name[0], SapModel_name[0],aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    for i in range(len(mySapObject_name)):
        ret = mySapObject_name[i].ApplicationExit(False)
        SapModel_name[i] = None
        mySapObject_name[i] = None


    out_put_result(pop_zhongqun_all, pop_zhongqun_all_2, pop_zhongqun_all_3,min_ru,weight_min,pop_all_fitness,pop_all_weight,time)



    return pop_zhongqun_all,pop_zhongqun_all_2,pop_zhongqun_all_3
#统计每一代数据
def out_put_result(pop1_all,pop2_all,pop3_all,fitness_all,weight_all,pop_all_fitness,pop_all_weight,time):
    wb1 = xlwt.Workbook()
    out_pop1_all = wb1.add_sheet('pop1_all')
    loc = 0
    for i in range(len(pop1_all)):
        out_pop1_all.write(loc, 0, f'{[i]}')
        for j in range(len(pop1_all[i])):
            loc += 1
            for z in range(len(pop1_all[i][j])):
                out_pop1_all.write(loc, z, pop1_all[i][j][z])
        loc += 1

    out_pop2_all = wb1.add_sheet('pop2_all')
    loc = 0
    for i in range(len(pop2_all)):
        out_pop2_all.write(loc, 0, f'{[i]}')
        for j in range(len(pop2_all[i])):
            loc += 1
            for z in range(len(pop2_all[i][j])):
                out_pop2_all.write(loc, z, pop2_all[i][j][z])
        loc += 1

    out_pop3_all = wb1.add_sheet('pop3_all')
    loc = 0
    for i in range(len(pop3_all)):
        out_pop3_all.write(loc, 0, f'{[i]}')
        for j in range(len(pop3_all[i])):
            loc += 1
            for z in range(len(pop3_all[i][j])):
                out_pop3_all.write(loc, z, pop3_all[i][j][z])
        loc += 1

    pop_all_fit = wb1.add_sheet('pop_all_fitness')
    loc = 0
    for i in range(len(pop_all_fitness)):
        pop_all_fit.write(loc, 0, f'{[i]}')
        loc += 1
        for j in range(len(pop_all_fitness[i])):
            pop_all_fit.write(loc, j, pop_all_fitness[i][j])
        loc += 1


    pop_all_wei = wb1.add_sheet('pop_all_weight')
    loc = 0
    for i in range(len(pop_all_weight)):
        pop_all_wei.write(loc, 0, f'{[i]}')
        loc += 1
        for j in range(len(pop_all_weight[0])):
            pop_all_wei.write(loc, j, pop_all_weight[i][j])
        loc += 1


    outmaxfitness = wb1.add_sheet('max_fitness')
    loc = 0
    outmaxfitness.write(loc, 0, 'max_fitness_all')
    loc += 1
    for i in range(len(fitness_all)):
        outmaxfitness.write(loc, i, fitness_all[i])
    loc += 1
    outmaxfitness.write(loc, 0, 'min_weight_all')
    loc += 1
    for i in range(len(weight_all)):
        outmaxfitness.write(loc, i, weight_all[i])


    APIPath = os.path.join(os.getcwd(), 'out_all_infor')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    path1 = os.path.join(APIPath, f'run_infor_{num_var}_{time}')


    wb1.save(f'{path1}.xls')

#统计记忆池
def out_put_memorize(memorize_pool,memorize_fit,memorize_weight,memorize_gx,memorize_loss,memorize_mae,memorize_gx_nor,memorize_num):
    wb1 = xlwt.Workbook()
    out_pop1_all = wb1.add_sheet('memorize_pool')
    loc = 0

    for i in range(len(memorize_pool)):
        pool_list = copy.deepcopy(memorize_pool[i])
        pool_list.tolist()
        for j in range(len(pool_list)):
            out_pop1_all.write(loc, j, pool_list[j])
        loc += 1

    pop_all_fit = wb1.add_sheet('memorize_fit')

    for i in range(len(memorize_fit)):
        pop_all_fit.write(i, 0, memorize_fit[i])

    pop_all_wei = wb1.add_sheet('memorize_weight')

    for i in range(len(memorize_weight)):
        pop_all_wei.write(i, 0, memorize_weight[i])


    memo_gx = wb1.add_sheet('memorize_gx')
    loc = 0
    for i in range(len(memorize_gx)):
        for j in range(len(memorize_gx[i])):
            memo_gx.write(loc, j, memorize_gx[i][j])
        loc += 1

    memo_loss = wb1.add_sheet('memorize_loss')

    for i in range(len(memorize_loss)):
        memo_loss.write(i, 0, memorize_loss[i])

    memo_mae = wb1.add_sheet('memorize_mae')

    for i in range(len(memorize_mae)):
        memo_mae.write(i, 0, memorize_mae[i])

    memo_gx_nor = wb1.add_sheet('memorize_gx_nor')
    loc = 0
    for i in range(len(memorize_gx_nor)):
        for j in range(len(memorize_gx_nor[i])):
            memo_gx_nor.write(loc, j, memorize_gx_nor[i][j])
        loc += 1

    memo_num = wb1.add_sheet('memorize_num')

    for i in range(len(memorize_num)):
        memo_num.write(i, 0, memorize_num[i])



    APIPath = os.path.join(os.getcwd(), 'out_all_memorize')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    path1 = os.path.join(APIPath, f'memorize_infor_{num_var}_{time}')

    wb1.save(f'{path1}.xls')


def draw_picture(name,title_name):
    value_str = []
    # filename = f'E:\wlq_room\optimization of structure\out_all_infor\\run_infor_0.xls',
    # formatting_info = True)

    wb = xlrd.open_workbook(
        filename=f'D:\desktop\os\optimization of structure\optimization of structure\optimization of structure\out_all_infor\\run_infor.xls',
        formatting_info=True)
    sheet1 = wb.sheet_by_index(5)
    for z in range(N_GENERATIONS):
        rows = sheet1.row_values(1)[z]
        value_str.append(rows)
    print(value_str[0])

    num1 = 0.8
    num2 = 0.75
    num3 = 3
    num4 = 0
    fig2 = plt.figure(num=1, figsize=(23, 30))
    ax2 = fig2.add_subplot(111)
    ax2.tick_params(labelsize=40)
    ax2.set_xlabel("Iteration",fontsize=50)  # 添加x轴坐标标签，后面看来没必要会删除它，这里只是为了演示一下。
    ax2.set_ylabel(title_name, fontsize=50)  # 添加y轴标签，设置字体大小为16，这里也可以设字体样式与颜色
    ax2.spines['bottom'].set_linewidth(4);###设置底部坐标轴的粗细
    ax2.spines['left'].set_linewidth(4)
    ax2.spines['right'].set_color('none')
    ax2.spines['top'].set_color('none')
    plt.ylim((0, 1000))
    info = copy.deepcopy(value_str)

    for j in range(len(info)):
        if info[j]>=500:
        # info[i][j] = 500+100*(m.log(info[i][j]))
            info[j] = 500 + info[j]/1000


    bbb = np.arange(0, len(info))
    ccc = info
    ax2.plot(bbb, ccc, label = name,linewidth=6)
    ax2.legend(bbox_to_anchor=(num1, num2), loc=num3, borderaxespad=num4,  handlelength=1.5, fontsize=30, shadow=False)

    plt.show()

#每层一种变量
def generate_DNA_coding_story1(num_var,num_room_type,x):
    all_room_num = story_num*4
    room_nu =np.linspace(1, 3, 3)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+all_room_num))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        sec.sort()
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        room_ty.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            # pop[i][j] = randint(1,3)
            pop[i][j] = 0
        for j in range(num_var+num_room_type,num_var+num_room_type+story_num*3):
            pop[i][j] = randint(0,num_var-1)
        for j in range(num_var+num_room_type+story_num*3,num_var+num_room_type+story_num*4):
            # pop[i][j] = randint(0,1)
            pop[i][j] = 0
    return pop

def decoding1(pop,num_var,num_room_type,labels):
    pop1_jiequ = pop[:,num_var+num_room_type:num_var+num_room_type+story_num*3]
    pop1_method = pop[:, num_var+num_room_type+story_num*3:num_var+num_room_type+story_num*4]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            posi = int(pop1_jiequ[i][j])
            pop_all[i][j] = pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_method[i][int(labels[j])-1])
                if posi == 0:
                    pop_room_label[i][j] = 0
                else:
                    # pop_room_label[i][j] = pop[i][num_var]
                    pop_room_label[i][j] = 0
    return pop_all,pop_room_label

#在最优个体附近生成种群
def generation_population(best_indivi,rate):
    best_in = copy.deepcopy(best_indivi)
    best_in.tolist()
    pop = np.zeros((POP_SIZE,len(best_in)))
    for i in range(len(pop)):
        for j in range(len(pop[i])):
            if np.random.rand() < rate:
                pop[i][j] = randint(0,num_var-1)
            else:
                pop[i][j] = best_in[j]
    return pop

#用于将GA_forDNN中的gx格式转换
def Gx_convert(fitness1):
    fitness3 = copy.deepcopy(fitness1)
    fitness4 = []  # 储存所有gx
    fitness2 = []  # 所有gx的和
    for j in range(len(fitness3)):
        fitness4.append(fitness3[j].tolist())
    fitness4=gx_nonNormalization(fitness4)
    for j in range(len(fitness3)):
        fitness2.append(sum(fitness4[j]))
    return fitness2

#用于GA_forDNN中的交叉变异
def crossover_and_mutation_GA_for_DNN(pop2,num_var,CROSSOVER_RATE,MUTATION_RATE):
    pop = pop2

    new_pop = np.zeros((len(pop),len(pop[0])))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_GA_for_DNN(child,num_var,MUTATION_RATE)
        new_pop[i] = child

    return new_pop

#用于GA_forDNN中的交叉变异
def mutation_GA_for_DNN(child,num_var,MUTATION_RATE):
    num_var = int(num_var)
    for j in range(len(child)):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,num_var-1)

#将gx中的数据归一化
def gx_Normalization(gx):
    gx_demo = copy.deepcopy(gx)
    for i in range(len(gx_demo)):
        if gx_demo[i][0]>=5:
            gx_demo[i][0]=1
        elif gx_demo[i][0]<=-1:
            gx_demo[i][0] = -1
        elif gx_demo[i][0]<=5 and gx_demo[i][0]>=-1:
            gx_demo[i][0]=(gx_demo[i][0]+1)/6
        if gx_demo[i][1]>=2:
            gx_demo[i][1]=1
        elif gx_demo[i][1]<=-1:
            gx_demo[i][1] = -1
        elif gx_demo[i][1]<=2 and gx_demo[i][1]>=-1:
            gx_demo[i][1]=(gx_demo[i][1]+1)/3
        if gx_demo[i][2] >= 0.05:
            gx_demo[i][2] = 0.05
        else:
            gx_demo[i][2] = gx_demo[i][2] / 0.05
        if gx_demo[i][3] >= 0.05:
            gx_demo[i][3] = 0.05
        else:
            gx_demo[i][3] = gx_demo[i][3] / 0.05
    return gx_demo

def gx_nonNormalization(gx):
    gx_demo = copy.deepcopy(gx)
    for i in range(len(gx_demo)):
        gx_demo[i][0]=gx_demo[i][0]*6-1
        gx_demo[i][1] = gx_demo[i][1] * 3-1
        gx_demo[i][2] = gx_demo[i][2] * 0.05
        gx_demo[i][3] = gx_demo[i][3] * 0.05
    return gx_demo
#用于神经网络训练的GA
def GA_for_DNN(run_time,pop2,model):
    for i in range(run_time):
        fitness1 = model.predict(pop2,verbose=0)
        fitness2 = Gx_convert(fitness1)#归一化还原，并将每个染色体对应的gx累加
        mm = fitness2.index(min(fitness2))
        min1 = min(fitness2)
        mm2_all = pop2[mm]
        #选择
        pop2 = select_2(pop2, fitness2)
        # 交叉变异
        pop2 = crossover_and_mutation_GA_for_DNN(pop2, num_var,CROSSOVER_RATE,MUTATION_RATE)
        fit_pred = model.predict(pop2,verbose=0)
        fit_pred2=Gx_convert(fit_pred)
        if min1 <= fit_pred2[0]:
            pop2[0] = mm2_all
    return pop2
#通过神经网络预测新个体
def DNN_GA(num_var,num_room_type,num_ind,best_indivi,run_time):
    #局部训练
    pool_local = copy.deepcopy(memorize_pool_local)
    x_train1_local = np.array(pool_local)
    x_train_local = x_train1_local[:,num_var+num_room_type:num_var+num_room_type+3*story_num]#提取用于训练的x_train部分
    gx_local = copy.deepcopy(memorize_gx_local)
    y_train_local = np.array(gx_local)
    y_train_local= gx_Normalization(y_train_local)#归一化
    model= create_model(len(x_train_local[0]), len(y_train_local[0]))#创建模型
    #verbose取消打印损失
    model.fit(x_train_local, y_train_local, epochs=100, batch_size=32,verbose=0)#训练模型

    #全局训练
    pool_global = copy.deepcopy(memorize_pool)
    gx_global = copy.deepcopy(memorize_gx)
    x_train1 = np.array(pool_global)
    x_train = x_train1[:,num_var+num_room_type:num_var+num_room_type+3*story_num]#提取用于训练的x_train部分
    y_train = np.array(gx_global)
    y_train = gx_Normalization(y_train)#归一化
    model = create_model(len(x_train[0]),len(y_train[0]))#创建模型
    history=model.fit(x_train, y_train, epochs=100, batch_size=32,verbose=0)#训练模型
    # history_loss.extend(history.history['loss'])
    # history_mae.extend(history.history['mae'])
    history_loss.append(history.history['loss'][len(history.history['loss'])-1])
    history_mae.append(history.history['mae'][len(history.history['loss'])-1])
    pop_best = []
    for i in range(num_ind):
        pop1 = generation_population(best_indivi, 0.2)#根据最好个体生成种群
        pop2 = pop1[:,num_var+num_room_type:num_var+num_room_type+3*story_num]
        pop2 = GA_for_DNN(run_time, pop2, model)
        pop_best.append(pop2[0].tolist())
    pop_best = np.array(pop_best)
    return pop_best

def GA_DNN_run(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time):
    pop2= generate_DNA_coding_story1(num_var, num_room_type, x)
    pop_decoe_1 = copy.deepcopy(pop2)
    pop1,pop3 = decoding1(pop_decoe_1,num_var,num_room_type,labels)

    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []
    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    col_up_all= []
    beam_up_all=[]
    pop_all_weight=[]
    pop_all_fitness=[]
    weight_min=[]
    min_ru = []
    sap_run_time = 0
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)

        # 计算fitness等参数
        fit = [0 for i in range(len(pop2))]
        weight = [0 for i in range(len(pop2))]
        clo_val = [0 for i in range(len(pop2))]
        beam_val = [0 for i in range(len(pop2))]
        result1,weight_pop,clo_up_1,beam_up_1=thread_sap(ModelPath_name,mySapObject_name,SapModel_name,num_thread, pop1, pop2, pop3, fit, weight, clo_val, beam_val)

        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        fitness2 =result1
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        min1 = min(fitness2)
        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        min_ru.append(min(fitness2))# 统计历代最小值
        #选择
        pop2 = select_2(pop2, fitness2)
        #交叉变异
        pop2 = crossover_and_mutation_coding_story5(pop2, num_var, num_room_type, CROSSOVER_RATE)

        # 引入新个体
        run_time +=1
        if run_time % 15 == 0:
            pop2_new = DNN_GA(num_var,num_room_type,int(0.3 * len(pop2)),pop2[0],200)
            exchange_num = int(0.3*len(pop2_new))
            for ex_num in range(exchange_num):
                for indi in range(len(pop2_new)):
                    pop2[len(pop1) - 1 - ex_num][indi+num_var+num_room_type] = pop2_new[ex_num][indi]
            memorize_num.append(len(memorize_pool))
            memorize_sum_loacl = []
            memorize_pool_loacl = []
            memorize_fit_loacl = []
            memorize_weight_loacl = []
            memorize_col_loacl = []
            memorize_beam_loacl = []
            memorize_gx_loacl = []

        if run_time %15==0:
            print(run_time)
            print(f'记忆池数量:{len(memorize_pool)}')
        pop1, pop3 = decoding1(pop2, num_var, num_room_type, labels)

        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if min1 <=GA_examine(ModelPath_name[0],mySapObject_name[0], SapModel_name[0],aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    for i in range(len(mySapObject_name)):
        ret = mySapObject_name[i].ApplicationExit(False)
        SapModel_name[i] = None
        mySapObject_name[i] = None


    out_put_result(pop_zhongqun_all, pop_zhongqun_all_2, pop_zhongqun_all_3,min_ru,weight_min,pop_all_fitness,pop_all_weight,time)



    return pop_zhongqun_all,pop_zhongqun_all_2,pop_zhongqun_all_3

def draw_loss(num_var,time):
    fig = plt.figure(1)
    ax1 = fig.add_subplot()
    dev_x = np.arange(0, len(history_loss))
    dev_y = history_loss
    ax1.set_xlabel("Epoch",fontsize=10)
    ax1.set_ylabel("loss", fontsize=10)

    APIPath = os.path.join(os.getcwd(), f'loss_mae')
    SpecifyPath = True
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass
    path1 = os.path.join(APIPath, f'loss{num_var}_{time}')
    ax1.plot(dev_x, dev_y)
    plt.savefig(path1, dpi=300)
    plt.close()
    #绘制mae图
    fig2 = plt.figure(2)
    ax2 = fig2.add_subplot()
    ax2.set_xlabel("Epoch",fontsize=10)
    ax2.set_ylabel("Mae", fontsize=10)
    path1 = os.path.join(APIPath, f'mae{num_var}_{time}')
    dev_x = np.arange(0, len(history_loss))
    dev_y = history_mae
    ax2.plot(dev_x, dev_y)
    plt.savefig(path1, dpi=300)
    plt.close()

'''model data'''
# modular size
modular_length = 8000
modular_width = [4000,4000,5400,3600,3600,4400,4400,4000]
modular_heigth = 3000
modular_length_num = 8
modular_dis = 400
story_num = 6
corridor_width = 4000

# steel section information
sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                  cfg_file_name="Steel_section_data.ini")

# generate model
model_data = dj.generate_model_data(modular_length,modular_width,modular_heigth,modular_length_num,modular_dis,story_num,corridor_width)
nodes = model_data[0]
edges_all = model_data[1]
labels = model_data[2]
cor_edges = model_data[3]
joint_hor = model_data[4]
joint_ver = model_data[5]
room_indx = model_data[6]


POP_SIZE =30
DNA_SIZE = story_num*3
CROSSOVER_RATE = 0.6
MUTATION_RATE = 0.1
N_GENERATIONS = 150
num_thread = 10
min_genera = []

x = np.linspace(0, 13, 14)
# x = np.array([2,4,6,8,10,12])
num_var = 2
num_room_type=1
#全局记忆池
memorize_pool = []
memorize_fit = []
memorize_weight = []
memorize_col = []
memorize_beam = []
memorize_sum = []
memorize_gx = []
memorize_gx_nor = []
memorize_num = []

#局部记忆池

memorize_sum_local=[]
memorize_pool_local=[]
memorize_fit_local=[]
memorize_weight_local=[]
memorize_col_local=[]
memorize_beam_local=[]
memorize_gx_local=[]
history_loss = []
history_mae = []
# label=[1,1,1,1,2,2,2,2]
# labels = []
# for i in range(12):
#     labels.extend(label)
labels = []
for i in range(1,7):
    for j in range(16):
        labels.append(i)


for num_var in [14]:
    for time in range(44,49):
        memorize_pool = []
        memorize_fit = []
        memorize_weight = []
        memorize_col = []
        memorize_beam = []
        memorize_sum = []
        memorize_gx = []
        memorize_num = []

        memorize_sum_local = []
        memorize_pool_local = []
        memorize_fit_local = []
        memorize_weight_local = []
        memorize_col_local = []
        memorize_beam_local = []
        memorize_gx_local = []
        history_loss = []
        history_mae = []
        mySapObject_name, ModelPath_name, SapModel_name =mulit_get_sap(num_thread)
        # zhan,jia,qi=run(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time)
        zhan, jia, qi = GA_DNN_run(ModelPath_name,mySapObject_name,SapModel_name,num_var,num_room_type,x,labels,time)
        out_put_memorize(memorize_pool, memorize_fit, memorize_weight, memorize_gx,history_loss,history_mae,memorize_num)
        draw_loss(num_var, time)
        gc.collect()

# draw_picture('name','title')

