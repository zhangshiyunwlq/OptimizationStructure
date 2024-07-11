import copy
import numpy as np
import pyvista as pv
import data_to_json as dj
import modular_utils as md
import model_to_sap as ms
import os
import sap_run as sr
import sys
import comtypes.client
import win32com.client
import xlwt
import math as m
import configparser
import time
import random
from random import randint
import xlwt
import xlrd
import gc
import openpyxl
import threading
import queue
from collections import Counter



import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D


def model_up_vis(model_data,all_up_indx_1):
    all_up_indx = []
    for i in range(len(all_up_indx_1)):
        all_up_indx.append(all_up_indx_1[i][0]*12+all_up_indx_1[i][1])
    p = pv.Plotter(shape=(1, 1))
    x = model_data[0][:, 0]
    y = model_data[0][:, 1]
    z = model_data[0][:, 2]
    model_data[4] = np.array(model_data[4])
    model_data[4] = model_data[4].astype(int)
    for i in range(len(model_data[4])):
        tube2 = pv.Tube((x[model_data[4][i, 0]], y[model_data[4][i, 0]], z[model_data[4][i, 0]]),
                        (x[model_data[4][i, 1]], y[model_data[4][i, 1]], z[model_data[4][i, 1]]), radius=100)
        p.add_mesh(tube2, color=[0.5, 0.5, 0.5], show_edges=False)

    model_data[5] = np.array(model_data[5])
    model_data[5] = model_data[5].astype(int)
    for i in range(len(model_data[5])):
        tube2 = pv.Tube((x[model_data[5][i, 0]], y[model_data[5][i, 0]], z[model_data[5][i, 0]]),
                        (x[model_data[5][i, 1]], y[model_data[5][i, 1]], z[model_data[5][i, 1]]), radius=100)
        p.add_mesh(tube2, color=[0.5, 0.5, 0.5], show_edges=False)

    model_data[1] = np.array(model_data[1])
    model_data[1] = model_data[1].astype(int)

    for i in range(len(model_data[1])):

        if i in all_up_indx:
            tube2 = pv.Tube((x[model_data[1][i, 0]], y[model_data[1][i, 0]], z[model_data[1][i, 0]]),
                            (x[model_data[1][i, 1]], y[model_data[1][i, 1]], z[model_data[1][i, 1]]), radius=100)
            p.add_mesh(tube2, color=[238, 64, 0], show_edges=False)

        else:
            tube2 = pv.Tube((x[model_data[1][i, 0]], y[model_data[1][i, 0]], z[model_data[1][i, 0]]),
                            (x[model_data[1][i, 1]], y[model_data[1][i, 1]], z[model_data[1][i, 1]]), radius=100)

            p.add_mesh(tube2, color=[0.5, 0.5, 0.5], show_edges=False)

    p.set_background('white')
    p.show()

def mulit_get_sap(num_thread):
    case_name = []
    APIPath_name = []
    mySapObject_name = []
    SapModel_name = []
    ModelPath_name = []
    for i in range(num_thread):
        case_name.append(f"cases{i}")
        APIPath_name.append(os.path.join(os.getcwd(), f"cases{i}"))
        mySapObject_name.append(f"mySapObject{i}")
        SapModel_name.append(f"SapModel{i}")
        ModelPath_name.append(f"ModelPath{i}")
        mySapObject_name[i], ModelPath_name[i], SapModel_name[i] = SAPanalysis_GA_run2(APIPath_name[i])
    return mySapObject_name,ModelPath_name,SapModel_name

def output_GA_in(info,var_num,num_room_type,num):
    pop_zhongqun_all = info[4]
    pop_zhongqun_all_2 = info[5]
    pop_all_fitness = info[7]
    pop_all_weight = info[8]
    col_up_all = info[9]
    beam_up_all = info[10]
    max_ru = info[2]
    weight_min = info[6]
    wb1 = xlwt.Workbook()
    outzhongqun = wb1.add_sheet('pop_zhongqun_all')
    loc = 0
    for i in range(len(pop_zhongqun_all)):
        outzhongqun.write(loc, 0, f'{[i]}')
        for j in range(len(pop_zhongqun_all[i])):
            loc += 1
            outzhongqun.write(loc, 0, f'{pop_zhongqun_all[i][j]}')
        loc += 1
    outzhongqunall = wb1.add_sheet('pop_zhongqun_all_2')
    loc = 0
    for i in range(len(pop_zhongqun_all_2)):
        outzhongqunall.write(loc, 0, f'{[i]}')
        for j in range(len(pop_zhongqun_all_2[i])):
            loc += 1
            outzhongqunall.write(loc, 0, f'{pop_zhongqun_all_2[i][j]}')
        loc += 1

    outpopfitness = wb1.add_sheet('pop_firness_all')
    loc = 0
    for i in range(len(pop_all_fitness)):
        outpopfitness.write(loc, 0, f'{[i]}')
        loc += 1
        for j in range(len(pop_all_fitness[0])):

            outpopfitness.write(loc, j, f'{pop_all_fitness[i][j]}')
        loc += 1

    outpopweight = wb1.add_sheet('pop_weight_all')
    loc = 0
    for i in range(len(pop_all_weight)):
        outpopweight.write(loc, 0, f'{[i]}')
        loc += 1
        for j in range(len(pop_all_weight[0])):
            outpopweight.write(loc, j, f'{pop_all_weight[i][j]}')
        loc += 1
    outpopcol = wb1.add_sheet('pop_col_up_all')
    loc = 0
    for i in range(len(col_up_all)):
        outpopcol.write(loc, 0, f'{[i]}')
        for z in range(len(col_up_all[0])):
            loc += 1
            for j in range(len(col_up_all[0][0])):
                outpopcol.write(loc, j, f'{col_up_all[i][z][j]}')
        loc += 1
    outpopbeam = wb1.add_sheet('pop_beam_up_all')
    loc = 0
    for i in range(len(beam_up_all)):
        outpopbeam.write(loc, 0, f'{[i]}')
        for z in range(len(beam_up_all[0])):
            loc += 1
            for j in range(len(beam_up_all[0][0])):
                outpopbeam.write(loc, j, f'{beam_up_all[i][z][j]}')
        loc += 1
    outmaxfitness = wb1.add_sheet('max_fitness')
    loc = 0
    outmaxfitness.write(loc, 0, 'max_fitness_all')
    loc += 1
    for i in range(len(max_ru)):
        outmaxfitness.write(loc, i, f'{max_ru[i]}')
    loc += 1
    outmaxfitness.write(loc, 0, 'min_weight_all')
    loc += 1
    for i in range(len(weight_min)):
        outmaxfitness.write(loc, i, f'{weight_min[i]}')

    wb1.save(f'out_sap_run_{var_num}_{num_room_type}_{num}.xls')

def crossover_and_mutation_1(pop2,num_var,num_room_type,CROSSOVER_RATE):
    pop = pop2
    room_nu = np.linspace(1, 12, 12)
    new_pop = np.zeros((len(pop),len(pop[0])))
    #前1/3
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1(child,x,num_var,num_room_type,MUTATION_RATE)




        new_pop[i] = child
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop[i][j+num_var+num_room_type])
            if num_var>=num_room_type:
                pop_all[i][j] = pop[i][posi]
                while posi>=num_room_type:
                    posi = posi-num_room_type
                pop_room_label[i][j] = pop[i][posi+num_var]
            else:
                pop_room_label[i][j] = pop[i][posi + num_var]
                while posi>=num_var:
                    posi = posi-num_var
                pop_all[i][j] = pop[i][posi]
    return pop_all,new_pop,pop_room_label
# 对截面库和房间库编码进行排序 变异及生成种群
def crossover_and_mutation_sort(pop2,num_var,num_room_type,CROSSOVER_RATE):
    pop = pop2
    room_nu = np.linspace(1, 12, 12)
    new_pop = np.zeros((len(pop),len(pop[0])))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1(child,x,num_var,num_room_type,MUTATION_RATE)
        new_pop[i] = child
    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var, num_var+num_room_type):
            room_sort.append(new_pop[i][j])
        room_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]
        for j in range(num_var,num_var+num_room_type):
            new_pop[i][j] = room_sort[j-num_var]
    pop1_jiequ = new_pop[:,num_var+num_room_type:num_var+num_room_type+DNA_SIZE]

    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop1_jiequ[i][j])
            while posi>= num_var:
                posi = posi- num_var
            pop_all[i][j] = new_pop[i][posi]
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop1_jiequ[i][j])
            while posi>= num_room_type:
                posi = posi- num_room_type
            pop_room_label[i][j] = new_pop[i][posi+num_var]
    return pop_all,new_pop,pop_room_label

def generate_DNA_sort_ku(num_var,num_room_type,x):
    room_nu =np.linspace(1, 12, 12)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+DNA_SIZE))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        sec.sort()
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        room_ty.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = room_ty[j-num_var]
        for j in range(num_var+num_room_type,num_var+num_room_type+DNA_SIZE):
            pop[i][j] = randint(0,max(num_var,num_room_type)-1)
    pop1_jiequ = pop[:,num_var+num_room_type:num_var+num_room_type+DNA_SIZE]

    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop1_jiequ[i][j])
            while posi>= num_var:
                posi = posi- num_var
            pop_all[i][j] = pop[i][posi]
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop1_jiequ[i][j])
            while posi>= num_room_type:
                posi = posi- num_room_type
            pop_room_label[i][j] = pop[i][posi+num_var]

    return pop_all,pop,pop_room_label

# 递进式交叉变异
def crossover_and_mutation_sort_progressive(pop2,num_var,num_room_type,CROSSOVER_RATE):
    pop = pop2
    room_nu = np.linspace(1, 12, 12)
    new_pop = np.zeros((len(pop),len(pop[0])))
    #前1/3
    for i in range(0,int(len(pop)/3)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE/50:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1_progressive_1(child,x,num_var,num_room_type,MUTATION_RATE,0)
        new_pop[i] = child
        #1/3--2/3
        for i in range(int(len(pop)/3),int(len(pop)/3*2)):
            father = pop[i]
            child = father
            if np.random.rand() < CROSSOVER_RATE:
                mother = pop[np.random.randint(POP_SIZE)]
                cross_points1 = np.random.randint(low=0, high=len(pop[0]))
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                while cross_points2 == cross_points1:
                    cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                exchan = []
                exchan.append(cross_points2)
                exchan.append(cross_points1)
                for j in range(min(exchan), max(exchan)):
                    child[j] = mother[j]
            mutation_1_progressive_1(child, x, num_var, num_room_type, MUTATION_RATE,0.5)
            new_pop[i] = child
        # 2/3--3/3
        for i in range(int(len(pop)/3*2),len(pop)):
            father = pop[i]
            child = father
            if np.random.rand() < CROSSOVER_RATE:
                mother = pop[np.random.randint(POP_SIZE)]
                cross_points1 = np.random.randint(low=0, high=len(pop[0]))
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                while cross_points2 == cross_points1:
                    cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                exchan = []
                exchan.append(cross_points2)
                exchan.append(cross_points1)
                for j in range(min(exchan), max(exchan)):
                    child[j] = mother[j]
            mutation_1_progressive_1(child, x, num_var, num_room_type, MUTATION_RATE,1)
            new_pop[i] = child
    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var, num_var+num_room_type):
            room_sort.append(new_pop[i][j])
        room_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]
        for j in range(num_var,num_var+num_room_type):
            new_pop[i][j] = room_sort[j-num_var]
    pop1_jiequ = new_pop[:,num_var+num_room_type:num_var+num_room_type+DNA_SIZE]

    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop1_jiequ[i][j])
            while posi>= num_var:
                posi = posi- num_var
            pop_all[i][j] = new_pop[i][posi]
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop1_jiequ[i][j])
            while posi>= num_room_type:
                posi = posi- num_room_type
            pop_room_label[i][j] = new_pop[i][posi+num_var]
    return pop_all,new_pop,pop_room_label

def mutation_1_progressive_1(child, x,num_var,num_room_type,MUTATION_RATE,rate):
    num_var = int(num_var)
    room_nu = np.linspace(1, 12, 12)
    num_room_type = int(num_room_type)
    for mutate_point in range(num_var):
        x_var = list(map(int, x.tolist()))
        for mutate_point_1 in range(num_var):
            if child[mutate_point_1] in x_var:
                x_var.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE*rate:  # 以MUTATION_RATE的概率进行变异
            child[mutate_point] = np.random.choice(x_var)
    for mutate_point in range(num_var,num_room_type+num_var):
        x_room = list(map(int, room_nu.tolist()))
        for mutate_point_1 in range(num_var, num_room_type + num_var):
            if child[mutate_point_1] in x_room:
                x_room.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE*rate:
            child[mutate_point] = np.random.choice(x_room)
    for j in range(num_room_type + num_var, num_room_type + num_var + DNA_SIZE):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,max(num_var,num_room_type)-1)

def mutation_1(child, x,num_var,num_room_type,MUTATION_RATE):
    num_var = int(num_var)
    room_nu = np.linspace(1, 12, 12)
    num_room_type = int(num_room_type)
    for mutate_point in range(num_var):
        x_var = list(map(int, x.tolist()))
        for mutate_point_1 in range(num_var):
            if child[mutate_point_1] in x_var:
                x_var.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE:  # 以MUTATION_RATE的概率进行变异
            child[mutate_point] = np.random.choice(x_var)
    for mutate_point in range(num_var,num_room_type+num_var):
        x_room = list(map(int, room_nu.tolist()))
        for mutate_point_1 in range(num_var, num_room_type + num_var):
            if child[mutate_point_1] in x_room:
                x_room.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE:
            child[mutate_point] = np.random.choice(x_room)
    for j in range(num_room_type + num_var, num_room_type + num_var + DNA_SIZE):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,max(num_var,num_room_type)-1)

def Get_fitness_1(result):
    fitness2 = []
    for i in range(len(result)):
        # fitness2.append(m.log(result[i]))
        fitness2.append(result[i])
    return fitness2

# 库数量+变化类型(5)+房间类型数量(30),库截面排序
def generate_DNA_sort_ku_30(num_var,num_room_type,x,labels):
    all_room_num = 5*story_num
    room_nu =np.linspace(1, 12, 12)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+4+all_room_num))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        sec.sort()
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        room_ty.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = room_ty[j-num_var]
        for j in range(num_var+num_room_type,num_var+num_room_type+4+all_room_num):
            pop[i][j] = randint(0,max(num_var,num_room_type)-1)
    pop1_jiequ = pop[:,num_var+num_room_type+4:num_var+num_room_type+4+all_room_num]
    pop1_method = pop[:, num_var + num_room_type :num_var + num_room_type + 4]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            for z in range(3):
                posi = int(pop1_jiequ[i][j]+pop1_method[i][z])
                while posi>= num_var:
                    posi = posi- num_var
                pop_all[i][j*3+z] = pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_jiequ[i][z*5+int(labels[j])-1]+pop1_method[i][3])
                while posi>= num_room_type:
                    posi = posi- num_room_type
                pop_room_label[i][j] = pop[i][posi+num_var]

    return pop_all,pop,pop_room_label

def crossover_and_mutation_sort_progressive_30(pop2,num_var,num_room_type,CROSSOVER_RATE):
    all_room_num = 5 * story_num
    pop = pop2
    room_nu = np.linspace(1, 12, 12)
    new_pop = np.zeros((len(pop),len(pop[0])))
    #前1/3
    for i in range(0,int(len(pop)/3)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE/50:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1_progressive_1_30(child,x,num_var,num_room_type,MUTATION_RATE,0,all_room_num)
        new_pop[i] = child
        #1/3--2/3
        for i in range(int(len(pop)/3),int(len(pop)/3*2)):
            father = pop[i]
            child = father
            if np.random.rand() < CROSSOVER_RATE:
                mother = pop[np.random.randint(POP_SIZE)]
                cross_points1 = np.random.randint(low=0, high=len(pop[0]))
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                while cross_points2 == cross_points1:
                    cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                exchan = []
                exchan.append(cross_points2)
                exchan.append(cross_points1)
                for j in range(min(exchan), max(exchan)):
                    child[j] = mother[j]
            mutation_1_progressive_1_30(child, x, num_var, num_room_type, MUTATION_RATE,0.5,all_room_num)
            new_pop[i] = child
        # 2/3--3/3
        for i in range(int(len(pop)/3*2),len(pop)):
            father = pop[i]
            child = father
            if np.random.rand() < CROSSOVER_RATE:
                mother = pop[np.random.randint(POP_SIZE)]
                cross_points1 = np.random.randint(low=0, high=len(pop[0]))
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                while cross_points2 == cross_points1:
                    cross_points2 = np.random.randint(low=0, high=len(pop[0]))
                exchan = []
                exchan.append(cross_points2)
                exchan.append(cross_points1)
                for j in range(min(exchan), max(exchan)):
                    child[j] = mother[j]
            mutation_1_progressive_1_30(child, x, num_var, num_room_type, MUTATION_RATE,1,all_room_num)
            new_pop[i] = child
    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var, num_var+num_room_type):
            room_sort.append(new_pop[i][j])
        room_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]
        for j in range(num_var,num_var+num_room_type):
            new_pop[i][j] = room_sort[j-num_var]

    pop1_jiequ = pop[:,num_var+num_room_type+4:num_var+num_room_type+4+all_room_num]
    pop1_method = pop[:, num_var + num_room_type :num_var + num_room_type + 4]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            for z in range(3):
                posi = int(pop1_jiequ[i][j]+pop1_method[i][z])
                while posi>= num_var:
                    posi = posi- num_var
                pop_all[i][j*3+z] = pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_jiequ[i][z*5+int(labels[j])-1]+pop1_method[i][3])
                while posi>= num_room_type:
                    posi = posi- num_room_type
                pop_room_label[i][j] = pop[i][posi+num_var]

    return pop_all,new_pop,pop_room_label

def mutation_1_progressive_1_30(child, x,num_var,num_room_type,MUTATION_RATE,rate,all_room_num):
    num_var = int(num_var)
    room_nu = np.linspace(1, 12, 12)
    num_room_type = int(num_room_type)
    for mutate_point in range(num_var):
        x_var = list(map(int, x.tolist()))
        for mutate_point_1 in range(num_var):
            if child[mutate_point_1] in x_var:
                x_var.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE*rate:  # 以MUTATION_RATE的概率进行变异
            child[mutate_point] = np.random.choice(x_var)
    for mutate_point in range(num_var,num_room_type+num_var):
        x_room = list(map(int, room_nu.tolist()))
        for mutate_point_1 in range(num_var, num_room_type + num_var):
            if child[mutate_point_1] in x_room:
                x_room.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE*rate:
            child[mutate_point] = np.random.choice(x_room)
    for j in range(num_room_type + num_var, num_var+num_room_type+4+all_room_num):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,max(num_var,num_room_type)-1)

# 每层分为两组，每组5变量
def generate_DNA_story5(num_var,num_room_type,x,labels):
    all_room_num = 2*story_num*5
    room_nu =np.linspace(1, 3, 3)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+all_room_num))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        sec.sort()
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        room_ty.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = randint(1,3)
        for j in range(num_var+num_room_type,num_var+num_room_type+2*story_num*3):
            pop[i][j] = randint(0,num_var-1)
        for j in range(num_var+num_room_type+2*story_num*3,num_var+num_room_type+2*story_num*5):
            pop[i][j] = randint(0,1)
    pop1_jiequ = pop[:,num_var+num_room_type:num_var+num_room_type+2*story_num*3]
    pop1_method = pop[:, num_var+num_room_type+2*story_num*3:num_var+num_room_type+2*story_num*5]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            posi = int(pop1_jiequ[i][j])
            pop_all[i][j] = pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_method[i][z*2+int(labels[j])-1])
                if posi == 0:
                    pop_room_label[i][j] = 0
                else:
                    pop_room_label[i][j] = pop[i][num_var]

    return pop_all,pop,pop_room_label

def crossover_and_mutation_story5(pop2,num_var,num_room_type,CROSSOVER_RATE):
    pop = pop2
    room_nu = np.linspace(1, 12, 12)
    new_pop = np.zeros((len(pop),len(pop[0])))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1(child,x,num_var,num_room_type,MUTATION_RATE)
        new_pop[i] = child

    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]

    pop1_jiequ = new_pop[:,num_var+num_room_type:num_var+num_room_type+2*story_num*3]
    pop1_method = new_pop[:, num_var+num_room_type+2*story_num*3:num_var+num_room_type+2*story_num*5]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            posi = int(pop1_jiequ[i][j])
            pop_all[i][j] = new_pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_method[i][z*2+int(labels[j])-1])
                if posi == 0:
                    pop_room_label[i][j] = 0
                else:
                    pop_room_label[i][j] = new_pop[i][num_var]

    return pop_all,new_pop,pop_room_label

# 每两层分组，编码解码分离
def generate_DNA_coding_story5(num_var,num_room_type,x):
    all_room_num = 2*story_num*5
    room_nu =np.linspace(1, 3, 3)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+all_room_num))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        sec.sort()
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        room_ty.sort()
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = randint(0,3)
        for j in range(num_var+num_room_type,num_var+num_room_type+2*story_num*3):
            pop[i][j] = randint(0,num_var-1)
        for j in range(num_var+num_room_type+2*story_num*3,num_var+num_room_type+2*story_num*5):
            pop[i][j] = randint(0,1)
    return pop

def crossover_and_mutation_coding_story5(pop2,num_var,num_room_type,CROSSOVER_RATE):
    pop = pop2
    room_nu = np.linspace(1, 12, 12)
    new_pop = np.zeros((len(pop),len(pop[0])))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=len(pop[0]))
            cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=len(pop[0]))
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation_1_stort5(child,x,num_var,num_room_type,MUTATION_RATE)
        new_pop[i] = child

    for i in range(len(new_pop)):
        sec_sort = []
        room_sort = []
        for j in range(num_var):
            sec_sort.append(new_pop[i][j])
        sec_sort.sort()
        for j in range(num_var):
            new_pop[i][j] = sec_sort[j]



    return new_pop

def decoding(pop,num_var,num_room_type,labels):
    pop1_jiequ = pop[:,num_var+num_room_type:num_var+num_room_type+2*story_num*3]
    pop1_method = pop[:, num_var+num_room_type+2*story_num*3:num_var+num_room_type+2*story_num*5]
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, len(labels)))
    for i in range(POP_SIZE):
        for j in range(len(pop1_jiequ[0])):
            posi = int(pop1_jiequ[i][j])
            pop_all[i][j] = pop[i][posi]
    for i in range(POP_SIZE):
        for z in range(story_num):
            for j in range(z*modular_length_num*2,(z+1)*modular_length_num*2):
                posi = int(pop1_method[i][z*2+int(labels[j])-1])
                if posi == 0:
                    pop_room_label[i][j] = 0
                else:
                    pop_room_label[i][j] = pop[i][num_var]
    return pop_all,pop_room_label

def mutation_1_stort5(child, x,num_var,num_room_type,MUTATION_RATE):
    num_var = int(num_var)
    room_nu = np.linspace(1, 12, 12)
    num_room_type = int(num_room_type)
    for mutate_point in range(num_var):
        x_var = list(map(int, x.tolist()))
        for mutate_point_1 in range(num_var):
            if child[mutate_point_1] in x_var:
                x_var.remove(child[mutate_point_1])
        if np.random.rand() < MUTATION_RATE:  # 以MUTATION_RATE的概率进行变异
            child[mutate_point] = np.random.choice(x_var)
    for j in range(num_room_type + num_var, num_room_type + num_var + DNA_SIZE):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,num_var-1)
    for j in range(num_var, num_room_type + num_var ):
        if np.random.rand() < MUTATION_RATE:
            child[j] = randint(0,3)

# 多线程运行
def mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop_room,pop_room_label):

    sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                      cfg_file_name="Steel_section_data.ini")
    modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
    # 按房间分好节点
    modulars_of_building = modular_building.building_modulars
    modular_nums = len(labels)
    modular_infos = {}
    # 每个房间定义梁柱截面信息
    sr.run_column_room_story5(labels,pop_room_label, modular_length_num * 2 * story_num, sections_data_c1, modular_infos, pop_room)
    #
    for i in range(len(modulars_of_building)):
        modulars_of_building[i].Add_Info_And_Update_Modular(
            # modular_infos[modulars_of_building[i].modular_label - 1])
            modular_infos[i])
    modular_building.Building_Assembling(modulars_of_building)

    all_data = ms.Run_GA_sap_2(mySapObject, ModelPath, SapModel, modular_building,pop_room_label,200,modular_length_num,story_num)
    aa, bb, cc, dd, ee, ff, gg, hh, ii = all_data

    ret = SapModel.SetModelIsLocked(False)
    return aa,bb,cc,dd,ee,ff,gg,hh,ii

def SAPanalysis_GA_run2(APIPath):


    cfg = configparser.ConfigParser()
    cfg.read("Configuration.ini", encoding='utf-8')
    ProgramPath = cfg['SAP2000PATH']['dirpath']
    if not os.path.exists(APIPath):
        try:
            os.makedirs(APIPath)
        except OSError:
            pass

    AttachToInstance = False
    SpecifyPath = True

    # ModelPath = os.path.join(APIPath, 'API_1-001.sdb')
    helper = comtypes.client.CreateObject('SAP2000v1.Helper')
    helper = helper.QueryInterface(comtypes.gen.SAP2000v1.cHelper)
    if AttachToInstance:
        # attach to a running instance of SAP2000
        try:
            # get the active SapObject
            mySapObject = helper.Getobject("CSI.SAP2000.API.SapObject")
        except (OSError, comtypes.COMError):
            print("No running instance of the program found or failed to attach.")
            sys.exit(-1)
    else:
        if SpecifyPath:
            try:
                # 'create an instance of the SAPObject from the specified path
                mySapObject = helper.CreateObject(ProgramPath)
            except (OSError, comtypes.COMError):
                print("Cannot start a new instance of the program from" + ProgramPath)
                sys.exit(-1)
        else:
            try:
                # create an instance of the SapObject from the latest installed SAP2000
                mySapObject = helper.CreateObjectProgID("CSI.SAP2000.API.SapObject")
            except (OSError, comtypes.COMError):
                print("Cannot start a new instance of the program")
                sys.exit(-1)

        # start SAP2000 application
        mySapObject.ApplicationStart()

    # create SapModel object
    SapModel = mySapObject.SapModel
    # initialize model
    SapModel.InitializeNewModel()
    ModelPath = os.path.join(APIPath, 'API_1-001.sdb')
    # create new blank model
    return mySapObject,ModelPath, SapModel

def mulitrun_GA(ModelPath,mySapObject, SapModel,pop1,pop_all,pop3,q,result,weight_1,col_up,beam_up,memory_pools_all,memory_pools_fit,memory_pools_weight,memory_pools_col,memory_pools_beam):
    while True:
        if q.empty():
            break

        time = q.get()
        pop = pop1[time]
        pop_room_label = pop3[time]
        pop2= pop_all[time]
        # 判断记忆池
        for j in range(len(memory_pools_all)):
            tf = []
            for z in range(len(pop2)):
                if pop2[z] != memory_pools_all[j][z]:
                    tf.append(z)
            if len(tf) == 0:
                break
        # 记忆池中存在满足条件的个体
        if len(tf) == 0:
            result[time] = memory_pools_fit[j]
            weight_1[time] = memory_pools_weight[j]
            col_up[time] = memory_pools_col[j]
            beam_up[time] = memory_pools_beam[j]
        # 记忆池中不存在满足条件的个体
        else:
            # pop_all.append(pop)
            we, co, be, r1, r2, r3, r4, dis_all, force_all = mulit_Sap_analy_allroom(ModelPath, mySapObject, SapModel, pop,
                                                                                     pop_room_label)
            res1, res2 = Fun_1(we, co, be, dis_all, force_all, 10000)
            # num3 += 1
            weight_1[time] = res2
            col_up[time] = co
            beam_up[time] = be
            result[time] = res1
        # 记忆池更新
            memory_pools_all.append(pop2)
            memory_pools_fit.append(res1)
            memory_pools_weight.append(res2)
            memory_pools_col.append(co)
            memory_pools_beam.append(be)

def thread_sap(mySapObject_name,SapModel_name,ModelPath_name,num,pop1,pop2,pop3,result,weight_1,col_up,beam_up,memory_pools_all,memory_pools_fit,memory_pools_weight,memory_pools_col,memory_pools_beam):

    # for i in range(2):
    #     fit1, weight1,clo_up_val, beam_up_val = mulit_GA(ModelPath_name[i],mySapObject_name[i],SapModel_name[i],[pop_room],[pop_room_label])
    pop_n = [0 for i in range(len(pop2[0]))]
    if len(memory_pools_all)==0:
        memory_pools_all.append(pop_n)
        memory_pools_fit.append(100000)
        memory_pools_weight.append(100000)
        memory_pools_col.append(100000)
        memory_pools_beam.append(100000)

    q = queue.Queue()
    threads = []
    for i in range(len(pop1)):
        q.put(i)
    for i in range(num_thread):
        t = threading.Thread(target=mulitrun_GA, args=(ModelPath_name[i],mySapObject_name[i], SapModel_name[i],pop1,pop2,pop3,q,
                            result,weight_1,col_up,beam_up,memory_pools_all,memory_pools_fit,memory_pools_weight,memory_pools_col,memory_pools_beam))
        t.start()
        threads.append(t)
    for i in threads:
        i.join()
    return result,weight_1,col_up,beam_up



def select_1(pop, fitness2):  # nature selection wrt pop's fitness
    max_num = max(fitness2)+1
    fitness = []
    for i in range(len(fitness2)):
        fitness.append(max_num-fitness2[i])
    idx = np.random.choice(np.arange(POP_SIZE), size=POP_SIZE, replace=True,
                           p=np.array(fitness) / (sum(fitness)))
    pop2 = np.zeros((POP_SIZE, len(pop[0])))
    for i in range(len(pop2)):
        pop2[i] = pop[int(idx[i])]
    return pop2

def generate_DNA(num_var,num_room_type,x):
    room_nu =np.linspace(1, 12, 12)
    pop = np.zeros((POP_SIZE,num_var+num_room_type+DNA_SIZE))
    for i in range(len(pop)):
        sec = list(map(int,random.sample(x.tolist(), num_var)))
        room_ty = list(map(int,random.sample(room_nu.tolist(), num_room_type)))
        for j in range(num_var):
            pop[i][j] = sec[j]
        for j in range(num_var,num_var+num_room_type):
            pop[i][j] = room_ty[j-num_var]
        for j in range(num_var+num_room_type,num_var+num_room_type+DNA_SIZE):
            pop[i][j] = randint(0,max(num_var,num_room_type)-1)
    pop_all = np.zeros((POP_SIZE,DNA_SIZE))
    pop_room_label = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(POP_SIZE):
        for j in range(DNA_SIZE):
            posi = int(pop[i][j+num_var+num_room_type])
            if num_var>=num_room_type:
                pop_all[i][j] = pop[i][posi]
                while posi>=num_room_type:
                    posi = posi-num_room_type
                pop_room_label[i][j] = pop[i][posi+num_var]
            else:
                pop_room_label[i][j] = pop[i][posi + num_var]
                while posi>=num_var:
                    posi = posi-num_var
                pop_all[i][j] = pop[i][posi]
    return pop_all,pop,pop_room_label



def get_minfitness(num_var,num_room_type):
    wb = xlrd.open_workbook(filename=f'out_sap_run_{num_var}_{num_room_type}.xls', formatting_info=True)
    sheet1 = wb.sheet_by_index(6)
    rows = sheet1.row_values(3)[1]
    return rows

def Fun(weight,g_col,g_beam,dis_all,all_force,u):
    g_col_all = 0
    g_beam_all = 0
    Y_dis_radio_all = 0
    Y_interdis_all = 0
    Y_interdis_radio_all = 0
    floor_radio = 0
    for i in range(len(g_col)):
        if g_col[i]<= 0:
            g_col[i] = 0
        else:
            g_col[i] = g_col[i]
        g_col_all += g_col[i]
    for i in range(len(g_beam)):
        if g_beam[i]<= 0:
            g_beam[i] = 0
        else:
            g_beam[i] = g_beam[i]
        g_beam_all += g_beam[i]
    #y dis ratio
    for i in range(len(dis_all[5])):
        if dis_all[5][i] <= 1.5 and dis_all[5][i] >= -1.5:
            dis_all[5][i] = 0
        else:
            dis_all[5][i] = dis_all[5][i]
        Y_dis_radio_all += dis_all[5][i]
    # y interdis max
    for i in range(len(dis_all[7])):
        if dis_all[7][i] <= 0.004 and dis_all[7][i] >= -0.004:
            dis_all[7][i] = 0
        else:
            dis_all[7][i] = dis_all[7][i]
        Y_interdis_all += dis_all[7][i]
    # y interdis radio
    for i in range(len(dis_all[11])):
        if dis_all[11][i] <= 1.5 and dis_all[11][i] >= -1.5:
            dis_all[11][i] = 0
        else:
            dis_all[11][i] = dis_all[11][i]
        Y_interdis_radio_all += dis_all[11][i]
    # x interdis ratio
    for i in range(len(all_force[10])):
        if all_force[10][i] <= 1.5 and all_force[10][i] >= -1.5:
            all_force[10][i] = 0
        else:
            all_force[10][i] = all_force[10][i]
        floor_radio += all_force[10][i]

    G_value=u * (abs(g_col_all) + abs(g_beam_all) + abs(Y_dis_radio_all) + abs(Y_interdis_all) + abs(Y_interdis_radio_all))
    if G_value==0:
        result = weight
    else:
        result = (weight + u * (abs(g_col_all) + abs(g_beam_all) + abs(Y_dis_radio_all) + abs(Y_interdis_all) + abs(
            Y_interdis_radio_all)))*1.5
    return result,weight

#u = 1000
def Fun_1(weight,g_col,g_beam,dis_all,all_force,u):
    g_col_all = 0
    g_beam_all = 0
    Y_dis_radio_all = 0
    Y_interdis_all = 0
    Y_interdis_radio_all = 0
    floor_radio = 0
    for i in range(len(g_col)):
        if g_col[i]<= 0:
            g_col[i] = 0
        else:
            g_col[i] = g_col[i]
        g_col_all += g_col[i]
    for i in range(len(g_beam)):
        if g_beam[i]<= 0:
            g_beam[i] = 0
        else:
            g_beam[i] = g_beam[i]
        g_beam_all += g_beam[i]
    #y dis ratio
    for i in range(len(dis_all[5])):
        if dis_all[5][i] <= 1.5 and dis_all[5][i] >= -1.5:
            dis_all[5][i] = 0
        else:
            dis_all[5][i] = dis_all[5][i]
        Y_dis_radio_all += dis_all[5][i]
    # y interdis max
    for i in range(len(dis_all[7])):
        if dis_all[7][i] <= 0.004 and dis_all[7][i] >= -0.004:
            dis_all[7][i] = 0
        else:
            dis_all[7][i] = dis_all[7][i]
        Y_interdis_all += dis_all[7][i]
    # y interdis radio
    for i in range(len(dis_all[11])):
        if dis_all[11][i] <= 1.5 and dis_all[11][i] >= -1.5:
            dis_all[11][i] = 0
        else:
            dis_all[11][i] = dis_all[11][i]
        Y_interdis_radio_all += dis_all[11][i]
    # x interdis ratio
    for i in range(len(all_force[10])):
        if all_force[10][i] <= 1.5 and all_force[10][i] >= -1.5:
            all_force[10][i] = 0
        else:
            all_force[10][i] = all_force[10][i]
        floor_radio += all_force[10][i]

    G_value=u * (abs(g_col_all) + abs(g_beam_all) + abs(Y_dis_radio_all) + abs(Y_interdis_all) + abs(Y_interdis_radio_all))
    result = weight + G_value

    return result,weight
# fit = 1-10
def select_2(pop, fitness):  # nature selection wrt pop's fitness

    fit_ini = copy.deepcopy(fitness)
    luyi = copy.deepcopy(fitness)
    luyi.sort(reverse=True)
    sort_num = []
    lst = list(range(1, len(fit_ini)+1))
    list_new = []
    for i in range(len(fit_ini)):
        sort_num.append(fit_ini.index(luyi[i]))
    for i in range(len(fit_ini)):
        list_new.append(lst[sort_num[i]])
    for i in range(len(list_new)):
        list_new[i] = m.e ** (list_new[i] * 0.5)
    idx = np.random.choice(np.arange(POP_SIZE), size=POP_SIZE, replace=True,
                           p=np.array(list_new) / (sum(list_new)))
    pop2 = np.zeros((POP_SIZE, len(pop[0])))
    for i in range(len(pop2)):
        pop2[i] = pop[int(idx[i])]
    return pop2


def Get_fitness(result):
    fitness1 = []
    fitness2 = []
    for i in range(len(result)):
        if result[i]>=800:
            fitness1.append((result[i]/100000)+800)
        else :
            fitness1.append(result[i])
        if fitness1[i]>=1100:
            fitness2.append(1)
        else:
            fitness2.append(1100-fitness1[i])
    # fitness1.append(m.exp(-result[i]))
    # fitness2.append((m.exp(-result[i])+0.001))
    return fitness1, fitness2

def select(pop, fitness):  # nature selection wrt pop's fitness
    idx = np.random.choice(np.arange(POP_SIZE), size=POP_SIZE, replace=True,
                           p=(fitness) / (sum(fitness)))
    pop2 = np.zeros((POP_SIZE, len(pop[0])))
    for i in range(len(pop2)):
        pop2[i] = pop[int(idx[i])]
    return pop2

def crossover_and_mutation(pop, CROSSOVER_RATE):
    new_pop = np.zeros((POP_SIZE,DNA_SIZE))
    for i in range(len(pop)):
        father = pop[i]
        child = father
        if np.random.rand() < CROSSOVER_RATE:
            mother = pop[np.random.randint(POP_SIZE)]
            cross_points1 = np.random.randint(low=0, high=DNA_SIZE)
            cross_points2 = np.random.randint(low=0, high=DNA_SIZE)
            while cross_points2==cross_points1:
                cross_points2 = np.random.randint(low=0, high=DNA_SIZE)
            exchan = []
            exchan.append(cross_points2)
            exchan.append(cross_points1)
            for j in range(min(exchan),max(exchan)):
                child[j] = mother[j]
        mutation(child)
        new_pop[i] = child
    return new_pop

def mutation(child, MUTATION_RATE):
    for mutate_point in range(DNA_SIZE):
        if np.random.rand() < MUTATION_RATE:  # 以MUTATION_RATE的概率进行变异
            # mutate_point = np.random.randint(0, DNA_SIZE)  # 随机产生一个实数，代表要变异基因的位置
            # if mutate_point == 2:
            #     child[mutate_point] = np.random.choice(z)
            # else:
            child[mutate_point] = np.random.choice(x)

def Sap_analy_allroom(pop_room,pop_room_label):

    sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                      cfg_file_name="Steel_section_data.ini")
    modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
    # 按房间分好节点
    modulars_of_building = modular_building.building_modulars
    modular_nums = len(labels)
    modular_infos = {}
    # 每个房间定义梁柱截面信息
    sr.run_column_room_story1(labels,pop_room_label, modular_length_num * 2 * story_num, sections_data_c1, modular_infos, pop_room)
    #
    for i in range(len(modulars_of_building)):
        modulars_of_building[i].Add_Info_And_Update_Modular(
            # modular_infos[modulars_of_building[i].modular_label - 1])
            modular_infos[i])
    modular_building.Building_Assembling(modulars_of_building)

    all_data = ms.Run_GA_sap_2(mySapObject, ModelPath, SapModel, modular_building,pop_room_label,200,modular_length_num,story_num)
    aa, bb, cc, dd, ee, ff, gg, hh, ii = all_data

    ret = SapModel.SetModelIsLocked(False)
    return aa,bb,cc,dd,ee,ff,gg,hh,ii

def GA(pop1,pop3):
    result = []
    num1 = 0
    num2 = 0
    num3 = 0
    weight_1 = []
    col_up = []
    beam_up = []
    # if len(pop_all) ==0:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         pop_all.append(pop)
    #         we,co,be,r1,r2,r3,r4 =Sap_analy(pop)
    #         res1,res2 = Fun(we, co, be, 10000)
    #         num3 += 1
    #         weight_1.append(res2)
    #         result.append(res1)
    #         pop_fun_all.append(res1)
    #         pop_weight_all.append(res2)
    # else:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         for i in range(len(pop_all)):
    #             if all(pop == pop_all[i]):
    #                 num1 = 1
    #                 num2 = i
    #                 # result.append(pop_fun_all[i])
    #         if num1 == 0:
    #             pop_all.append(pop)
    #             we, co, be, r1, r2, r3, r4 = Sap_analy(pop)
    #             res1, res2 = Fun(we, co, be, 10000)
    #             num3 += 1
    #             weight_1.append(res2)
    #             pop_weight_all.append(res2)
    #             result.append(res1)
    #             pop_fun_all.append(res1)
    #         elif num1 == 1:
    #             result.append(pop_fun_all[num2])
    #             weight_1.append(pop_weight_all[num2])
    wb2_examine_ind = openpyxl.Workbook()
    # wb2_examine_ind = openpyxl.load_workbook('examine_individual.xlsx')
    wb2_pop1_indivi = wb2_examine_ind.create_sheet('pop1_indivi', index=0)
    wb2_pop3_indivi = wb2_examine_ind.create_sheet('pop3_indivi', index=1)
    loc_3 = 1
    for time in range(len(pop1)):
        pop = pop1[time]
        pop_room_label = pop3[time]
        _ = wb2_pop1_indivi.cell(row=loc_3, column=1, value=f'{pop1[time]}')
        _ = wb2_pop3_indivi.cell(row=loc_3, column=1, value=f'{pop3[time]}')
        loc_3 += 1
        wb2_examine_ind.save('examine_individual.xlsx')
        # pop_all.append(pop)
        we,co,be,r1,r2,r3,r4,dis_all,force_all =Sap_analy_allroom(pop,pop_room_label)
        res1,res2 = Fun_1(we, co, be,dis_all,force_all, 10000)
        # num3 += 1
        weight_1.append(res2)
        col_up.append(co)
        beam_up.append(be)
        result.append(res1)
        # pop_fun_all.append(res1)
        # pop_weight_all.append(res2)

    wb_clear_in1 = openpyxl.load_workbook('examine_individual.xlsx')
    ws_clear_in1 = wb_clear_in1['pop1_indivi']
    for row in ws_clear_in1:
        for cell in row:
            cell.value = None
    ws_clear_in3 = wb_clear_in1['pop3_indivi']
    for row in ws_clear_in3:
        for cell in row:
            cell.value = None
    wb2_examine_ind.save('examine_individual.xlsx')
    return result,weight_1,col_up,beam_up

def GA_examine(ModelPath,mySapObject, SapModel,pop1,pop3):
    result = []
    num1 = 0
    num2 = 0
    num3 = 0
    weight_1 = []
    col_up = []
    beam_up = []
    # if len(pop_all) ==0:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         pop_all.append(pop)
    #         we,co,be,r1,r2,r3,r4 =Sap_analy(pop)
    #         res1,res2 = Fun(we, co, be, 10000)
    #         num3 += 1
    #         weight_1.append(res2)
    #         result.append(res1)
    #         pop_fun_all.append(res1)
    #         pop_weight_all.append(res2)
    # else:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         for i in range(len(pop_all)):
    #             if all(pop == pop_all[i]):
    #                 num1 = 1
    #                 num2 = i
    #                 # result.append(pop_fun_all[i])
    #         if num1 == 0:
    #             pop_all.append(pop)
    #             we, co, be, r1, r2, r3, r4 = Sap_analy(pop)
    #             res1, res2 = Fun(we, co, be, 10000)
    #             num3 += 1
    #             weight_1.append(res2)
    #             pop_weight_all.append(res2)
    #             result.append(res1)
    #             pop_fun_all.append(res1)
    #         elif num1 == 1:
    #             result.append(pop_fun_all[num2])
    #             weight_1.append(pop_weight_all[num2])
    wb2_examine_ind = openpyxl.Workbook()
    # wb2_examine_ind = openpyxl.load_workbook('examine_individual.xlsx')
    wb2_pop1_indivi = wb2_examine_ind.create_sheet('pop1_indivi', index=0)
    wb2_pop3_indivi = wb2_examine_ind.create_sheet('pop3_indivi', index=1)
    loc_3 = 1
    for time in range(len(pop1)):
        pop = pop1[time]
        pop_room_label = pop3[time]
        _ = wb2_pop1_indivi.cell(row=loc_3, column=1, value=f'{pop1[time]}')
        _ = wb2_pop3_indivi.cell(row=loc_3, column=1, value=f'{pop3[time]}')
        loc_3 += 1
        wb2_examine_ind.save('examine_individual.xlsx')
        # pop_all.append(pop)
        we,co,be,r1,r2,r3,r4,dis_all,force_all =mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop,pop_room_label)
        res1,res2 = Fun_1(we, co, be,dis_all,force_all, 10000)
        # num3 += 1
        weight_1.append(res2)
        col_up.append(co)
        beam_up.append(be)
        result.append(res1)
        # pop_fun_all.append(res1)
        # pop_weight_all.append(res2)

    wb_clear_in1 = openpyxl.load_workbook('examine_individual.xlsx')
    ws_clear_in1 = wb_clear_in1['pop1_indivi']
    for row in ws_clear_in1:
        for cell in row:
            cell.value = None
    ws_clear_in3 = wb_clear_in1['pop3_indivi']
    for row in ws_clear_in3:
        for cell in row:
            cell.value = None
    wb2_examine_ind.save('examine_individual.xlsx')
    return result,weight_1,col_up,beam_up


def GA_1(pop1,pop3):
    result = []
    num1 = 0
    num2 = 0
    num3 = 0
    weight_1 = []
    col_up = []
    beam_up = []
    # if len(pop_all) ==0:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         pop_all.append(pop)
    #         we,co,be,r1,r2,r3,r4 =Sap_analy(pop)
    #         res1,res2 = Fun(we, co, be, 10000)
    #         num3 += 1
    #         weight_1.append(res2)
    #         result.append(res1)
    #         pop_fun_all.append(res1)
    #         pop_weight_all.append(res2)
    # else:
    #     for time in range(len(pop1)):
    #         pop = pop1[time]
    #         for i in range(len(pop_all)):
    #             if all(pop == pop_all[i]):
    #                 num1 = 1
    #                 num2 = i
    #                 # result.append(pop_fun_all[i])
    #         if num1 == 0:
    #             pop_all.append(pop)
    #             we, co, be, r1, r2, r3, r4 = Sap_analy(pop)
    #             res1, res2 = Fun(we, co, be, 10000)
    #             num3 += 1
    #             weight_1.append(res2)
    #             pop_weight_all.append(res2)
    #             result.append(res1)
    #             pop_fun_all.append(res1)
    #         elif num1 == 1:
    #             result.append(pop_fun_all[num2])
    #             weight_1.append(pop_weight_all[num2])
    # wb2_examine_ind = openpyxl.Workbook()
    # # wb2_examine_ind = openpyxl.load_workbook('examine_individual.xlsx')
    # wb2_pop1_indivi = wb2_examine_ind.create_sheet('pop1_indivi', index=0)
    # wb2_pop3_indivi = wb2_examine_ind.create_sheet('pop3_indivi', index=1)
    loc_3 = 1
    for time in range(len(pop1)):
        pop = pop1[time]
        pop_room_label = pop3[time]
        # _ = wb2_pop1_indivi.cell(row=loc_3, column=1, value=f'{pop1[time]}')
        # _ = wb2_pop3_indivi.cell(row=loc_3, column=1, value=f'{pop3[time]}')
        # loc_3 += 1
        # wb2_examine_ind.save('examine_individual.xlsx')
        # pop_all.append(pop)
        we,co,be,r1,r2,r3,r4,dis_all,force_all =Sap_analy_allroom(pop,pop_room_label)
        res1,res2 = Fun(we, co, be,dis_all,force_all, 10)
        # num3 += 1
        weight_1.append(res2)
        col_up.append(co)
        beam_up.append(be)
        result.append(res1)
        # pop_fun_all.append(res1)
        # pop_weight_all.append(res2)

    wb_clear_in1 = openpyxl.load_workbook('examine_individual.xlsx')
    ws_clear_in1 = wb_clear_in1['pop1_indivi']
    for row in ws_clear_in1:
        for cell in row:
            cell.value = None
    ws_clear_in3 = wb_clear_in1['pop3_indivi']
    for row in ws_clear_in3:
        for cell in row:
            cell.value = None
    # wb2_examine_ind.save('examine_individual.xlsx')
    return result,weight_1,col_up,beam_up

def Run_GA_allstory(POP_SIZE_1,DNA_SIZE_1,CROSSOVER_RATE_1,MUTATION_RATE_1,N_GENERATIONS_1,xx1,mySapObject):

    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop1 = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(len(pop1)):
        for j in range(len(pop1[0])):
            pop1[i][j] = np.random.choice(x)
            # pop1[i][j] = 15
    for _ in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1)
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness1, fitness2 = Get_fitness(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(max(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = max(fitness2)
        mm2 = pop1[mm]
        max_ru.append(1100 - max(fitness2))
        pop1 = select(pop1, fitness2)
        pop1 = np.array(crossover_and_mutation(pop1, CROSSOVER_RATE))
        aaa = []
        aaa.append(pop1[0])
        pop200 = pop_all
        if max1 >= Get_fitness(GA(aaa)[0])[1][0]:
            sap_run_time += 1
            pop1[0] = mm2

    print(f"最小值位置为", pop1[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0],max_ru[len(max_ru) - 1],max_ru,pop_all,pop_zhongqun_all,weight_min,pop_all_fitness,pop_all_weight,
                 col_up_all,beam_up_all]
    return all_infor


def Run_GA_allstory_1(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject, num_var):
    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    mm1 = xx1
    x = []
    for i in range(len(num_var)):
        x.append(np.random.choice(mm1))
        
    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop1 = np.zeros((POP_SIZE, DNA_SIZE))
    for i in range(len(pop1)):
        for j in range(len(pop1[0])):
            pop1[i][j] = np.random.choice(x)
            # pop1[i][j] = 15
    for _ in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1)
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness1, fitness2 = Get_fitness(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(max(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = max(fitness2)
        mm2 = pop1[mm]
        max_ru.append(1100 - max(fitness2))
        pop1 = select(pop1, fitness2)
        pop1 = np.array(crossover_and_mutation(pop1, CROSSOVER_RATE))
        aaa = []
        aaa.append(pop1[0])
        pop200 = pop_all
        if max1 >= Get_fitness(GA(aaa)[0])[1][0]:
            sap_run_time += 1
            pop1[0] = mm2

    print(f"最小值位置为", pop1[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor


def Run_GA_allstory_2(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):



    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    pop1, pop2,pop3 = generate_DNA(num_var,num_room_type,x)

    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1


    for _ in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值
        pop2 = select_1(pop2, fitness2)
        pop1,pop2,pop3 = crossover_and_mutation_1(pop2,num_var,num_room_type, CROSSOVER_RATE)

        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值位置为", pop1[0])
    print(f"最小值位置为2", pop2[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor

def Run_GA_allstory_3(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):



    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    # pop1 section_coda pop2 all pop3 label code
    pop1, pop2,pop3 = generate_DNA(num_var,num_room_type,x)

    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1

    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        if len(memory_pools_all) == 0:
            memory_pools_all.extend(pop2)
            fit, weight,clo_val, beam_val = GA(pop1,pop3)
            memory_pools_fit.extend(fit)
            memory_pools_weight.extend(weight)
            memory_pools_col.extend(clo_val)
            memory_pools_beam.extend(beam_val)
        else:
            fit = []
            weight = []
            num_cho = []
            clo_val = []
            beam_val = []
            for i in range(len(pop2)):
                for j in range(len(memory_pools_all)):
                    panduan = 0
                    tf = []
                    for z in range(len(pop2[i])):
                        if pop2[i][z] != memory_pools_all[j][z]:
                            tf.append(z)
                    if len(tf) == 0:
                        break
                if len(tf) == 0:
                    fit.append(memory_pools_fit[j])
                    weight.append(memory_pools_weight[j])
                    clo_val.append(memory_pools_col[j])
                    beam_val.append(memory_pools_beam[j])
                else:
                    num_cho.append(i)
                    fit1, weight1,clo_up_val, beam_up_val = GA([pop1[i]],[pop3[i]])
                    fit.append(fit1[0])
                    weight.append(weight1[0])
                    clo_val.append(clo_up_val[0])
                    beam_val.append(beam_up_val[0])
            for i in range(len(num_cho)):
                memory_pools_all.append(pop2[num_cho[i]])
                memory_pools_fit.append(fit[num_cho[i]])
                memory_pools_weight.append(weight[num_cho[i]])
                memory_pools_col.append(clo_val[num_cho[i]])
                memory_pools_beam.append(beam_val[num_cho[i]])
        # result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        weight_pop = weight
        result1 =fit
        clo_up_1 = clo_val
        beam_up_1 = beam_val
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值
        pop2 = select_2(pop2, fitness2)
        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop1_new, pop2_new, pop3_new = generate_DNA(num_var, num_room_type, x)
            exchange_num = int(0.5*len(pop1_new))
            for ex_num in range(exchange_num):
                pop1[len(pop1)-1-ex_num] = pop1_new[ex_num]
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]
                pop3[len(pop1) - 1 - ex_num] = pop3_new[ex_num]

        pop1,pop2,pop3 = crossover_and_mutation_1(pop2,num_var,num_room_type, CROSSOVER_RATE)

        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA(aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值位置为", pop1[0])
    print(f"最小值位置为2", pop2[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor

def Run_GA_allstory_trad(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):



    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    # pop1 section_coda pop2 all pop3 label code
    pop1, pop2,pop3 = generate_DNA(num_var,num_room_type,x)

    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1
    for _ in range(N_GENERATIONS):
        result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1, pop3)
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值
        pop2 = select_2(pop2, fitness2)
        # 引入新个体

        pop1,pop2,pop3 = crossover_and_mutation_1(pop2,num_var,num_room_type, CROSSOVER_RATE)

        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA(aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值位置为", pop1[0])
    print(f"最小值位置为2", pop2[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor
# 排序,分段变异，染色体长度108+
def Run_GA_allstory_3_sort(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):

    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    # pop1 section_coda pop2 all pop3 label code
    pop1, pop2,pop3 = generate_DNA_sort_ku(num_var,num_room_type,x)
    pop8 = copy.deepcopy(pop2)


    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1

    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        if len(memory_pools_all) == 0:
            memory_pools_all.extend(pop2)
            fit, weight,clo_val, beam_val = GA(pop1,pop3)
            memory_pools_fit.extend(fit)
            memory_pools_weight.extend(weight)
            memory_pools_col.extend(clo_val)
            memory_pools_beam.extend(beam_val)
        else:
            fit = []
            weight = []
            num_cho = []
            clo_val = []
            beam_val = []
            for i in range(len(pop2)):
                for j in range(len(memory_pools_all)):
                    panduan = 0
                    tf = []
                    for z in range(len(pop2[i])):
                        if pop2[i][z] != memory_pools_all[j][z]:
                            tf.append(z)
                    if len(tf) == 0:
                        break
                if len(tf) == 0:
                    fit.append(memory_pools_fit[j])
                    weight.append(memory_pools_weight[j])
                    clo_val.append(memory_pools_col[j])
                    beam_val.append(memory_pools_beam[j])
                else:
                    num_cho.append(i)
                    fit1, weight1,clo_up_val, beam_up_val = GA([pop1[i]],[pop3[i]])
                    fit.append(fit1[0])
                    weight.append(weight1[0])
                    clo_val.append(clo_up_val[0])
                    beam_val.append(beam_up_val[0])
            for i in range(len(num_cho)):
                memory_pools_all.append(pop2[num_cho[i]])
                memory_pools_fit.append(fit[num_cho[i]])
                memory_pools_weight.append(weight[num_cho[i]])
                memory_pools_col.append(clo_val[num_cho[i]])
                memory_pools_beam.append(beam_val[num_cho[i]])
        # result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        weight_pop = weight
        result1 =fit
        clo_up_1 = clo_val
        beam_up_1 = beam_val
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值

        pop2 = select_2(pop2, fitness2)
        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop1_new, pop2_new, pop3_new = generate_DNA(num_var, num_room_type, x)
            exchange_num = int(0.5*len(pop1_new))
            for ex_num in range(exchange_num):
                pop1[len(pop1)-1-ex_num] = pop1_new[ex_num]
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]
                pop3[len(pop1) - 1 - ex_num] = pop3_new[ex_num]

        pop1,pop2,pop3 = crossover_and_mutation_sort(pop2,num_var,num_room_type, CROSSOVER_RATE)

        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA(aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值位置为", pop1[0])
    print(f"最小值位置为2", pop2[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor
# 排序,分段变异，染色体长度45+
def Run_GA_allstory_3_sort_30(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):

    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    # pop1 section_coda pop2 all pop3 label code
    pop1, pop2,pop3 = generate_DNA_sort_ku_30(num_var,num_room_type,x,labels)
    pop8 = copy.deepcopy(pop2)


    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1

    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        if len(memory_pools_all) == 0:
            memory_pools_all.extend(pop2)
            fit, weight,clo_val, beam_val = GA(pop1,pop3)
            memory_pools_fit.extend(fit)
            memory_pools_weight.extend(weight)
            memory_pools_col.extend(clo_val)
            memory_pools_beam.extend(beam_val)
        else:
            fit = []
            weight = []
            num_cho = []
            clo_val = []
            beam_val = []
            for i in range(len(pop2)):
                for j in range(len(memory_pools_all)):
                    panduan = 0
                    tf = []
                    for z in range(len(pop2[i])):
                        if pop2[i][z] != memory_pools_all[j][z]:
                            tf.append(z)
                    if len(tf) == 0:
                        break
                if len(tf) == 0:
                    fit.append(memory_pools_fit[j])
                    weight.append(memory_pools_weight[j])
                    clo_val.append(memory_pools_col[j])
                    beam_val.append(memory_pools_beam[j])
                else:
                    num_cho.append(i)
                    fit1, weight1,clo_up_val, beam_up_val = GA([pop1[i]],[pop3[i]])
                    fit.append(fit1[0])
                    weight.append(weight1[0])
                    clo_val.append(clo_up_val[0])
                    beam_val.append(beam_up_val[0])
            for i in range(len(num_cho)):
                memory_pools_all.append(pop2[num_cho[i]])
                memory_pools_fit.append(fit[num_cho[i]])
                memory_pools_weight.append(weight[num_cho[i]])
                memory_pools_col.append(clo_val[num_cho[i]])
                memory_pools_beam.append(beam_val[num_cho[i]])
        # result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        weight_pop = weight
        result1 =fit
        clo_up_1 = clo_val
        beam_up_1 = beam_val
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值

        pop2 = select_2(pop2, fitness2)
        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop1_new, pop2_new, pop3_new = generate_DNA_sort_ku_30(num_var, num_room_type, x,labels)
            exchange_num = int(0.5*len(pop1_new))
            for ex_num in range(exchange_num):
                pop1[len(pop1)-1-ex_num] = pop1_new[ex_num]
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]
                pop3[len(pop1) - 1 - ex_num] = pop3_new[ex_num]

        pop1,pop2,pop3 = crossover_and_mutation_sort_progressive_30(pop2,num_var,num_room_type, CROSSOVER_RATE)

        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA(aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值位置为", pop1[0])
    print(f"最小值位置为2", pop2[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor
# 每层分两组
def Run_GA_allstory2(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):

    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    # pop1 section_coda pop2 all pop3 label code
    pop1, pop2,pop3 = generate_DNA_story5(num_var, num_room_type, x, labels)
    pop8 = copy.deepcopy(pop2)


    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1

    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        if len(memory_pools_all) == 0:
            memory_pools_all.extend(pop2)
            fit, weight,clo_val, beam_val = GA(pop1,pop3)
            memory_pools_fit.extend(fit)
            memory_pools_weight.extend(weight)
            memory_pools_col.extend(clo_val)
            memory_pools_beam.extend(beam_val)
        else:
            fit = []
            weight = []
            num_cho = []
            clo_val = []
            beam_val = []
            for i in range(len(pop2)):
                for j in range(len(memory_pools_all)):
                    panduan = 0
                    tf = []
                    for z in range(len(pop2[i])):
                        if pop2[i][z] != memory_pools_all[j][z]:
                            tf.append(z)
                    if len(tf) == 0:
                        break
                if len(tf) == 0:
                    fit.append(memory_pools_fit[j])
                    weight.append(memory_pools_weight[j])
                    clo_val.append(memory_pools_col[j])
                    beam_val.append(memory_pools_beam[j])
                else:
                    num_cho.append(i)
                    fit1, weight1,clo_up_val, beam_up_val = GA([pop1[i]],[pop3[i]])
                    fit.append(fit1[0])
                    weight.append(weight1[0])
                    clo_val.append(clo_up_val[0])
                    beam_val.append(beam_up_val[0])
            for i in range(len(num_cho)):
                memory_pools_all.append(pop2[num_cho[i]])
                memory_pools_fit.append(fit[num_cho[i]])
                memory_pools_weight.append(weight[num_cho[i]])
                memory_pools_col.append(clo_val[num_cho[i]])
                memory_pools_beam.append(beam_val[num_cho[i]])
        # result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        weight_pop = weight
        result1 =fit
        clo_up_1 = clo_val
        beam_up_1 = beam_val
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值

        pop2 = select_2(pop2, fitness2)

        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop1_new, pop2_new, pop3_new = generate_DNA_story5(num_var, num_room_type, x, labels)
            exchange_num = int(0.5*len(pop1_new))
            for ex_num in range(exchange_num):
                pop1[len(pop1)-1-ex_num] = pop1_new[ex_num]
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]
                pop3[len(pop1) - 1 - ex_num] = pop3_new[ex_num]

        pop1,pop2,pop3 = crossover_and_mutation_story5(pop2, num_var, num_room_type, CROSSOVER_RATE)

        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA(aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值截面编码", pop1[0])
    print(f"最小值染色体", pop2[0])
    print(f"最小值房间编码", pop3[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,
                 col_up_all, beam_up_all]
    return all_infor
#每层分两组，编码解码分离
def Run_GA_allstory2_divided(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1, mySapObject,
                      num_var,num_room_type):

    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []
    # pop1 section_coda pop2 all pop3 label code
    pop2= generate_DNA_coding_story5(num_var, num_room_type, x)
    pop_decoe_1 = copy.deepcopy(pop2)
    pop1,pop3 = decoding(pop_decoe_1,num_var,num_room_type,labels)
    pop8 = copy.deepcopy(pop2)


    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1

    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)
        if len(memory_pools_all) == 0:
            memory_pools_all.extend(pop2)
            fit, weight,clo_val, beam_val = GA(pop1,pop3)
            memory_pools_fit.extend(fit)
            memory_pools_weight.extend(weight)
            memory_pools_col.extend(clo_val)
            memory_pools_beam.extend(beam_val)
        else:
            fit = []
            weight = []
            num_cho = []
            clo_val = []
            beam_val = []
            for i in range(len(pop2)):
                for j in range(len(memory_pools_all)):
                    panduan = 0
                    tf = []
                    for z in range(len(pop2[i])):
                        if pop2[i][z] != memory_pools_all[j][z]:
                            tf.append(z)
                    if len(tf) == 0:
                        break
                if len(tf) == 0:
                    fit.append(memory_pools_fit[j])
                    weight.append(memory_pools_weight[j])
                    clo_val.append(memory_pools_col[j])
                    beam_val.append(memory_pools_beam[j])
                else:
                    num_cho.append(i)
                    fit1, weight1,clo_up_val, beam_up_val = GA([pop1[i]],[pop3[i]])
                    fit.append(fit1[0])
                    weight.append(weight1[0])
                    clo_val.append(clo_up_val[0])
                    beam_val.append(beam_up_val[0])
            for i in range(len(num_cho)):
                memory_pools_all.append(pop2[num_cho[i]])
                memory_pools_fit.append(fit[num_cho[i]])
                memory_pools_weight.append(weight[num_cho[i]])
                memory_pools_col.append(clo_val[num_cho[i]])
                memory_pools_beam.append(beam_val[num_cho[i]])
        # result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        weight_pop = weight
        result1 =fit
        clo_up_1 = clo_val
        beam_up_1 = beam_val
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值

        pop2 = select_2(pop2, fitness2)


        # 交叉变异
        pop2 = crossover_and_mutation_coding_story5(pop2, num_var, num_room_type, CROSSOVER_RATE)

        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop2_new = generate_DNA_coding_story5(num_var, num_room_type, x)
            exchange_num = int(0.3*len(pop2_new))
            for ex_num in range(exchange_num):
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]

        if run_time %30==0:
            print(run_time)
        pop1, pop3 = decoding(pop2, num_var, num_room_type, labels)
        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA(aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值截面编码", pop1[0])
    print(f"最小值染色体", pop2[0])
    print(f"最小值房间编码", pop3[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    ret = mySapObject.ApplicationExit(False)
    SapModel = None
    mySapObject = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,col_up_all, beam_up_all,pop2[0],pop3[0],pop_zhongqun_all_3]
    return all_infor

#每层分两组，编码解码分离、多线程运行
def Run_GA_allstory2_divided_mulit(POP_SIZE_1, DNA_SIZE_1, CROSSOVER_RATE_1, MUTATION_RATE_1, N_GENERATIONS_1, xx1,
                      num_var,num_room_type,num_thread,mySapObject_name,SapModel_name,ModelPath_name):

    # run GA
    start = time.perf_counter()
    POP_SIZE = POP_SIZE_1
    DNA_SIZE = DNA_SIZE_1
    x = xx1

    CROSSOVER_RATE = CROSSOVER_RATE_1
    MUTATION_RATE = MUTATION_RATE_1
    N_GENERATIONS = N_GENERATIONS_1
    max_ru = []  # 记录每代fitness 最值
    pop_all = []  # 记录所有计算过的种群（不重复）
    pop_fun_all = []  # 记录所有计算过的种群fitness（不重复）
    pop_zhongqun_all = []  # 记录每代种群（不重复）
    pop_weight_all = []  # 记录所有计算过的种群重量（不重复）
    sap_run_time = 0  # 记录sap运行次数
    weight_min = []  # 记录每代最小重量
    pop_all_fitness = []  # 记录每代种群fitness
    pop_all_weight = []  # 记录每代种群重量
    col_up_all = []  # 记录每代每层柱最大约束
    beam_up_all = []  # 记录每代每层梁最大约束
    pop_zhongqun_all_2 = []#记录种群所有
    pop_zhongqun_all_3 = []
    # pop1 section_coda pop2 all pop3 label code
    pop2= generate_DNA_coding_story5(num_var, num_room_type, x)
    pop_decoe_1 = copy.deepcopy(pop2)
    pop1,pop3 = decoding(pop_decoe_1,num_var,num_room_type,labels)
    pop8 = copy.deepcopy(pop2)


    wb2_examine = openpyxl.Workbook()
    # wb2_examine = openpyxl.load_workbook('examine_run.xlsx')
    wb2_pop1_all = wb2_examine.create_sheet('pop1_all', index=0)
    wb2_pop3_all = wb2_examine.create_sheet('pop3_all', index=1)

    loc_2 = 1
    for i in range(len(pop1)):
        _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
        _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
        loc_2 += 1
    loc_2 += 1

    memory_pools_all = []
    memory_pools_fit = []
    memory_pools_weight = []
    memory_pools_col = []
    memory_pools_beam = []
    for run_time in range(N_GENERATIONS):
        pop_zhongqun_all.append(pop1)
        pop_zhongqun_all_2.append(pop2)
        pop_zhongqun_all_3.append(pop3)
        # 计算fitness等参数
        fit = [0 for i in range(len(pop2))]
        weight = [0 for i in range(len(pop2))]
        clo_val = [0 for i in range(len(pop2))]
        beam_val = [0 for i in range(len(pop2))]
        result1,weight_pop,clo_up_1,beam_up_1=thread_sap(mySapObject_name,SapModel_name,ModelPath_name,num_thread, pop1, pop2, pop3, fit, weight, clo_val, beam_val, memory_pools_all, memory_pools_fit,
                   memory_pools_weight, memory_pools_col, memory_pools_beam)

        # result1, weight_pop, clo_up_1, beam_up_1 = GA(pop1,pop3)
        col_up_all.append(clo_up_1)
        beam_up_all.append(beam_up_1)
        pop_all_weight.append(weight_pop)
        # sap_run_time += run_time
        fitness2 = Get_fitness_1(result1)
        pop_all_fitness.append(fitness2)
        mm = fitness2.index(min(fitness2))
        weight_min.append(weight_pop[mm])
        max1 = min(fitness2)

        mm2 = pop1[mm]# 最小值对应pop1编码
        mm2_all = pop2[mm]# 最小值对应pop2编码
        mm2_all3 = pop3[mm]  # 最小值对应pop2编码
        max_ru.append(min(fitness2))# 统计历代最小值

        pop2 = select_2(pop2, fitness2)


        # 交叉变异
        pop2 = crossover_and_mutation_coding_story5(pop2, num_var, num_room_type, CROSSOVER_RATE)

        # 引入新个体
        run_time +=1
        if run_time % 5 == 0:
            pop2_new = generate_DNA_coding_story5(num_var, num_room_type, x)
            exchange_num = int(0.3*len(pop2_new))
            for ex_num in range(exchange_num):
                pop2[len(pop1) - 1 - ex_num] = pop2_new[ex_num]

        if run_time %5==0:
            print(run_time)
        pop1, pop3 = decoding(pop2, num_var, num_room_type, labels)
        for i in range(len(pop1)):
            _ = wb2_pop1_all.cell(row=loc_2, column=1, value=f'{pop1[i]}')
            _ = wb2_pop3_all.cell(row=loc_2, column=1, value=f'{pop3[i]}')
            loc_2 += 1
        loc_2 += 1
        wb2_examine.save('examine_run.xlsx')
        aaa = []
        aaa.append(pop1[0])
        pop3_ga = []
        pop3_ga.append(pop3[0])
        pop200 = pop_all
        # if max1 <= m.log(GA(aaa,pop3_ga)[0][0]):
        if max1 <=GA_examine(ModelPath_name[0],mySapObject_name[0], SapModel_name[0],aaa, pop3_ga)[0][0]:
            sap_run_time += 1
            pop1[0] = mm2
            pop2[0] = mm2_all
            pop3[0] = mm2_all3

    wb_clear_1 = openpyxl.load_workbook('examine_run.xlsx')
    ws_clear_1 = wb_clear_1['pop1_all']
    for row in ws_clear_1:
        for cell in row:
            cell.value = None
    ws_clear_3 = wb_clear_1['pop3_all']
    for row in ws_clear_3:
        for cell in row:
            cell.value = None
    wb2_examine.save('examine_run.xlsx')

    print(f"最小值截面编码", pop1[0])
    print(f"最小值染色体", pop2[0])
    print(f"最小值房间编码", pop3[0])
    print(f"最小值为", max_ru[len(max_ru) - 1])
    print(f"最小重量为",weight_min[len(max_ru) - 1])
    end = time.perf_counter()
    runTime = end - start
    print("运行时间：", runTime)
    for i in range(len(mySapObject_name)):
        ret = mySapObject_name[i].ApplicationExit(False)
        SapModel_name[i] = None
        mySapObject_name[i] = None

    all_infor = [pop1[0], max_ru[len(max_ru) - 1], max_ru, pop_all, pop_zhongqun_all,pop_zhongqun_all_2, weight_min, pop_all_fitness,
                 pop_all_weight,col_up_all, beam_up_all,pop2[0],pop3[0],pop_zhongqun_all_3]
    return all_infor

def Fun_1(weight,g_col,g_beam,dis_all,all_force,u):
    g_col_all = 0
    g_beam_all = 0
    Y_dis_radio_all = 0
    Y_interdis_all = 0
    Y_interdis_radio_all = 0
    floor_radio = 0
    g_col1 = copy.deepcopy(g_col)
    g_beam1 = copy.deepcopy(g_beam)
    dis_all5 = copy.deepcopy(dis_all[5])
    dis_all7 = copy.deepcopy(dis_all[7])
    for i in range(len(g_col1)):
        if g_col1[i]<= 0:
            g_col1[i] = 0
        else:
            g_col1[i] = g_col1[i]
        g_col_all += g_col1[i]
    for i in range(len(g_beam1)):
        if g_beam1[i]<= 0:
            g_beam1[i] = 0
        else:
            g_beam1[i] = g_beam1[i]
        g_beam_all += g_beam1[i]
    #y dis ratio
    for i in range(len(dis_all5)):
        if dis_all5[i] <= 0.00167 and dis_all5[i] >= -0.00167:
            dis_all5[i] = 0
        else:
            dis_all5[i] = dis_all5[i]
        # Y_dis_radio_all += dis_all5[i]
    Y_dis_radio_all = max(dis_all5)
    Y_dis_radio_all = Y_dis_radio_all*100
    # y interdis max
    for i in range(len(dis_all7)):
        if dis_all7[i] <= 0.004 and dis_all7[i] >= -0.004:
            dis_all7[i] = 0
        else:
            dis_all7[i] = dis_all7[i]
        # Y_interdis_all += dis_all7[i]
    Y_interdis_all = max(dis_all7)
    Y_interdis_all = Y_interdis_all*100
    # # y interdis radio
    # for i in range(len(dis_all[11])):
    #     if dis_all[11][i] <= 1.5 and dis_all[11][i] >= -1.5:
    #         dis_all[11][i] = 0
    #     else:
    #         dis_all[11][i] = dis_all[11][i]
    #     Y_interdis_radio_all += dis_all[11][i]
    # # x interdis ratio
    # for i in range(len(all_force[10])):
    #     if all_force[10][i] <= 1.5 and all_force[10][i] >= -1.5:
    #         all_force[10][i] = 0
    #     else:
    #         all_force[10][i] = all_force[10][i]
    #     floor_radio += all_force[10][i]
    g_col_max= max(g_col)
    g_beam_max = max(g_beam)
    dis_all_max = max(dis_all[5])
    interdis_max = max(dis_all[7])
    g_all_max = max(g_col_max,g_beam_max)
    G_value=u * (abs(g_col_all) + abs(g_beam_all) + abs(Y_dis_radio_all) + abs(Y_interdis_all) + abs(Y_interdis_radio_all))
    gx = [g_col_max,g_beam_max,abs(dis_all_max),abs(interdis_max)]
    # gx_Normalization = [g_col_all,g_beam_all,Y_dis_radio_all,Y_interdis_all]
    result = weight + G_value

    gx_demo = copy.deepcopy(gx)
    if gx_demo[0]>=5:
        gx_demo[0]=1
    elif gx_demo[0]<=-1:
        gx_demo[0] = -1
    elif gx_demo[0]<=5 and gx_demo[0]>=-1:
        gx_demo[0]=(gx_demo[0]+1)/6
    if gx_demo[1]>=2:
        gx_demo[1]=1
    elif gx_demo[1]<=-1:
        gx_demo[1] = -1
    elif gx_demo[1]<=2 and gx_demo[1]>=-1:
        gx_demo[1]=(gx_demo[1]+1)/3
    if gx_demo[2] >= 0.05:
        gx_demo[2] = 0.05
    else:
        gx_demo[2] = gx_demo[2] / 0.05
    if gx_demo[3] >= 0.05:
        gx_demo[3] = 0.05
    else:
        gx_demo[3] = gx_demo[3] / 0.05
    return result,weight,gx,gx_demo

def mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop_room,pop_room_label):
    # 建立房间信息
    sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                      cfg_file_name="Steel_section_data_I_cube.ini")
    modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
    # 按房间分好节点
    modulars_of_building = modular_building.building_modulars

    modular_infos = {}
    # 每个房间定义梁柱截面信息
    sr.run_column_room_modular_section(labels,pop_room_label, modular_length_num * 2 * story_num, sections_data_c1, modular_infos, pop_room,story_num,zone_num)
    #
    for i in range(len(modulars_of_building)):
        modulars_of_building[i].Add_Info_And_Update_Modular(
            # modular_infos[modulars_of_building[i].modular_label - 1])
            modular_infos[i])
    modular_building.Building_Assembling(modulars_of_building)

    all_data = ms.Run_GA_sap_2(mySapObject, ModelPath, SapModel, modular_building,pop_room_label,200,modular_length_num,story_num)
    aa, bb, cc, dd, ee, ff, gg, hh, ii = all_data

    ret = SapModel.SetModelIsLocked(False)
    return aa,bb,cc,dd,ee,ff,gg,hh,ii


'''model data'''
# modular size
modular_length = 8000
modular_width = [4000,4000,5400,3600,3600,4400,4400,4000]
modular_heigth = 3000
modular_length_num = 8
modular_dis = 400
corridor_width = 4000

story_num = 6
story_zone = 4#每组模块的分区数量
story_group = 3#每组模块的楼层数
modular_num = 6#整个建筑的模块种类

zone_num = int(story_num / story_group * story_zone)
section_num = 3 * modular_num
brace_num = modular_num
group_num = int(story_num / story_group)
modular_all = modular_length_num * 2 *story_num


# steel section information
sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                  cfg_file_name="Steel_section_data_I_cube.ini")

# generate model
model_data = dj.generate_model_data(modular_length,modular_width,modular_heigth,modular_length_num,modular_dis,story_num,corridor_width)
nodes = model_data[0]
edges_all = model_data[1]
labels = model_data[2]
cor_edges = model_data[3]
joint_hor = model_data[4]
joint_ver = model_data[5]
room_indx = model_data[6]

'''model visvalization'''
# dj.model_vis(model_data)


'''model to sap2000'''
#获得模型信息

# modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
#生成模型
# modulars_of_building = modular_building.building_modulars
# modular_nums = len(labels)
# modular_infos = {}
# APIPath = os.path.join(os.getcwd(), 'cases')
# mySapObject, ModelPath, SapModel = ms.SAPanalysis_GA_run(APIPath)
# for i in range(modular_nums):
#     modular_infos[i] = ms.Modular_Info_Initialization(type='regular', top_edge=sections_data_c1[0],
#                                                       bottom_edge=sections_data_c1[8],
#                                                       column_edge=sections_data_c1[17])
# for i in range(len(modulars_of_building)):
#     modulars_of_building[i].Add_Info_And_Update_Modular(
#         modular_infos[modulars_of_building[i].modular_label - 1])
# modular_building.Building_Assembling(modulars_of_building)
# # run sap2000
# all_data = ms.Run_GA_sap(mySapObject, ModelPath, SapModel, modular_building,200,modular_length_num,story_num)

sections_data_c1, type_keys_c1, sections_c1 = ms.get_section_info(section_type='c0',
                                                                  cfg_file_name="Steel_section_data_I_cube.ini")
modular_building = md.ModularBuilding(nodes, room_indx, edges_all, labels, joint_hor, joint_ver, cor_edges)
# 按房间分好节点
modulars_of_building = modular_building.building_modulars



'''GA to model'''
# case 1 按照区域分组优化
POP_SIZE = 50
DNA_SIZE = 2*story_num*3
CROSSOVER_RATE = 0.4
MUTATION_RATE = 0.15
N_GENERATIONS = 100
num_thread = 10
min_genera = []
x = np.linspace(0, 12, 13)
# for i in range(3):
#     for num_var in [8,9,10]:
#         for num_room_type in [1]:
#             # 创建多线程sap
#             mySapObject_name, ModelPath_name, SapModel_name =mulit_get_sap(num_thread)
#             all_GA_infor = Run_GA_allstory2_divided_mulit(POP_SIZE,DNA_SIZE,CROSSOVER_RATE,MUTATION_RATE,N_GENERATIONS,x,
#                                              num_var,num_room_type,num_thread,mySapObject_name,SapModel_name,ModelPath_name)
#             output_GA_in(all_GA_infor, num_var,num_room_type,i)
#             gc.collect()
#

#     min_genera.append(all_GA_infor[1])
# weight_fin = all_GA_infor[0]
# pop_fin = all_GA_infor[1]
# max_ru = all_GA_infor[2]
# pop_all= all_GA_infor[3]
# pop_zhongqun_all= all_GA_infor[4]
# weight_min= all_GA_infor[5]
# pop_all_fitness= all_GA_infor[6]
# pop_all_weight= all_GA_infor[7]
# col_up_all= all_GA_infor[8]
# beam_up_all= all_GA_infor[9]
# pop_room = weight_fin

#使用xlrd读取信息 从1开始不是0开始
pop_room = []
pop_room_label = []
# wb = xlrd.open_workbook(
#     filename=f'D:\desktop\os\optimization of structure\optimization of structure\optimization of structure\out_all_infor\\run_infor_8_941.xlsx',
#     formatting_info=True)
# sheet1 = wb.sheet_by_index(0)
# for z in range(18):
#     rows = sheet1.row_values(1830)[z]
#     pop_room.append(rows)
# sheet1 = wb.sheet_by_index(2)
# for z in range(96):
#     rows = sheet1.row_values(1830)[z]
#     pop_room_label.append(rows)

#使用openxyxl读取信息

wb = openpyxl.load_workbook(
    filename=f'D:\desktop\os\optimization of structure\out_all_infor_case4\\run_infor_7_3_2.xlsx',
    )
# sheet1 = wb['pop1_all']
# for z in range(48):
#     rows = sheet1.cell(4311,z+1).value
#     pop_room.append(rows)
# sheet1 = wb['pop3_all']
# for z in range(modular_length_num*2*story_num):
#     rows = sheet1.cell(4311,z+1).value
#     pop_room_label.append(rows)



# pop_room = []
# pop_room_label = []
for i in range(24):
    pop_room.append(10)
# pop_room = all_GA_infor[0]
for i in range(modular_length_num*2*story_num):
    pop_room_label.append(0)



# #
# pop_room = np.array(pop_room)
# pop_room = np.array(all_GA_infor[0])
APIPath = os.path.join(os.getcwd(), 'cases')
mySapObject, ModelPath, SapModel = ms.SAPanalysis_GA_run(APIPath)
# ret = SapModel.File.Save("D:\图片文件夹\结构分析模型.sdb")
weight1,g_col,g_beam,reaction_all,section_all,all_up_name,all_up_data,Joint_dis,all_force = mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop_room,pop_room_label)


# weight1,g_col,g_beam,reaction_all,section_all,all_up_name,all_up_data,Joint_dis,all_force=mulit_Sap_analy_allroom(ModelPath,mySapObject, SapModel,pop_room,pop_room_label)


# model_up_vis(model_data,all_force[11][1])
#
#
X_dis = Joint_dis[0]
Y_dis = Joint_dis[1]
X_dis_ave = Joint_dis[2]
Y_dis_ave = Joint_dis[3]
X_dis_radio = Joint_dis[4]
Y_dis_radio = Joint_dis[5]
X_interdis = Joint_dis[6]
Y_interdis = Joint_dis[7]
X_interdis_ave = Joint_dis[8]
Y_interdis_ave = Joint_dis[9]
X_interdis_radio = Joint_dis[10]
Y_interdis_radio = Joint_dis[11]
floor_mass = all_force[9]
floor_mass_radio = all_force[10]
for i in range(len(floor_mass_radio)):
    floor_mass_radio[i] = round(floor_mass_radio[i], 3)
max_floor_radio = max(floor_mass_radio)
max_floor_label =floor_mass_radio.index(max(floor_mass_radio))
word_infor = []
word_1th = []
word_1th.append(weight1)
word_1th.append(max_floor_radio)
word_1th.append(max_floor_label)
word_1th.append(max(Y_interdis_radio))
word_1th.append(Y_interdis_radio.index(max(Y_interdis_radio)))
word_3th = []
word_3th.append(Y_dis)
word_3th.append(Y_interdis)
word_3th.append(Y_interdis_ave)
word_3th.append(Y_interdis_radio)
word_infor.append(word_1th)
word_infor.append(word_3th)

res1, res2,gx,gx_demo=Fun_1(weight1,g_col,g_beam,Joint_dis,all_force,10000)

# yangtingting = [i for i in range(50)]
# luyiwen = []
# for i in range(len(yangtingting)):
#     luyiwen.append( m.e ** (yangtingting[i] * 0.5))
#     # luyiwen.append((yangtingting[i]))
# # luyiwen[49]=luyiwen[49]*30
# zhanjiaqi = 0
# for i in range(len(luyiwen)):
#     zhanjiaqi+=luyiwen[i]
#
# zhanhuang = []
# for i in range(len(luyiwen)):
#     zhanhuang.append(luyiwen[i]/zhanjiaqi)
#

# import copy
# import math as m
# import random
# from random import randint
# import numpy as np
# from collections import Counter
#
# POP_SIZE = 50
# pop = [[] for i in range(50)]
# fitness = []
# for i in range(50):
#     pop[i].append(i*i)
#
#     fitness.append(random.randint(1,50))
#
#
# def select_2(pop, fitness):  # nature selection wrt pop's fitness
#
#     fit_ini = copy.deepcopy(fitness)
#     luyi = copy.deepcopy(fitness)
#     luyi.sort(reverse=True)
#     sort_num = []
#     for i in range(len(fit_ini)):
#         sort_num.append(luyi.index(fit_ini[i]))
#
#
#     list_n = []
#     for i in range(len(lst)):
#         list_n.append(lst[sort_num[i]])
#     list_new = []
#     for i in range(len(fit_ini)):
#         sort_num.append(fit_ini.index(luyi[i]))
#     for i in range(len(fit_ini)):
#         list_new.append(lst[sort_num[i]])
#     for i in range(len(list_new)):
#         list_new[i] = m.e ** (list_new[i] * 1.2)
#     idx = np.random.choice(np.arange(POP_SIZE), size=POP_SIZE, replace=True,
#                            p=np.array(list_new) / (sum(list_new)))
#     pop2 = np.zeros((POP_SIZE, len(pop[0])))
#     for i in range(len(pop2)):
#         pop2[i] = pop[int(idx[i])]
#     return pop2
#
# yangtingting = select_2(pop, fitness)
# zhanhuang= yangtingting.tolist()
# zhanjiahuang = []
# for i in range(len(zhanhuang)):
#     zhanjiahuang.append(zhanhuang[i][0])
# zhanjiaqi = Counter(zhanjiahuang)
#
#
